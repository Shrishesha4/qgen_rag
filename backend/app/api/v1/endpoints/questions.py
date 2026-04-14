"""
Question generation API endpoints.
"""

import json
import os
import asyncio
from typing import Optional, List, AsyncGenerator
import uuid
from datetime import datetime, timezone
import re
import logging
from contextlib import asynccontextmanager

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from pydantic import BaseModel

from app.core.database import get_db
from app.core.database import AsyncSessionLocal
from app.core.config import settings
from app.schemas.question import (
    QuestionGenerationRequest,
    QuestionResponse,
    QuestionListResponse,
    QuestionRatingRequest,
    GenerationProgress,
    GenerationSessionResponse,
    QuickGenerateProgress,
)
from app.services.question_service import QuestionGenerationService
from app.services.provider_service import get_provider_service, create_question_service_for_provider
from app.services.document_service import DocumentService
from app.api.v1.deps import get_current_user, rate_limit, ensure_user_can_generate, ensure_user_can_vet
from app.models.user import User
from app.models.question import Question, GenerationSession
from app.models.document import Document
from app.models.subject import Subject, Topic
from app.models.generation_run import GenerationRun
from app.services.generation_status_service import GenerationStatusService, broadcast_generation_update
from app.services.activity_service import safe_record_activity


# MIME type mapping for allowed extensions
MIME_TYPE_MAPPING = {
    ".pdf": "application/pdf",
    ".txt": "text/plain",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


async def _build_provider_aware_question_service(
    db_session: AsyncSession,
    provider_key: Optional[str] = None,
) -> QuestionGenerationService:
    """Create a QuestionGenerationService backed by an admin-configured provider.

    If ``provider_key`` is given the matching enabled provider is used; otherwise
    the preferred provider (from settings.LLM_PROVIDER) or the first enabled
    provider is chosen.  Falls back to the default QuestionGenerationService when
    no providers are configured or on any error.
    """
    try:
        ps = get_provider_service()
        enabled = await ps.get_enabled_providers()
        if enabled:
            if provider_key:
                provider = next((p for p in enabled if p.key == provider_key), enabled[0])
            else:
                preferred = (getattr(settings, "LLM_PROVIDER", "") or "").strip().lower()
                provider = next((p for p in enabled if p.key == preferred), enabled[0])
            return await create_question_service_for_provider(db_session, provider)
    except Exception as provider_exc:
        logger.warning(
            "Provider-aware generation setup failed, falling back to default: %s",
            provider_exc,
        )
    return QuestionGenerationService(db_session)


router = APIRouter()
_BACKGROUND_GENERATION_TASKS: dict[str, asyncio.Task] = {}
_BACKGROUND_GENERATION_STATUS: dict[str, dict] = {}
_BACKGROUND_GENERATION_QUEUE: list[dict] = []  # Global queue for all users
_QUEUE_PROCESSING_LOCK = asyncio.Lock()  # Prevent race conditions in queue processing
_MAX_CONCURRENT_GENERATIONS = 2  # Max concurrent generations across all users
_TASK_CLEANUP_INTERVAL = 300  # Clean up completed tasks every 5 minutes
_QUEUE_PROCESSING_INTERVAL = 10  # Process queue every 10 seconds
_MAX_RETRY_ATTEMPTS = 3  # Max retry attempts for failed queue items
_GENERATION_REFILL_MAX_ROUNDS = 4  # Top-up passes if generation ends below target
_GENERATION_REFILL_ATTEMPT_MULTIPLIER = 4  # Attempts per missing question in each refill round

# Configure logging
logger = logging.getLogger(__name__)


def _bg_gen_task_key(user_id: str, subject_id: str) -> str:
    return f"{user_id}:{subject_id}"


def _get_queue_position(user_id: str, subject_id: str) -> int:
    """Get position in queue for this user/subject combination.
    Returns 0 if not found, 1-based position if found."""
    for i, item in enumerate(_BACKGROUND_GENERATION_QUEUE):
        if item["user_id"] == user_id and item["subject_id"] == subject_id:
            return i + 1
    return 0


def _add_to_queue(user_id: str, subject_id: str, count: int, request_data: dict) -> int:
    """Add to queue and return position (1-based)"""
    queue_item = {
        "user_id": user_id,
        "subject_id": subject_id,
        "count": count,
        "request_data": request_data,
        "added_at": datetime.now(timezone.utc),
        "run_id": str(uuid.uuid4()),
        "retry_count": 0,
    }
    _BACKGROUND_GENERATION_QUEUE.append(queue_item)
    logger.info(f"Added to queue: user_id={user_id}, subject_id={subject_id}, position={len(_BACKGROUND_GENERATION_QUEUE)}")
    return len(_BACKGROUND_GENERATION_QUEUE)


def _get_next_from_queue() -> dict | None:
    """Get next item from queue"""
    if not _BACKGROUND_GENERATION_QUEUE:
        return None
    return _BACKGROUND_GENERATION_QUEUE.pop(0)


def _get_current_running_count() -> int:
    """Get count of currently running generations and clean up completed tasks"""
    count = 0
    completed_tasks = []
    
    for key, task in _BACKGROUND_GENERATION_TASKS.items():
        if task.done():
            completed_tasks.append(key)
        else:
            count += 1
    
    # Clean up completed tasks to prevent memory leak
    for key in completed_tasks:
        del _BACKGROUND_GENERATION_TASKS[key]
        logger.debug(f"Cleaned up completed task: {key}")
    
    return count


def _cleanup_completed_tasks():
    """Clean up all completed tasks from the task dictionary"""
    completed_tasks = []
    for key, task in _BACKGROUND_GENERATION_TASKS.items():
        if task.done():
            completed_tasks.append(key)
    
    for key in completed_tasks:
        del _BACKGROUND_GENERATION_TASKS[key]
        logger.debug(f"Cleaned up completed task: {key}")
    
    if completed_tasks:
        logger.info(f"Cleaned up {len(completed_tasks)} completed tasks")


async def _process_queue():
    """Process queue when there's capacity with proper locking to prevent race conditions"""
    async with _QUEUE_PROCESSING_LOCK:
        # Double-check capacity with lock held
        current_running = _get_current_running_count()
        if current_running >= _MAX_CONCURRENT_GENERATIONS:
            logger.debug(f"Queue processing skipped: {current_running}/{_MAX_CONCURRENT_GENERATIONS} tasks running")
            return
        
        next_item = _get_next_from_queue()
        if not next_item:
            logger.debug("Queue processing skipped: no items in queue")
            return
        
        logger.info(f"Processing queue item: user_id={next_item['user_id']}, subject_id={next_item['subject_id']}, attempt={next_item.get('retry_count', 0) + 1}")
        
        # Start the queued generation
        try:
            await _run_background_subject_generation(
                user_id=next_item["user_id"],
                subject_id=next_item["subject_id"],
                count=next_item["count"],
                types=next_item["request_data"]["types"],
                difficulty=next_item["request_data"]["difficulty"],
                run_id=next_item["run_id"],
                topic_id=next_item["request_data"].get("topic_id"),
                topic_ids=next_item["request_data"].get("topic_ids"),
                allow_without_reference=next_item["request_data"].get("allow_without_reference", False),
            )
            logger.info(f"Successfully started queued generation for user_id={next_item['user_id']}, subject_id={next_item['subject_id']}")
        except Exception as e:
            logger.error(f"Error processing queued generation for user_id={next_item['user_id']}, subject_id={next_item['subject_id']}: {e}")
            
            # Retry logic for failed queue items
            retry_count = next_item.get("retry_count", 0)
            if retry_count < _MAX_RETRY_ATTEMPTS:
                next_item["retry_count"] = retry_count + 1
                _BACKGROUND_GENERATION_QUEUE.insert(0, next_item)  # Put back at front of queue
                logger.warning(f"Retrying queue item (attempt {retry_count + 1}/{_MAX_RETRY_ATTEMPTS}): user_id={next_item['user_id']}, subject_id={next_item['subject_id']}")
            else:
                logger.error(f"Queue item failed after {_MAX_RETRY_ATTEMPTS} attempts: user_id={next_item['user_id']}, subject_id={next_item['subject_id']}")
                # Update status to show failure
                task_key = _bg_gen_task_key(next_item["user_id"], next_item["subject_id"])
                _BACKGROUND_GENERATION_STATUS[task_key] = {
                    "in_progress": False,
                    "run_id": next_item["run_id"],
                    "status": "failed",
                    "progress": 0,
                    "current_question": 0,
                    "total_questions": next_item["count"],
                    "message": f"Failed after {_MAX_RETRY_ATTEMPTS} retry attempts: {str(e)}",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
            
            # Continue processing queue if there are more items
            asyncio.create_task(_process_queue())


async def _wait_for_subject_docs_ready(
    db: AsyncSession,
    user_id: str,
    subject_id: str,
    timeout_seconds: int = 30 * 60,
    poll_interval_seconds: float = 3.0,
) -> bool:
    """Wait until subject reference docs are no longer processing and at least one is ready."""
    ready_statuses = {"completed", "complete", "processed", "ready", "indexed"}
    failed_statuses = {"failed", "error"}
    started_at = asyncio.get_event_loop().time()

    while asyncio.get_event_loop().time() - started_at < timeout_seconds:
        res = await db.execute(
            select(Document.processing_status).where(
                Document.subject_id == subject_id,
                Document.index_type.in_(["reference_book", "template_paper"]),
            )
        )
        statuses = [(s or "").lower() for s in res.scalars().all()]

        if not statuses:
            await asyncio.sleep(poll_interval_seconds)
            continue

        processing_count = sum(1 for s in statuses if s in {"pending", "processing"})
        ready_count = sum(1 for s in statuses if s in ready_statuses)
        failed_count = sum(1 for s in statuses if s in failed_statuses)

        if ready_count > 0 and processing_count == 0:
            return True
        if failed_count == len(statuses):
            return False

        await asyncio.sleep(poll_interval_seconds)

    return False


async def _delayed_status_cleanup(task_key: str, delay_seconds: float = 60.0) -> None:
    """Remove a terminal status entry after a delay so the frontend has time to poll it."""
    await asyncio.sleep(delay_seconds)
    _BACKGROUND_GENERATION_STATUS.pop(task_key, None)


async def _get_generated_count_since_start(
    db: AsyncSession,
    subject_id: str,
    started_total_questions: int,
) -> int:
    """Return number of newly saved questions for this subject since generation started."""
    subject_total_res = await db.execute(
        select(Subject.total_questions).where(Subject.id == subject_id)
    )
    current_subject_total = int(subject_total_res.scalar_one_or_none() or started_total_questions)
    return max(0, current_subject_total - started_total_questions)


async def _refill_missing_questions_to_target(
    db: AsyncSession,
    question_service: QuestionGenerationService,
    *,
    user_id: str,
    subject_id: str,
    count: int,
    types: list[str],
    difficulty: str,
    marks_by_type: dict[str, int],
    context: str,
    topic_id: Optional[str],
    selected_topic_ids: list[str],
    allow_without_reference: bool,
    task_key: str,
    run_id: str,
    started_total_questions: int,
    target_total_questions: int,
) -> int:
    """
    Refill strategy: if generated count is below requested count, run bounded retry rounds
    with per-question attempts until target is reached or retry budget is exhausted.
    """
    generated_count = min(count, await _get_generated_count_since_start(db, subject_id, started_total_questions))
    if generated_count >= count:
        return generated_count

    candidate_topic_ids = [tid for tid in selected_topic_ids if tid]
    if topic_id and topic_id not in candidate_topic_ids:
        candidate_topic_ids = [topic_id]

    for refill_round in range(1, _GENERATION_REFILL_MAX_ROUNDS + 1):
        generated_count = min(count, await _get_generated_count_since_start(db, subject_id, started_total_questions))
        remaining = count - generated_count
        if remaining <= 0:
            break

        max_attempts_this_round = max(remaining * _GENERATION_REFILL_ATTEMPT_MULTIPLIER, remaining)
        _BACKGROUND_GENERATION_STATUS[task_key] = {
            "in_progress": True,
            "run_id": run_id,
            "status": "refilling",
            "progress": min(99, int((generated_count / max(1, count)) * 100)),
            "current_question": generated_count,
            "total_questions": count,
            "started_total_questions": started_total_questions,
            "target_total_questions": target_total_questions,
            "message": f"Auto-refill round {refill_round}: generating {remaining} missing questions",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }

        round_success = 0
        for attempt_idx in range(max_attempts_this_round):
            chosen_topic_id = topic_id
            if not chosen_topic_id and candidate_topic_ids:
                chosen_topic_id = candidate_topic_ids[attempt_idx % len(candidate_topic_ids)]

            produced = False
            try:
                refill_generator = question_service.quick_generate_from_subject(
                    user_id=user_id,
                    subject_id=subject_id,
                    context=context,
                    count=1,
                    types=types,
                    difficulty=difficulty,
                    marks_by_type=marks_by_type,
                    topic_id=chosen_topic_id,
                    existing_session_id=None,
                    allow_without_reference=allow_without_reference,
                )

                async for refill_event in refill_generator:
                    if (refill_event.current_question or 0) >= 1 or refill_event.question is not None:
                        produced = True

                await db.commit()
            except Exception as refill_exc:
                logger.warning(
                    "Refill attempt failed for subject_id=%s topic_id=%s attempt=%s: %s",
                    subject_id,
                    chosen_topic_id,
                    attempt_idx + 1,
                    refill_exc,
                )

            if produced:
                round_success += 1

            generated_count = min(count, await _get_generated_count_since_start(db, subject_id, started_total_questions))
            if generated_count >= count:
                break

        generated_count = min(count, await _get_generated_count_since_start(db, subject_id, started_total_questions))
        if generated_count >= count:
            break
        if round_success == 0:
            # Stop early if a full refill round produced no net new question.
            break

    return generated_count


async def _run_background_subject_generation(
    user_id: str,
    subject_id: str,
    count: int,
    types: list[str],
    difficulty: str,
    run_id: Optional[str] = None,
    topic_id: Optional[str] = None,
    topic_ids: Optional[list[str]] = None,
    allow_without_reference: bool = False,
) -> None:
    task_key = _bg_gen_task_key(user_id, subject_id)
    error_occurred = False
    error_message = ""
    final_current_question = 0
    started_total_questions = 0
    target_total_questions = count
    run_id_value = run_id or str(uuid.uuid4())
    db_run_created = False
    effective_topic_id = topic_id  # Track the actual topic being generated
    try:
        async with AsyncSessionLocal() as task_db:
            subject_res = await task_db.execute(
                select(Subject).where(Subject.id == subject_id)
            )
            subject = subject_res.scalar_one_or_none()
            if not subject:
                return
            started_total_questions = int(subject.total_questions or 0)
            target_total_questions = started_total_questions + count

            docs_ready = True
            if not allow_without_reference:
                _BACKGROUND_GENERATION_STATUS[task_key] = {
                    "in_progress": True,
                    "run_id": run_id_value,
                    "status": "waiting_for_documents",
                    "progress": 0,
                    "current_question": 0,
                    "total_questions": count,
                    "started_total_questions": started_total_questions,
                    "target_total_questions": target_total_questions,
                    "message": "Waiting for reference documents to finish processing",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                docs_ready = await _wait_for_subject_docs_ready(task_db, user_id, subject_id)
                if not docs_ready:
                    return

            selected_topic_ids = [tid for tid in (topic_ids or []) if tid]
            if topic_id and topic_id not in selected_topic_ids:
                selected_topic_ids = [topic_id]

            if topic_id and len(selected_topic_ids) <= 1:
                topic_res = await task_db.execute(
                    select(Topic).where(Topic.id == topic_id, Topic.subject_id == subject_id)
                )
                topic = topic_res.scalar_one_or_none()
                if not topic:
                    return
                context = f"{subject.name}: {topic.name}"
                effective_topic_id = topic_id
            else:
                topics_res = await task_db.execute(
                    select(Topic.name).where(Topic.subject_id == subject_id).order_by(Topic.order_index.asc())
                )
                topic_names = [name for name in topics_res.scalars().all() if name]
                context = f"{subject.name}: {', '.join(topic_names)}" if topic_names else subject.name

            # Create database record for cross-device visibility (only if we have a specific topic)
            if effective_topic_id:
                try:
                    gen_status_service = GenerationStatusService(task_db)
                    await gen_status_service.create_run(
                        run_id=run_id_value,
                        subject_id=subject_id,
                        topic_id=effective_topic_id,
                        user_id=user_id,
                        target_count=count,
                        status="generating",
                        message="Starting question generation",
                    )
                    db_run_created = True
                except Exception as db_err:
                    logger.warning(f"Failed to create generation run in DB: {db_err}")

            marks_by_type = {"mcq": 1, "short_answer": 2, "long_answer": 5}
            question_service = await _build_provider_aware_question_service(task_db)

            if len(selected_topic_ids) > 1:
                topic_rows_res = await task_db.execute(
                    select(Topic).where(Topic.subject_id == subject_id, Topic.id.in_(selected_topic_ids))
                )
                topic_rows = topic_rows_res.scalars().all()
                topic_name_by_id = {topic_row.id: topic_row.name for topic_row in topic_rows}
                valid_topic_ids = [tid for tid in selected_topic_ids if tid in topic_name_by_id]
                if not valid_topic_ids:
                    return

                parallel_workers = max(
                    1,
                    min(
                        count,
                        int(settings.QUICK_GENERATE_PARALLEL_WORKERS or 1),
                    ),
                )
                max_attempts_per_slot = 4
                total_multi_attempts = 0
                max_multi_attempts = count * max_attempts_per_slot

                # Build initial slot queue with even distribution across topics
                multi_slot_queue = []
                questions_per_topic = count // len(valid_topic_ids)
                remaining_questions = count % len(valid_topic_ids)
                
                # Create an even distribution list
                topic_distribution = []
                for i, topic_id in enumerate(valid_topic_ids):
                    # Add base questions per topic
                    base_questions = questions_per_topic + (1 if i < remaining_questions else 0)
                    for _ in range(base_questions):
                        topic_distribution.append(topic_id)
                
                # Shuffle to avoid predictable patterns
                import random
                random.shuffle(topic_distribution)

                # Build weighted provider assignment list (e.g. deepseek×15 + xai×15 → shuffled)
                _slot_provider_assignments: list[str] = []
                try:
                    _ps = get_provider_service()
                    _allocations = await _ps.allocate_batch(count)
                    for _alloc in _allocations:
                        _slot_provider_assignments.extend(
                            [_alloc.provider.key] * _alloc.question_count
                        )
                    random.shuffle(_slot_provider_assignments)
                except Exception as _palloc_exc:
                    logger.warning(
                        "Provider allocation failed, slots will use default provider: %s",
                        _palloc_exc,
                    )
                
                # Create slots from the shuffled distribution
                for slot_idx, chosen_tid in enumerate(topic_distribution):
                    assigned_provider = (
                        _slot_provider_assignments[slot_idx]
                        if slot_idx < len(_slot_provider_assignments)
                        else None
                    )
                    multi_slot_queue.append({
                        "slot_id": slot_idx,
                        "topic_id": chosen_tid,
                        "attempts": 0,
                        "provider_key": assigned_provider,
                    })

                async def _run_single_topic_slot(slot: dict) -> dict:
                    chosen_topic_id = slot["topic_id"]
                    chosen_topic_name = topic_name_by_id.get(chosen_topic_id, "Selected Topic")
                    slot_context = f"{subject.name}: {chosen_topic_name}"
                    produced_count = 0
                    last_message = ""

                    try:
                        async with AsyncSessionLocal() as worker_db:
                            worker_question_service = await _build_provider_aware_question_service(
                                worker_db, slot.get("provider_key")
                            )
                            generator = worker_question_service.quick_generate_from_subject(
                                user_id=user_id,
                                subject_id=subject_id,
                                context=slot_context,
                                count=1,
                                types=types,
                                difficulty=difficulty,
                                marks_by_type=marks_by_type,
                                topic_id=chosen_topic_id,
                                existing_session_id=None,
                                allow_without_reference=allow_without_reference,
                                skip_generation_lock=True,
                            )

                            async for progress_event in generator:
                                if progress_event.message:
                                    last_message = progress_event.message
                                if (progress_event.current_question or 0) >= 1 or progress_event.question is not None:
                                    produced_count = 1

                            await worker_db.commit()

                        return {
                            "ok": produced_count >= 1,
                            "topic_id": chosen_topic_id,
                            "message": last_message,
                        }
                    except Exception as slot_exc:
                        return {
                            "ok": False,
                            "topic_id": chosen_topic_id,
                            "message": str(slot_exc),
                        }

                completed_slots = 0
                while (
                    multi_slot_queue
                    and completed_slots < count
                    and total_multi_attempts < max_multi_attempts
                ):
                    batch_size = min(
                        parallel_workers,
                        len(multi_slot_queue),
                        max(1, count - completed_slots),
                    )
                    batch_slots = [multi_slot_queue.pop(0) for _ in range(batch_size)]
                    total_multi_attempts += len(batch_slots)

                    _BACKGROUND_GENERATION_STATUS[task_key] = {
                        "in_progress": True,
                        "run_id": run_id_value,
                        "status": "generating",
                        "progress": min(99, int((completed_slots / max(1, count)) * 100)),
                        "current_question": completed_slots,
                        "total_questions": count,
                        "started_total_questions": started_total_questions,
                        "target_total_questions": target_total_questions,
                        "message": (
                            f"Generating questions {completed_slots + 1}-{min(count, completed_slots + batch_size)} "
                            f"of {count} using {batch_size} workers"
                        ),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }

                    batch_results = await asyncio.gather(*[_run_single_topic_slot(slot) for slot in batch_slots])

                    succeeded_in_batch = 0
                    for slot, result in zip(batch_slots, batch_results):
                        if result.get("ok"):
                            succeeded_in_batch += 1
                            continue

                        slot["attempts"] += 1
                        if (
                            slot["attempts"] < max_attempts_per_slot
                            and total_multi_attempts < max_multi_attempts
                        ):
                            retry_topic = valid_topic_ids[(total_multi_attempts + slot["slot_id"]) % len(valid_topic_ids)]
                            slot["topic_id"] = retry_topic
                            multi_slot_queue.append(slot)

                    completed_slots += succeeded_in_batch

                    # Keep status in sync with persisted DB totals to avoid pill lag.
                    subject_total_res = await task_db.execute(
                        select(Subject.total_questions).where(
                            Subject.id == subject_id,
                        )
                    )
                    current_subject_total = int(subject_total_res.scalar_one_or_none() or started_total_questions)
                    db_generated = max(0, current_subject_total - started_total_questions)
                    current_generated = min(count, max(completed_slots, db_generated))

                    _BACKGROUND_GENERATION_STATUS[task_key] = {
                        "in_progress": True,
                        "run_id": run_id_value,
                        "status": "generating",
                        "progress": min(99, int((current_generated / max(1, count)) * 100)),
                        "current_question": current_generated,
                        "total_questions": count,
                        "started_total_questions": started_total_questions,
                        "target_total_questions": target_total_questions,
                        "message": f"Generated {current_generated}/{count} questions",
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }
                    
                    # Broadcast progress via WebSocket for cross-device visibility
                    if db_run_created and effective_topic_id:
                        try:
                            await broadcast_generation_update(
                                subject_id=subject_id,
                                topic_id=effective_topic_id,
                                status_data={
                                    "run_id": run_id_value,
                                    "topic_id": effective_topic_id,
                                    "in_progress": True,
                                    "status": "generating",
                                    "progress": min(99, int((current_generated / max(1, count)) * 100)),
                                    "current_question": current_generated,
                                    "total_questions": count,
                                    "message": f"Generated {current_generated}/{count} questions",
                                },
                            )
                        except Exception:
                            pass  # Don't fail generation due to WebSocket issues

                subject_total_res = await task_db.execute(
                    select(Subject.total_questions).where(
                        Subject.id == subject_id,
                    )
                )
                current_subject_total = int(subject_total_res.scalar_one_or_none() or started_total_questions)
                db_generated = max(0, current_subject_total - started_total_questions)
                final_current_question = min(count, max(completed_slots, db_generated))

            else:
                generator = question_service.quick_generate_from_subject(
                    user_id=user_id,
                    subject_id=subject_id,
                    context=context,
                    count=count,
                    types=types,
                    difficulty=difficulty,
                    marks_by_type=marks_by_type,
                    topic_id=topic_id,
                    existing_session_id=None,
                    allow_without_reference=allow_without_reference,
                )
                async for progress_event in generator:
                    # Use current_question (saved/validated questions) for progress,
                    # not progress_event.progress which includes duplicate-filtered attempts
                    saved_count = progress_event.current_question or 0
                    total = progress_event.total_questions or count
                    saved_progress = min(99, int((saved_count / max(1, total)) * 100)) if saved_count > 0 else (progress_event.progress or 0)
                    # For non-generating statuses (processing, complete, error), use the event's own progress
                    if progress_event.status not in ("generating",):
                        saved_progress = progress_event.progress or 0
                    _BACKGROUND_GENERATION_STATUS[task_key] = {
                        "in_progress": True,
                        "run_id": run_id_value,
                        "status": progress_event.status,
                        "progress": saved_progress,
                        "current_question": saved_count,
                        "total_questions": total,
                        "started_total_questions": started_total_questions,
                        "target_total_questions": target_total_questions,
                        "message": f"Generated {saved_count}/{total} questions" if saved_count > 0 and progress_event.status == "generating" else (progress_event.message or "Generating questions"),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }

                    # Broadcast progress via WebSocket for cross-device visibility
                    if db_run_created and effective_topic_id:
                        try:
                            await broadcast_generation_update(
                                subject_id=subject_id,
                                topic_id=effective_topic_id,
                                status_data={
                                    "run_id": run_id_value,
                                    "topic_id": effective_topic_id,
                                    "in_progress": True,
                                    "status": progress_event.status,
                                    "progress": saved_progress,
                                    "current_question": saved_count,
                                    "total_questions": total,
                                    "message": f"Generated {saved_count}/{total} questions" if saved_count > 0 and progress_event.status == "generating" else (progress_event.message or "Generating questions"),
                                },
                            )
                        except Exception:
                            pass  # Don't fail generation due to WebSocket issues
                final_current_question = _BACKGROUND_GENERATION_STATUS.get(task_key, {}).get("current_question", count)

            final_current_question = await _refill_missing_questions_to_target(
                task_db,
                question_service,
                user_id=user_id,
                subject_id=subject_id,
                count=count,
                types=types,
                difficulty=difficulty,
                marks_by_type=marks_by_type,
                context=context,
                topic_id=topic_id,
                selected_topic_ids=selected_topic_ids,
                allow_without_reference=allow_without_reference,
                task_key=task_key,
                run_id=run_id_value,
                started_total_questions=started_total_questions,
                target_total_questions=target_total_questions,
            )

            _BACKGROUND_GENERATION_STATUS[task_key] = {
                "in_progress": True,
                "run_id": run_id_value,
                "status": "generating",
                "progress": min(99, int((final_current_question / max(1, count)) * 100)),
                "current_question": final_current_question,
                "total_questions": count,
                "started_total_questions": started_total_questions,
                "target_total_questions": target_total_questions,
                "message": f"Generated {final_current_question}/{count} questions",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await task_db.commit()
    except Exception as exc:
        error_occurred = True
        error_message = str(exc)
    finally:
        # Remove the task reference immediately
        _BACKGROUND_GENERATION_TASKS.pop(task_key, None)
        # Set a terminal status so the frontend can see completion/error
        if error_occurred:
            _BACKGROUND_GENERATION_STATUS[task_key] = {
                "run_id": run_id_value,
                "status": "error",
                "progress": 0,
                "current_question": final_current_question,
                "total_questions": count,
                "started_total_questions": started_total_questions,
                "target_total_questions": target_total_questions,
                "message": f"Generation failed: {error_message}",
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "in_progress": False,
            }
            logger.error(f"Background generation failed for task_key={task_key}: {error_message}")
            
            # Update database and broadcast failure via WebSocket
            if db_run_created and effective_topic_id:
                try:
                    async with AsyncSessionLocal() as cleanup_db:
                        gen_status_service = GenerationStatusService(cleanup_db)
                        await gen_status_service.fail_run(run_id_value, error_message)
                except Exception as db_err:
                    logger.warning(f"Failed to update generation run failure in DB: {db_err}")
        else:
            final_current_question = max(0, min(count, int(final_current_question or 0)))
            terminal_progress = 100 if final_current_question >= count else int((final_current_question / max(1, count)) * 100)
            _BACKGROUND_GENERATION_STATUS[task_key] = {
                "run_id": run_id_value,
                "status": "completed",
                "progress": terminal_progress,
                "current_question": final_current_question,
                "total_questions": count,
                "started_total_questions": started_total_questions,
                "target_total_questions": target_total_questions,
                "message": (
                    "Generation complete"
                    if final_current_question >= count
                    else f"Generation complete ({final_current_question}/{count})"
                ),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "in_progress": False,
            }
            logger.info(f"Background generation completed for task_key={task_key}: {final_current_question}/{count} questions")
            
            # Update database and broadcast completion via WebSocket
            if db_run_created and effective_topic_id:
                try:
                    async with AsyncSessionLocal() as cleanup_db:
                        gen_status_service = GenerationStatusService(cleanup_db)
                        await gen_status_service.complete_run(
                            run_id_value,
                            final_current_question,
                            status="completed",
                            message=(
                                "Generation complete"
                                if final_current_question >= count
                                else f"Generation complete ({final_current_question}/{count})"
                            ),
                        )
                except Exception as db_err:
                    logger.warning(f"Failed to update generation run completion in DB: {db_err}")
        
        # Process queue after task completes (both success and error cases)
        asyncio.create_task(_process_queue())
        
        # Schedule cleanup after 60s so the frontend has time to poll the terminal status
        asyncio.create_task(_delayed_status_cleanup(task_key, delay_seconds=60.0))


async def _periodic_cleanup():
    """Periodic cleanup of completed tasks and queue processing"""
    while True:
        try:
            # Clean up completed tasks
            _cleanup_completed_tasks()
            
            # Process queue if there's capacity
            await _process_queue()
            
            logger.debug("Periodic cleanup completed")
        except Exception as e:
            logger.error(f"Error in periodic cleanup: {e}")
        
        await asyncio.sleep(_TASK_CLEANUP_INTERVAL)


# Start periodic cleanup when the module is loaded
_cleanup_task = None

def start_periodic_cleanup():
    """Start the periodic cleanup task if not already running"""
    global _cleanup_task
    if _cleanup_task is None or _cleanup_task.done():
        _cleanup_task = asyncio.create_task(_periodic_cleanup())
        logger.info("Started periodic cleanup task")


async def sse_with_heartbeat(
    generator: AsyncGenerator[str, None],
    interval: float = 15.0,
) -> AsyncGenerator[str, None]:
    """
    Wraps an SSE generator and injects an SSE comment (': heartbeat\\n\\n') every
    `interval` seconds while waiting for the next real event.

    This keeps the TCP connection alive through Cloudflare Tunnels (and other
    proxies) that would otherwise buffer or drop the stream due to inactivity.
    SSE comments are ignored by EventSource clients but flush Cloudflare's buffer.
    """
    gen = generator.__aiter__()
    next_item: asyncio.Future = asyncio.ensure_future(gen.__anext__())

    while True:
        done, _ = await asyncio.wait({next_item}, timeout=interval)
        if next_item in done:
            try:
                yield next_item.result()
                next_item = asyncio.ensure_future(gen.__anext__())
            except StopAsyncIteration:
                break
        else:
            # Timeout — send a keepalive comment to flush Cloudflare's buffer
            yield ": heartbeat\n\n"


# ── Question Starter Diversity Pool ──────────────────────────────────────────
# Organised by question type so the LLM gets type-appropriate starters.
# Each batch cycles through these to ensure no two questions share the same opening.
# IMPORTANT: Avoid scenario setups (A student, A clinician, etc.) as they lead to repetitive patterns.
_QUESTION_STARTERS: dict[str, list[str]] = {
    "mcq": [
        "Which", "What", "How", "Why", "When", "Where", "Who",
        "Select", "Identify", "Choose", "Determine", "Classify",
        "Given that", "In the context of", "If", "During",
        "From the following", "Among the following", "Consider",
    ],
    "short_answer": [
        "Explain", "Describe", "Define", "Outline", "Summarize",
        "Distinguish", "Compare", "Contrast", "Illustrate", "Justify",
        "State", "List", "Mention", "Write", "Give",
        "What is meant by", "What are the", "How does", "Why is",
        "In what way", "What happens when", "What are the effects of",
        "Briefly explain", "With an example", "Using a diagram",
    ],
    "long_answer": [
        "Analyze", "Evaluate", "Discuss", "Examine", "Elaborate",
        "Critically analyze", "Compare and contrast", "With the help of",
        "Derive", "Prove", "Show that", "Sketch and explain",
        "Design", "Propose", "Formulate", "Develop",
        "Assess", "Critique", "Investigate", "Synthesize",
        "With suitable examples", "With a neat diagram",
    ],
    # default pool used when type is unknown
    "default": [
        "What", "How", "Why", "Which", "Describe", "Explain",
        "Analyze", "Define", "Discuss", "Evaluate", "Identify",
        "Compare", "Illustrate", "Justify", "State",
    ],
}


def _pick_starter(q_type: str, used_starters: list[str]) -> str:
    """
    Pick the next unused starter word for a question type, cycling through
    the pool to guarantee variety across a generation batch.
    """
    import random
    pool = _QUESTION_STARTERS.get(q_type, _QUESTION_STARTERS["default"])
    # Prefer starters not yet used in this batch
    available = [s for s in pool if s not in used_starters]
    if not available:
        # All used — reset and pick randomly
        used_starters.clear()
        available = pool
    chosen = random.choice(available)
    used_starters.append(chosen)
    return chosen


# ─────────────────────────────────────────────────────────────────────────────

import logging as _logging

_dedupe_logger = _logging.getLogger(__name__)
_bg_status_logger = _logging.getLogger(__name__ + ".bg_status")

DEDUPE_SIMILARITY_THRESHOLD = 0.895


async def _preload_subject_embeddings(
    db: AsyncSession,
    subject_id: str,
    limit: int = 2000,
) -> list:
    """
    Preload existing question embeddings for a subject from the DB.
    Used to detect duplicates against the full question bank, not just the
    current generation session.
    """
    try:
        emb_res = await db.execute(
            select(Question.question_embedding).where(
                Question.subject_id == subject_id,
                Question.is_archived == False,
                Question.question_embedding.isnot(None),
            ).limit(limit)
        )
        embeddings = [r[0] for r in emb_res.all()]
        _dedupe_logger.info(f"Preloaded {len(embeddings)} existing embeddings for subject {subject_id}")
        return embeddings
    except Exception as e:
        _dedupe_logger.warning(f"Could not preload existing embeddings for subject {subject_id}: {e}")
        return []


def _is_duplicate_embedding(
    candidate_embedding: list,
    existing_embeddings: list,
    generated_embeddings: list,
    threshold: float = DEDUPE_SIMILARITY_THRESHOLD,
) -> bool:
    """
    Check whether a candidate embedding is too similar to any existing question
    (from the DB) or any question generated in this session.

    Returns True if the candidate should be rejected as a duplicate.
    """
    import numpy as np

    all_embeddings = existing_embeddings + generated_embeddings
    if not all_embeddings:
        return False

    query = np.array(candidate_embedding)
    matrix = np.array(all_embeddings)

    query_norm = query / (np.linalg.norm(query) or 1.0)
    norms = np.linalg.norm(matrix, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    matrix_norms = matrix / norms

    similarities = np.dot(matrix_norms, query_norm)
    max_sim = float(np.max(similarities))
    if max_sim > threshold:
        _dedupe_logger.debug(f"Duplicate detected: max_similarity={max_sim:.4f} > {threshold}")
        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────

def _sanitize_question_fields(kwargs: dict) -> dict:
    """
    Sanitize Question field values to fit within DB column limits.
    Extracts short IDs from verbose LLM responses (e.g. 'LO2 - description...' → 'LO2').
    """
    import re as _re

    # Extract short LO ID from verbose string
    lo = kwargs.get("learning_outcome_id")
    if lo and isinstance(lo, str) and len(lo) > 50:
        # Try to extract 'LO1', 'LO2', etc.
        m = _re.match(r"(LO\d+)", lo)
        kwargs["learning_outcome_id"] = m.group(1) if m else lo[:50]
    elif lo and isinstance(lo, str):
        kwargs["learning_outcome_id"] = lo[:50]

    # Sanitize CO mapping keys (truncate verbose keys) and cap at 2 COs
    co_map = kwargs.get("course_outcome_mapping")
    if co_map and isinstance(co_map, dict):
        clean_map = {}
        for k, v in co_map.items():
            if isinstance(k, str) and len(k) > 20:
                m = _re.match(r"(CO\d+)", k)
                k = m.group(1) if m else k[:20]
            clean_map[k] = v
        # Hard-cap: keep at most the 2 COs with the highest Bloom level value
        if len(clean_map) > 2:
            clean_map = dict(sorted(clean_map.items(), key=lambda x: x[1], reverse=True)[:2])
        kwargs["course_outcome_mapping"] = clean_map

    # Normalize numeric bloom taxonomy levels to named levels
    # LLMs sometimes return the Bloom number (1/2/3) used in CO mapping instead of the word
    _bloom_num_map = {"1": "remember", "2": "apply", "3": "evaluate"}
    bloom_val = kwargs.get("bloom_taxonomy_level")
    if bloom_val is not None:
        bloom_key = str(bloom_val).strip()
        if bloom_key in _bloom_num_map:
            kwargs["bloom_taxonomy_level"] = _bloom_num_map[bloom_key]

    # Truncate/coerce other VARCHAR fields to their column limits
    _limits = {
        "question_type": 50,
        "difficulty_level": 20,
        "bloom_taxonomy_level": 30,
        "vetting_status": 20,
        "similarity_source": 50,
        "generation_status": 20,
        "user_difficulty_rating": 20,
    }
    for field, limit in _limits.items():
        val = kwargs.get(field)
        if val is not None:
            if not isinstance(val, str):
                val = str(val)
                kwargs[field] = val
            if len(val) > limit:
                kwargs[field] = val[:limit]

    return kwargs


def _normalize_llm_response(response: dict) -> dict:
    """
    Normalize LLM response keys to match expected schema.
    Handles common variations like 'question' -> 'question_text', 'answer' -> 'correct_answer'.
    """
    if not response:
        return response
    
    key_mappings = {
        "question": "question_text",
        "answer": "correct_answer",
    }
    
    for old_key, new_key in key_mappings.items():
        if old_key in response and new_key not in response:
            response[new_key] = response.pop(old_key)
    
    return response


@router.post("/generate")
async def generate_questions(
    request: QuestionGenerationRequest,
    current_user: User = Depends(rate_limit(requests=100, window_seconds=86400)),  # 100/day
    db: AsyncSession = Depends(get_db),
):
    """
    Generate questions from a document using RAG.
    Returns Server-Sent Events (SSE) stream with progress updates.
    
    The system automatically:
    1. Retrieves relevant document chunks
    2. Blacklists previously generated questions
    3. Generates unique questions avoiding duplicates
    4. Validates quality and stores results
    """
    ensure_user_can_generate(current_user)
    question_service = QuestionGenerationService(db)
    
    async def event_generator():
        """Generate SSE events."""
        async for progress in question_service.generate_questions(
            user_id=current_user.id,
            request=request,
        ):
            yield f"data: {progress.model_dump_json()}\n\n"
    
    return StreamingResponse(
        sse_with_heartbeat(event_generator()),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/quick-generate")
async def quick_generate_questions(
    file: UploadFile = File(..., description="PDF file to generate questions from"),
    context: str = Form(..., min_length=3, max_length=500, description="Context or title (e.g., 'Chapter 5: Data Structures')"),
    count: int = Form(default=5, ge=1, le=200, description="Number of questions to generate"),
    types: Optional[str] = Form(default="mcq,short_answer", description="Comma-separated question types: mcq, short_answer, long_answer"),
    difficulty: str = Form(default="medium", description="Difficulty level: easy, medium, hard"),
    bloom_levels: Optional[str] = Form(default=None, description="Comma-separated Bloom levels: remember, understand, apply, analyze, evaluate, create"),
    marks_mcq: Optional[int] = Form(default=1, ge=1, le=100, description="Marks for MCQ questions"),
    marks_short: Optional[int] = Form(default=2, ge=1, le=100, description="Marks for short answer questions"),
    marks_long: Optional[int] = Form(default=5, ge=1, le=100, description="Marks for long answer/essay questions"),
    subject_id: Optional[str] = Form(default=None, description="Subject ID to link questions to"),
    topic_id: Optional[str] = Form(default=None, description="Topic/Chapter ID to link questions to"),
    existing_session_id: Optional[str] = Form(default=None, description="Existing session ID to attach retry questions to"),
    current_user: User = Depends(rate_limit(requests=50, window_seconds=86400)),  # 50/day for quick generate
    db: AsyncSession = Depends(get_db),
):
    """
    Quick question generation from a PDF upload with just context/title.
    
    This simplified endpoint allows you to:
    1. Upload any PDF document
    2. Provide a brief context/title describing the content
    3. Get questions generated automatically using RAG + LLM
    
    No rubrics, subjects, or detailed configuration required.
    Returns Server-Sent Events (SSE) stream with progress updates.
    
    Example:
    - Upload: "chapter5_datastructures.pdf"
    - Context: "Binary Trees and Tree Traversal Algorithms"
    - Get: 5 questions about binary trees
    """
    ensure_user_can_generate(current_user)

    # Validate file extension
    filename = file.filename or "document.pdf"
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Unsupported file type. Allowed: {', '.join(settings.ALLOWED_EXTENSIONS)}",
        )
    
    # Read file content
    content = await file.read()
    
    # Validate file size
    if len(content) > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Maximum size: {settings.MAX_UPLOAD_SIZE_MB}MB",
        )
    
    # Parse types and bloom_levels from comma-separated strings
    type_list = [t.strip() for t in types.split(",")] if types else ["mcq", "short_answer"]
    bloom_list = [b.strip() for b in bloom_levels.split(",")] if bloom_levels else None
    
    # Validate types
    valid_types = {"mcq", "short_answer", "long_answer"}
    for t in type_list:
        if t not in valid_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid question type: {t}. Valid types: {', '.join(valid_types)}",
            )
    
    # Validate difficulty
    if difficulty not in {"easy", "medium", "hard"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid difficulty. Must be: easy, medium, or hard",
        )
    
    # Validate bloom levels if provided
    if bloom_list:
        valid_blooms = {"remember", "understand", "apply", "analyze", "evaluate", "create"}
        for b in bloom_list:
            if b not in valid_blooms:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid bloom level: {b}. Valid levels: {', '.join(valid_blooms)}",
                )
    
    mime_type = MIME_TYPE_MAPPING.get(ext, "application/octet-stream")

    parsed_existing_session_id: Optional[str] = None
    if existing_session_id:
        try:
            parsed_existing_session_id = str(existing_session_id)
        except ValueError:
            pass

    async def event_generator():
        """Generate SSE events for quick generation."""
        import logging
        logger = logging.getLogger(__name__)
        
        document_service = DocumentService(db)
        question_service = QuestionGenerationService(db)
        
        # Step 1: Upload and process document (run enhance in parallel)
        logger.info(f"Quick generate: Starting for user {current_user.id}, context: {context}")
        yield f"data: {QuickGenerateProgress(status='uploading', progress=5, message='Uploading document...').model_dump_json()}\n\n"
        
        try:
            import asyncio
            from app.services.llm_service import LLMService
            llm = LLMService()

            # Run context enhancement and document upload concurrently
            enhanced_context_task = asyncio.create_task(_enhance_focus_prompt(context, llm))

            yield f"data: {QuickGenerateProgress(status='processing', progress=8, message='Processing document content...').model_dump_json()}\n\n"

            # Upload and process synchronously
            logger.info(f"Quick generate: Processing document {filename}")
            document = await document_service.upload_and_process_document(
                user_id=current_user.id,
                filename=filename,
                file_content=content,
                mime_type=mime_type,
                context=context,
                subject_id=parsed_subject_id,
            )

            # By now enhance should be done (upload takes much longer)
            enhanced_context = await enhanced_context_task
            logger.info(f"Quick generate: original context='{context}' enhanced='{enhanced_context}'")

            # Extract document_id immediately to avoid session issues
            doc_id = document.id
            doc_chunks = document.total_chunks
            
            logger.info(f"Quick generate: Document processed, {doc_chunks} chunks")
            yield f"data: {QuickGenerateProgress(status='processing', progress=20, message=f'Document processed: {doc_chunks} sections extracted', document_id=doc_id).model_dump_json()}\n\n"
            
            # Step 2: Generate questions
            logger.info(f"Quick generate: Starting question generation, count={count}, types={type_list}")
            generation_started = False
            
            # Build marks by type dictionary
            marks_by_type = {
                "mcq": marks_mcq,
                "short_answer": marks_short,
                "long_answer": marks_long,
            }
            
            # Parse subject_id and topic_id
            parsed_subject_id = str(subject_id) if subject_id else None
            parsed_topic_id = str(topic_id) if topic_id else None
            
            try:
                generator = question_service.quick_generate(
                    user_id=current_user.id,
                    document_id=doc_id,
                    context=enhanced_context,
                    count=count,
                    types=type_list,
                    difficulty=difficulty,
                    bloom_levels=bloom_list,
                    marks_by_type=marks_by_type,
                    subject_id=parsed_subject_id,
                    topic_id=parsed_topic_id,
                    existing_session_id=parsed_existing_session_id,
                )
                
                async for progress in generator:
                    generation_started = True
                    # Adjust progress to account for document processing phase
                    if progress.status == "generating":
                        adjusted_progress = 20 + int(progress.progress * 0.8)
                        progress.progress = min(adjusted_progress, 100)
                    logger.info(f"Quick generate: Progress {progress.status} - {progress.message}")
                    yield f"data: {progress.model_dump_json()}\n\n"
                
                logger.info("Quick generate: Complete")
            except GeneratorExit:
                # Client disconnected during generation
                logger.warning(f"Quick generate: Client disconnected, cleaning up locks")
                if generation_started:
                    # Force cleanup the lock if generation was in progress
                    try:
                        redis_svc = question_service.redis_service
                        await redis_svc.release_generation_lock(str(current_user.id), str(doc_id))
                        logger.info(f"Quick generate: Force released lock for document {doc_id}")
                    except Exception as e:
                        logger.error(f"Quick generate: Failed to release lock on disconnect: {e}")
                raise
                
        except ValueError as e:
            logger.error(f"Quick generate: ValueError - {e}")
            yield f"data: {QuickGenerateProgress(status='error', progress=0, message=str(e)).model_dump_json()}\n\n"
        except GeneratorExit:
            # Re-raise to let FastAPI handle it
            raise
        except Exception as e:
            logger.exception(f"Quick generate: Exception - {e}")
            yield f"data: {QuickGenerateProgress(status='error', progress=0, message=f'An error occurred: {str(e)}').model_dump_json()}\n\n"
    
    return StreamingResponse(
        sse_with_heartbeat(event_generator()),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


async def _enhance_focus_prompt(context: str, llm_service) -> str:
    """Use LLM to improve the user's focus prompt before generation."""
    try:
        enhanced = await llm_service.generate(
            prompt=(
                f'Improve this academic topic description for question generation:\n'
                f'"{context}"\n\n'
                f'Rules: make it specific and concise (1-2 sentences), preserve the original intent, '
                f'use clear academic language. Return ONLY the improved text, no explanations.'
            ),
            temperature=0.3,
            max_tokens=150,
        )
        result = enhanced.strip().strip('"').strip("'")
        return result if result else context
    except Exception:
        return context


@router.post("/quick-generate-from-subject")
async def quick_generate_from_subject(
    subject_id: str = Form(..., description="Subject ID to generate questions from"),
    context: str = Form(..., min_length=3, max_length=500, description="Focus topic or description"),
    count: int = Form(default=5, ge=1, le=200, description="Number of questions to generate"),
    types: Optional[str] = Form(default="mcq,short_answer", description="Comma-separated question types"),
    difficulty: str = Form(default="medium", description="Difficulty level: easy, medium, hard"),
    marks_mcq: Optional[int] = Form(default=1, ge=1, le=100),
    marks_short: Optional[int] = Form(default=2, ge=1, le=100),
    marks_long: Optional[int] = Form(default=5, ge=1, le=100),
    topic_id: Optional[str] = Form(default=None, description="Topic/Chapter ID (optional)"),
    existing_session_id: Optional[str] = Form(default=None, description="Existing session ID to attach retry questions to"),
    allow_without_reference: bool = Form(default=False, description="Allow generating from topic/syllabus context without PDFs"),
    current_user: User = Depends(rate_limit(requests=50, window_seconds=86400)),
    db: AsyncSession = Depends(get_db),
):
    """
    Quick question generation from a subject's indexed documents (no file upload needed).

    The LLM will first enhance the focus prompt before generating questions.
    Returns Server-Sent Events (SSE) stream with progress updates.
    """
    ensure_user_can_generate(current_user)

    import logging
    logger = logging.getLogger(__name__)

    # Validate inputs
    type_list = [t.strip() for t in types.split(",")] if types else ["mcq", "short_answer"]
    valid_types = {"mcq", "short_answer", "long_answer"}
    for t in type_list:
        if t not in valid_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid question type: {t}. Valid types: {', '.join(valid_types)}",
            )

    if difficulty not in {"easy", "medium", "hard"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid difficulty. Must be: easy, medium, or hard",
        )

    try:
        parsed_subject_id = str(subject_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subject_id format")

    parsed_topic_id: Optional[str] = None
    if topic_id:
        try:
            parsed_topic_id = str(topic_id)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid topic_id format")

    parsed_existing_session_id_subj: Optional[str] = None
    if existing_session_id:
        try:
            parsed_existing_session_id_subj = str(existing_session_id)
        except ValueError:
            pass

    marks_by_type = {"mcq": marks_mcq, "short_answer": marks_short, "long_answer": marks_long}

    async def event_generator():
        question_service = await _build_provider_aware_question_service(db)

        yield f"data: {QuickGenerateProgress(status='processing', progress=5, message='Searching subject content...').model_dump_json()}\n\n"

        try:
            generation_started = False
            generator = question_service.quick_generate_from_subject(
                user_id=current_user.id,
                subject_id=parsed_subject_id,
                context=context,
                count=count,
                types=type_list,
                difficulty=difficulty,
                marks_by_type=marks_by_type,
                topic_id=parsed_topic_id,
                existing_session_id=parsed_existing_session_id_subj,
                allow_without_reference=allow_without_reference,
            )

            async for progress in generator:
                generation_started = True
                if progress.status == "generating":
                    adjusted = 10 + int(progress.progress * 0.9)
                    progress.progress = min(adjusted, 100)
                logger.info(f"Quick generate from subject: {progress.status} - {progress.message}")
                yield f"data: {progress.model_dump_json()}\n\n"

            logger.info("Quick generate from subject: Complete")

        except ValueError as e:
            logger.error(f"Quick generate from subject: ValueError - {e}")
            yield f"data: {QuickGenerateProgress(status='error', progress=0, message=str(e)).model_dump_json()}\n\n"
        except GeneratorExit:
            raise
        except Exception as e:
            logger.exception(f"Quick generate from subject: Exception - {e}")
            yield f"data: {QuickGenerateProgress(status='error', progress=0, message=f'An error occurred: {str(e)}').model_dump_json()}\n\n"

    return StreamingResponse(
        sse_with_heartbeat(event_generator()),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/cancel-generation")
async def cancel_generation(
    subject_id: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Release the generation lock for a subject, allowing a new generation to start."""
    ensure_user_can_generate(current_user)

    from app.services.redis_service import RedisService
    redis = RedisService()
    await redis.release_generation_lock(str(current_user.id), subject_id)
    return {"message": "Generation cancelled", "subject_id": subject_id}


@router.get("/estimate-capacity/{subject_id}")
async def estimate_question_capacity(
    subject_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Estimate the maximum number of questions that can be generated from a subject's content.
    This helps determine the optimal question count for training.
    """
    try:
        parsed_subject_id = str(subject_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subject_id format")

    # Get subject
    subject_res = await db.execute(
        select(Subject)
        .where(Subject.id == parsed_subject_id)
    )
    subject = subject_res.scalar_one_or_none()
    if not subject:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Subject not found")

    # Match the same document discovery strategy used by subject generation.
    # Strategy 1: directly linked completed primary documents.
    direct_primary_res = await db.execute(
        select(Document).where(
            Document.subject_id == parsed_subject_id,
            Document.index_type == "primary",
            Document.processing_status == "completed",
        )
    )
    direct_primary_docs = list(direct_primary_res.scalars().all())

    completed_docs: List[Document] = direct_primary_docs
    discovery_strategy = "direct_primary"

    # Strategy 2: completed documents used in prior generation sessions for this subject.
    if not completed_docs:
        session_docs_res = await db.execute(
            select(Document)
            .distinct()
            .join(GenerationSession, GenerationSession.document_id == Document.id)
            .where(
                GenerationSession.subject_id == parsed_subject_id,
                Document.processing_status == "completed",
            )
        )
        completed_docs = list(session_docs_res.scalars().all())
        if completed_docs:
            discovery_strategy = "generation_sessions"

    # Strategy 3: any completed document linked to this subject.
    if not completed_docs:
        any_completed_res = await db.execute(
            select(Document).where(
                Document.subject_id == parsed_subject_id,
                Document.processing_status == "completed",
            )
        )
        completed_docs = list(any_completed_res.scalars().all())
        if completed_docs:
            discovery_strategy = "all_completed_subject_docs"

    primary_docs = [doc for doc in completed_docs if doc.index_type == "primary"]
    reference_docs = [doc for doc in completed_docs if doc.index_type == "reference_book"]
    template_docs = [doc for doc in completed_docs if doc.index_type in {"template_paper", "reference_questions"}]
    total_chunks = sum(doc.total_chunks or 0 for doc in completed_docs)
    
    return {
        "subject_id": str(subject_id),
        "primary_documents": len(primary_docs),
        "completed_documents": len(completed_docs),
        "reference_documents": len(reference_docs),
        "template_documents": len(template_docs),
        "total_chunks": total_chunks,
        "discovery_strategy": discovery_strategy,
    }


@router.post("/schedule-background-generation")
async def schedule_background_generation(
    request: Request,
    subject_id: str = Form(..., description="Subject ID to generate questions for"),
    count: int = Form(default=10, ge=1, le=200, description="Target number of questions"),
    types: str = Form(default="mcq", description="Comma-separated question types"),
    difficulty: str = Form(default="medium", description="easy, medium, hard"),
    topic_id: Optional[str] = Form(default=None, description="Optional topic ID to generate only for a single topic"),
    topic_ids: Optional[str] = Form(default=None, description="Optional comma-separated topic IDs for multi-topic generation"),
    allow_without_reference: bool = Form(default=False, description="Allow setup completion without reference PDFs"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Schedule server-side generation that can continue after the teacher leaves the page."""
    ensure_user_can_generate(current_user)

    # Start periodic cleanup if not already running
    start_periodic_cleanup()
    valid_types = {"mcq", "short_answer", "long_answer"}
    type_list = [t.strip() for t in types.split(",") if t.strip()]
    if not type_list:
        type_list = ["mcq"]
    if any(t not in valid_types for t in type_list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid types. Allowed: {', '.join(sorted(valid_types))}",
        )
    if difficulty not in {"easy", "medium", "hard"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid difficulty. Must be easy, medium, or hard",
        )

    try:
        parsed_subject_id = str(subject_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subject_id format")

    subject_res = await db.execute(
        select(Subject).where(Subject.id == parsed_subject_id)
    )
    subject = subject_res.scalar_one_or_none()
    if not subject:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Subject not found")

    parsed_topic_id: Optional[str] = None
    parsed_topic_ids: list[str] = []
    if topic_id:
        try:
            parsed_topic_id = str(topic_id)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid topic_id format")

        topic_res = await db.execute(
            select(Topic).where(Topic.id == parsed_topic_id, Topic.subject_id == parsed_subject_id)
        )
        if not topic_res.scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Topic not found for this subject")

    if topic_ids:
        for raw_id in [item.strip() for item in topic_ids.split(",") if item.strip()]:
            try:
                parsed_topic_ids.append(str(raw_id))
            except ValueError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid topic_ids format")

        if parsed_topic_ids:
            topic_rows_res = await db.execute(
                select(Topic.id).where(Topic.subject_id == parsed_subject_id, Topic.id.in_(parsed_topic_ids))
            )
            valid_topic_ids = set(topic_rows_res.scalars().all())
            if len(valid_topic_ids) != len(set(parsed_topic_ids)):
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="One or more topic_ids are invalid for this subject")

    if parsed_topic_id and parsed_topic_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Use either topic_id or topic_ids, not both")

    no_reference_fallback = allow_without_reference
    all_topic_ids = ([parsed_topic_id] if parsed_topic_id else []) + parsed_topic_ids
    topic_name_rows = []
    if all_topic_ids:
        topic_name_result = await db.execute(
            select(Topic.id, Topic.name).where(Topic.id.in_(all_topic_ids))
        )
        topic_name_rows = topic_name_result.all()
    topic_name_map = {topic_row[0]: topic_row[1] for topic_row in topic_name_rows}
    topic_names = [topic_name_map[topic_value] for topic_value in all_topic_ids if topic_value in topic_name_map]

    task_key = _bg_gen_task_key(current_user.id, parsed_subject_id)
    existing_task = _BACKGROUND_GENERATION_TASKS.get(task_key)
    if existing_task and not existing_task.done():
        existing_status = _BACKGROUND_GENERATION_STATUS.get(task_key, {})
        return {
            "status": "already_running",
            "message": "Background generation is already running for this subject",
            "subject_id": str(parsed_subject_id),
            "run_id": existing_status.get("run_id"),
            "count": count,
            "progress": existing_status.get("progress", 0),
            "current_question": existing_status.get("current_question", 0),
            "total_questions": existing_status.get("total_questions", count),
        }

    # Check if already queued
    queue_position = _get_queue_position(current_user.id, parsed_subject_id)
    if queue_position > 0:
        return {
            "status": "queued",
            "message": f"Background generation is queued for this subject (position {queue_position})",
            "subject_id": str(parsed_subject_id),
            "queue_position": queue_position,
            "count": count,
        }

    # Check if we have capacity to run immediately or need to queue
    current_running = _get_current_running_count()
    request_data = {
        "types": type_list,
        "difficulty": difficulty,
        "topic_id": parsed_topic_id,
        "topic_ids": parsed_topic_ids or None,
        "allow_without_reference": no_reference_fallback,
    }

    if current_running >= _MAX_CONCURRENT_GENERATIONS:
        # Queue the request
        queue_position = _add_to_queue(current_user.id, parsed_subject_id, count, request_data)
        run_id = str(uuid.uuid4())
        
        # Set queued status
        _BACKGROUND_GENERATION_STATUS[task_key] = {
            "in_progress": True,
            "run_id": run_id,
            "status": "queued",
            "progress": 0,
            "current_question": 0,
            "total_questions": count,
            "started_total_questions": int(subject.total_questions or 0),
            "target_total_questions": int(subject.total_questions or 0) + count,
            "message": f"Queued for background generation (position {queue_position})",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }

        await safe_record_activity(
            user=current_user,
            action_key="generation_started",
            action_label="Started Generation",
            category="generation",
            source_area="teacher_subjects",
            entity_type="subject",
            entity_id=subject.id,
            entity_name=subject.name,
            subject_id=subject.id,
            subject_name=subject.name,
            topic_id=parsed_topic_id,
            topic_name=topic_names[0] if len(topic_names) == 1 else None,
            details={
                "count": count,
                "types": type_list,
                "difficulty": difficulty,
                "status": "queued",
                "queue_position": queue_position,
                "topic_ids": all_topic_ids or None,
                "topic_names": topic_names or None,
                "allow_without_reference": no_reference_fallback,
            },
            request=request,
        )
        
        return {
            "status": "queued",
            "message": f"Background generation queued (position {queue_position}) due to high demand",
            "subject_id": str(parsed_subject_id),
            "run_id": run_id,
            "queue_position": queue_position,
            "count": count,
            "types": type_list,
            "difficulty": difficulty,
        }

    # Run immediately
    run_id = str(uuid.uuid4())

    _BACKGROUND_GENERATION_STATUS[task_key] = {
        "in_progress": True,
        "run_id": run_id,
        "status": "scheduled",
        "progress": 0,
        "current_question": 0,
        "total_questions": count,
        "started_total_questions": int(subject.total_questions or 0),
        "target_total_questions": int(subject.total_questions or 0) + count,
        "message": "Background generation scheduled",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    task = asyncio.create_task(
        _run_background_subject_generation(
            user_id=current_user.id,
            subject_id=parsed_subject_id,
            count=count,
            types=type_list,
            difficulty=difficulty,
            run_id=run_id,
            topic_id=parsed_topic_id,
            topic_ids=parsed_topic_ids or None,
            allow_without_reference=no_reference_fallback,
        )
    )
    _BACKGROUND_GENERATION_TASKS[task_key] = task

    await safe_record_activity(
        user=current_user,
        action_key="generation_started",
        action_label="Started Generation",
        category="generation",
        source_area="teacher_subjects",
        entity_type="subject",
        entity_id=subject.id,
        entity_name=subject.name,
        subject_id=subject.id,
        subject_name=subject.name,
        topic_id=parsed_topic_id,
        topic_name=topic_names[0] if len(topic_names) == 1 else None,
        details={
            "count": count,
            "types": type_list,
            "difficulty": difficulty,
            "status": "scheduled",
            "run_id": run_id,
            "topic_ids": all_topic_ids or None,
            "topic_names": topic_names or None,
            "allow_without_reference": no_reference_fallback,
        },
        request=request,
    )
    
    # Process queue after task completes (this will be handled in the finally block of _run_background_subject_generation)

    return {
        "status": "scheduled",
        "message": (
            "Background generation scheduled (no PDF fallback enabled)"
            if no_reference_fallback
            else "Background generation scheduled"
        ),
        "subject_id": str(parsed_subject_id),
        "run_id": run_id,
        "count": count,
        "types": type_list,
        "difficulty": difficulty,
        "topic_id": str(parsed_topic_id) if parsed_topic_id else None,
        "topic_ids": [str(item) for item in parsed_topic_ids] if parsed_topic_ids else None,
    }


@router.get("/background-generation-statuses")
async def get_background_generation_statuses(
    subject_ids: str = Query(..., description="Comma-separated subject IDs"),
    current_user: User = Depends(get_current_user),
):
    """Get in-progress background generation status for multiple subjects."""
    raw_ids = [item.strip() for item in subject_ids.split(",") if item.strip()]
    if not raw_ids:
        return {"statuses": {}}

    statuses: dict[str, dict] = {}
    for subject_id_str in raw_ids:
        try:
            parsed_subject_id = str(subject_id_str)
        except ValueError:
            continue

        task_key = _bg_gen_task_key(current_user.id, parsed_subject_id)
        status_payload = _BACKGROUND_GENERATION_STATUS.get(task_key)
        task = _BACKGROUND_GENERATION_TASKS.get(task_key)

        # Check for terminal status first (task finished, status still cached)
        if status_payload and status_payload.get("in_progress") is False:
            # Return the terminal status so the frontend can see completion/error
            statuses[str(parsed_subject_id)] = {
                "in_progress": False,
                "run_id": status_payload.get("run_id"),
                "status": status_payload.get("status", "completed"),
                "progress": status_payload.get("progress", 100),
                "current_question": status_payload.get("current_question", 0),
                "total_questions": status_payload.get("total_questions"),
                "started_total_questions": status_payload.get("started_total_questions"),
                "target_total_questions": status_payload.get("target_total_questions"),
                "message": status_payload.get("message", "Generation complete"),
                "updated_at": status_payload.get("updated_at"),
            }
            continue

        if status_payload and status_payload.get("in_progress") is True:
            statuses[str(parsed_subject_id)] = {
                "in_progress": True,
                "run_id": status_payload.get("run_id"),
                "status": status_payload.get("status", "generating"),
                "progress": status_payload.get("progress", 0),
                "current_question": status_payload.get("current_question", 0),
                "total_questions": status_payload.get("total_questions"),
                "started_total_questions": status_payload.get("started_total_questions"),
                "target_total_questions": status_payload.get("target_total_questions"),
                "message": status_payload.get("message", "Background generation in progress"),
                "updated_at": status_payload.get("updated_at"),
            }
            if not task:
                _bg_status_logger.warning(
                    "Returning in-progress payload without task entry",
                    extra={
                        "user_id": str(current_user.id),
                        "subject_id": str(parsed_subject_id),
                        "task_key": task_key,
                        "status": status_payload.get("status"),
                        "progress": status_payload.get("progress"),
                    },
                )
            continue

        if not task:
            if status_payload:
                _bg_status_logger.warning(
                    "Skipping payload with missing task and non-in-progress flag",
                    extra={
                        "user_id": str(current_user.id),
                        "subject_id": str(parsed_subject_id),
                        "task_key": task_key,
                        "payload_status": status_payload.get("status"),
                        "payload_in_progress": status_payload.get("in_progress"),
                    },
                )
            continue

        if task.done():
            _bg_status_logger.info(
                "Task finished before status payload switched to terminal",
                extra={
                    "user_id": str(current_user.id),
                    "subject_id": str(parsed_subject_id),
                    "task_key": task_key,
                    "has_status_payload": bool(status_payload),
                    "payload_status": (status_payload or {}).get("status") if status_payload else None,
                },
            )
            _BACKGROUND_GENERATION_TASKS.pop(task_key, None)
            # Don't pop status here — the finally block in the task sets a terminal status
            continue

        statuses[str(parsed_subject_id)] = {
            "in_progress": True,
            "run_id": (status_payload or {}).get("run_id"),
            "status": (status_payload or {}).get("status", "generating"),
            "progress": (status_payload or {}).get("progress", 0),
            "current_question": (status_payload or {}).get("current_question", 0),
            "total_questions": (status_payload or {}).get("total_questions"),
            "started_total_questions": (status_payload or {}).get("started_total_questions"),
            "target_total_questions": (status_payload or {}).get("target_total_questions"),
            "message": (status_payload or {}).get("message", "Background generation in progress"),
            "updated_at": (status_payload or {}).get("updated_at"),
        }

    return {"statuses": statuses}


@router.get("/topic-generation-statuses")
async def get_topic_generation_statuses(
    topic_ids: str = Query(..., description="Comma-separated topic IDs"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get generation status for specific topics from the database.
    This provides cross-device visibility of generation status.
    """
    raw_ids = [item.strip() for item in topic_ids.split(",") if item.strip()]
    if not raw_ids:
        return {"statuses": {}}

    statuses: dict[str, dict] = {}
    
    # Query the database for active generation runs
    result = await db.execute(
        select(GenerationRun).where(
            GenerationRun.topic_id.in_(raw_ids),
            GenerationRun.in_progress == True,
        )
    )
    active_runs = result.scalars().all()
    
    for run in active_runs:
        statuses[run.topic_id] = run.to_status_dict()
    
    # Also check for recently completed runs (within last 60 seconds) for topics without active runs
    from datetime import timedelta
    recent_cutoff = datetime.now(timezone.utc) - timedelta(seconds=60)
    
    missing_topic_ids = [tid for tid in raw_ids if tid not in statuses]
    if missing_topic_ids:
        recent_result = await db.execute(
            select(GenerationRun).where(
                GenerationRun.topic_id.in_(missing_topic_ids),
                GenerationRun.in_progress == False,
                GenerationRun.completed_at >= recent_cutoff,
            ).order_by(GenerationRun.completed_at.desc())
        )
        recent_runs = recent_result.scalars().all()
        
        # Only include the most recent completed run per topic
        seen_topics = set()
        for run in recent_runs:
            if run.topic_id not in seen_topics:
                statuses[run.topic_id] = run.to_status_dict()
                seen_topics.add(run.topic_id)
    
    return {"statuses": statuses}


@router.post("/backfill-topic-mapping")
async def backfill_topic_mapping(
    subject_id: str = Form(..., description="Subject ID to backfill topic mappings for"),
    limit: int = Form(default=500, ge=1, le=5000, description="Max unmapped questions to scan"),
    dry_run: bool = Form(default=False, description="If true, compute mappings without writing to DB"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """One-click backfill: auto-map existing unmapped subject questions to topics."""
    try:
        parsed_subject_id = str(subject_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subject_id format")

    service = QuestionGenerationService(db)
    try:
        result = await service.backfill_unmapped_questions_for_subject(
            user_id=current_user.id,
            subject_id=parsed_subject_id,
            limit=limit,
            dry_run=dry_run,
        )
        return {"status": "ok", **result}
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Backfill failed: {str(e)}",
        )


@router.get("", response_model=QuestionListResponse)
async def list_questions(
    document_id: Optional[str] = Query(None, description="Document ID to get questions for"),
    subject_id: Optional[str] = Query(None, description="Subject ID to filter by"),
    topic_id: Optional[str] = Query(None, description="Topic ID to filter by"),
    vetting_status: Optional[str] = Query(None, description="Filter by vetting status (pending, approved, rejected)"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    question_type: Optional[str] = Query(None, description="Filter by question type"),
    difficulty: Optional[str] = Query(None, description="Filter by difficulty"),
    show_archived: bool = Query(False, description="Show only archived questions when true"),
    include_all_versions: bool = Query(False, description="Include all versions including replaced questions"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    List questions with pagination and filtering.
    Can filter by document, subject, topic, or vetting status.
    Set show_archived=true to view archived questions only.
    Set include_all_versions=true to see all versions (including old regenerated ones).
    """
    question_service = QuestionGenerationService(db)
    
    try:
        questions, pagination = await question_service.get_questions(
            user_id=current_user.id,
            document_id=document_id,
            subject_id=subject_id,
            topic_id=topic_id,
            vetting_status=vetting_status,
            page=page,
            limit=limit,
            question_type=question_type,
            difficulty=difficulty,
            show_archived=show_archived,
            include_all_versions=include_all_versions,
        )
        
        return QuestionListResponse(
            questions=[QuestionResponse.from_orm_with_sources(q) for q in questions],
            pagination=pagination,
        )
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e),
        )


@router.get("/{question_id}", response_model=QuestionResponse)
async def get_question(
    question_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get a specific question by ID.
    """
    from sqlalchemy import select, or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
            )
        )
    )
    question = result.scalar_one_or_none()
    
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    return QuestionResponse.from_orm_with_sources(question)


@router.get("/{question_id}/versions", response_model=List[QuestionResponse])
async def get_question_versions(
    question_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get all versions of a question (version history / regeneration chain).
    Returns questions ordered by version_number ascending (oldest first).
    """
    from sqlalchemy import select, or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    
    # First verify the question exists and user has access
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
            )
        )
    )
    question = result.scalar_one_or_none()
    
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    # Walk the version chain backwards to find the original (first version)
    original_id = question.id
    current = question
    while current.replaces_id:
        res = await db.execute(select(Question).where(Question.id == current.replaces_id))
        prev = res.scalar_one_or_none()
        if prev:
            original_id = prev.id
            current = prev
        else:
            break
    
    # Now collect all versions starting from the original
    versions = []
    current_id = original_id
    while current_id:
        res = await db.execute(select(Question).where(Question.id == current_id))
        q = res.scalar_one_or_none()
        if q:
            versions.append(q)
            current_id = q.replaced_by_id
        else:
            break
    
    return [QuestionResponse.from_orm_with_sources(v) for v in versions]


@router.post("/{question_id}/promote-version", response_model=QuestionResponse)
async def promote_question_version(
    question_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Promote an older version of a question to be the current (latest) version.
    This creates a new version that copies the content from the specified question
    and marks it as is_latest=True, while setting is_latest=False on the previous current version.
    """
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    from sqlalchemy import or_
    from datetime import datetime, timezone
    
    # Load the question to promote
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
            )
        )
    )
    source_question = result.scalar_one_or_none()
    
    if not source_question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    # If this is already the latest version, just return it
    if source_question.is_latest:
        return QuestionResponse.from_orm_with_sources(source_question)
    
    # Find the current latest version in the chain by walking forward
    current_latest = source_question
    while True:
        # Find questions that replace the current one
        next_res = await db.execute(
            select(Question).where(Question.replaces_id == current_latest.id)
        )
        next_q = next_res.scalar_one_or_none()
        if next_q:
            current_latest = next_q
        else:
            break
    
    # Mark the current latest as not latest
    current_latest.is_latest = False
    
    # Determine the next version number
    next_version = (current_latest.version_number or 1) + 1
    
    # Create a new question that copies content from source_question
    # but is a new version in the chain
    new_question = Question(
        document_id=source_question.document_id,
        subject_id=source_question.subject_id,
        topic_id=source_question.topic_id,
        session_id=source_question.session_id,
        question_text=source_question.question_text,
        question_embedding=source_question.question_embedding,
        question_type=source_question.question_type,
        marks=source_question.marks,
        difficulty_level=source_question.difficulty_level,
        bloom_taxonomy_level=source_question.bloom_taxonomy_level,
        options=source_question.options,
        correct_answer=source_question.correct_answer,
        explanation=source_question.explanation,
        topic_tags=source_question.topic_tags,
        source_chunk_ids=source_question.source_chunk_ids,
        learning_outcome_id=source_question.learning_outcome_id,
        course_outcome_mapping=source_question.course_outcome_mapping,
        answerability_score=source_question.answerability_score,
        specificity_score=source_question.specificity_score,
        generation_confidence=source_question.generation_confidence,
        generation_metadata={"promoted_from": str(source_question.id)},
        generated_at=datetime.now(timezone.utc),
        vetting_status="pending",  # Reset to pending for review
        is_archived=False,
        replaces_id=current_latest.id,
        version_number=next_version,
        is_latest=True,
    )
    
    db.add(new_question)
    await db.commit()
    await db.refresh(new_question)
    
    return QuestionResponse.from_orm_with_sources(new_question)


@router.post("/{question_id}/rate", response_model=QuestionResponse)
async def rate_question(
    question_id: str,
    rating_data: QuestionRatingRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Rate a question for quality feedback.
    """
    question_service = QuestionGenerationService(db)
    
    try:
        question = await question_service.rate_question(
            question_id=question_id,
            user_id=current_user.id,
            rating=rating_data.rating,
            difficulty_rating=rating_data.difficulty_rating,
        )
        
        return QuestionResponse.from_orm_with_sources(question)
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e),
        )


@router.delete("/{question_id}")
async def archive_question(
    question_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Archive a question (soft delete).
    Archived questions remain in blacklist but won't be shown.
    """
    question_service = QuestionGenerationService(db)
    
    success = await question_service.archive_question(
        question_id=question_id,
        user_id=current_user.id,
    )
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    return {"message": "Question archived successfully"}


@router.post("/{question_id}/unarchive")
async def unarchive_question(
    question_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Unarchive a question (restore it).
    Restores an archived question to make it visible again.
    """
    question_service = QuestionGenerationService(db)
    
    success = await question_service.unarchive_question(
        question_id=question_id,
        user_id=current_user.id,
    )
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Archived question not found",
        )
    
    return {"message": "Question unarchived successfully"}


# ============== Excel/CSV Import Endpoint ==============

@router.post("/import")
async def import_questions_from_file(
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV file with questions"),
    subject_id: str = Form(..., description="Subject ID to link questions to"),
    topic_id: Optional[str] = Form(default=None, description="Topic/Chapter ID to link questions to"),
    current_user: User = Depends(rate_limit(requests=50, window_seconds=86400)),
    db: AsyncSession = Depends(get_db),
):
    """
    Import questions from an Excel (.xlsx) or CSV file.
    
    Expected format (row 2 = headers):
    question | option_a | option_b | option_c | option_d | option_correct | CO | LO mapping | difficulty | marks | course_code | topic | Reference
    
    If option_a through option_d are present → MCQ, otherwise → short_answer.
    """
    import csv
    import io
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Validate file extension
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type: {ext}. Only .xlsx and .csv are supported.",
        )
    
    # Parse subject/topic IDs
    try:
        parsed_subject_id = str(subject_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid subject_id")
    parsed_topic_id = str(topic_id) if topic_id else None
    
    # Read file content
    content = await file.read()
    
    rows = []
    headers = []
    
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            
            # Find headers — check row 1 and row 2
            row1_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
            row2_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[2]] if ws.max_row >= 2 else []
            
            # If row 2 looks like headers (contains 'question'), use row 2
            if any("question" in v for v in row2_values):
                headers = row2_values
                data_start_row = 3
            elif any("question" in v for v in row1_values):
                headers = row1_values
                data_start_row = 2
            else:
                raise HTTPException(status_code=400, detail="Could not find 'question' column in headers (checked rows 1-2)")
            
            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                if row and any(v is not None for v in row):
                    rows.append([str(v).strip() if v is not None else "" for v in row])
        
        elif ext == ".csv":
            text = content.decode("utf-8-sig")  # handle BOM
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            
            if not all_rows:
                raise HTTPException(status_code=400, detail="CSV file is empty")
            
            # Check first two rows for headers
            row1 = [v.strip().lower() for v in all_rows[0]]
            row2 = [v.strip().lower() for v in all_rows[1]] if len(all_rows) > 1 else []
            
            if any("question" in v for v in row2):
                headers = row2
                rows = [[v.strip() for v in r] for r in all_rows[2:] if any(v.strip() for v in r)]
            elif any("question" in v for v in row1):
                headers = row1
                rows = [[v.strip() for v in r] for r in all_rows[1:] if any(v.strip() for v in r)]
            else:
                raise HTTPException(status_code=400, detail="Could not find 'question' column in headers")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to parse file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")
    
    # Map column indices
    def find_col(names):
        """Find column index matching any of the given names."""
        for name in names:
            for i, h in enumerate(headers):
                if name in h:
                    return i
        return None
    
    col_question = find_col(["question"])
    col_opt_a = find_col(["option_a"])
    col_opt_b = find_col(["option_b"])
    col_opt_c = find_col(["option_c"])
    col_opt_d = find_col(["option_d"])
    col_correct = find_col(["option_correct", "correct", "answer"])
    col_co = find_col(["co", "course_outcome"])
    col_lo = find_col(["lo", "learning"])
    col_difficulty = find_col(["difficulty"])
    col_marks = find_col(["marks", "mark"])
    col_topic = find_col(["topic"])
    col_reference = find_col(["reference"])
    
    if col_question is None:
        raise HTTPException(status_code=400, detail="Required column 'question' not found in headers")
    
    # Create generation session for this import
    session_id = uuid.uuid4()
    gen_session = GenerationSession(
        id=session_id,
        user_id=current_user.id,
        subject_id=parsed_subject_id,
        topic_id=parsed_topic_id,
        generation_method="import",
        requested_count=len(rows),
        status="completed",
        started_at=datetime.now(timezone.utc),
        completed_at=datetime.now(timezone.utc),
        generation_config={"source_file": filename, "format": ext},
    )
    db.add(gen_session)
    
    # Import questions
    imported = 0
    skipped = 0
    
    for row_data in rows:
        def get_val(col_idx):
            if col_idx is not None and col_idx < len(row_data):
                return row_data[col_idx]
            return ""
        
        question_text = get_val(col_question)
        if not question_text:
            skipped += 1
            continue
        
        # Determine options
        opt_a = get_val(col_opt_a)
        opt_b = get_val(col_opt_b)
        opt_c = get_val(col_opt_c)
        opt_d = get_val(col_opt_d)
        
        options = [o for o in [opt_a, opt_b, opt_c, opt_d] if o]
        is_mcq = len(options) >= 2
        
        # Correct answer
        correct = get_val(col_correct)
        correct_answer = None
        if correct:
            # Map letter answer to full text
            answer_map = {"a": opt_a, "b": opt_b, "c": opt_c, "d": opt_d}
            correct_answer = answer_map.get(correct.lower().strip(), correct)
        
        # CO mapping
        co_text = get_val(col_co)
        co_mapping = None
        if co_text:
            # Parse "CO1" or "CO1, CO2" format
            co_mapping = {}
            for co in re.findall(r'CO\d+', co_text, re.IGNORECASE):
                co_mapping[co.upper()] = 1
        
        # Marks
        marks_val = get_val(col_marks)
        marks = None
        if marks_val:
            try:
                marks = int(float(marks_val))
            except (ValueError, TypeError):
                pass
        
        # Difficulty
        difficulty = get_val(col_difficulty).lower() or None
        if difficulty and difficulty not in ("easy", "medium", "hard"):
            difficulty = "medium"
        
        # Topic tags
        topic_text = get_val(col_topic)
        topic_tags = [topic_text] if topic_text else None
        
        # LO mapping
        lo_text = get_val(col_lo)
        learning_outcome_id = None
        if lo_text:
            lo_match = re.search(r'LO\d+', lo_text, re.IGNORECASE)
            learning_outcome_id = lo_match.group(0).upper() if lo_match else lo_text
        
        question = Question(
            id=uuid.uuid4(),
            subject_id=parsed_subject_id,
            topic_id=parsed_topic_id,
            session_id=session_id,
            question_text=question_text,
            question_type="mcq" if is_mcq else "short_answer",
            options=options if is_mcq else None,
            correct_answer=correct_answer,
            marks=marks,
            difficulty_level=difficulty,
            bloom_taxonomy_level=None,
            course_outcome_mapping=co_mapping,
            learning_outcome_id=learning_outcome_id,
            topic_tags=topic_tags,
            vetting_status="pending",
            generation_metadata={"source": "import", "file": filename},
        )
        db.add(question)
        imported += 1
    
    gen_session.questions_generated = imported
    gen_session.questions_failed = skipped
    
    await db.commit()
    
    logger.info(f"Import complete: {imported} questions imported, {skipped} skipped from {filename}")
    
    return {
        "message": f"Successfully imported {imported} questions",
        "imported": imported,
        "skipped": skipped,
        "session_id": str(session_id),
        "subject_id": subject_id,
    }


@router.get("/sessions/list")
async def list_generation_sessions(
    document_id: Optional[str] = Query(None, description="Filter by document"),
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    List question generation sessions.
    """
    from app.models.subject import Subject, Topic
    
    query = (
        select(
            GenerationSession,
            Subject.name.label("subject_name"),
            Subject.code.label("subject_code"),
            Topic.name.label("topic_name"),
        )
        .outerjoin(Subject, GenerationSession.subject_id == Subject.id)
        .outerjoin(Topic, GenerationSession.topic_id == Topic.id)
        .where(GenerationSession.user_id == current_user.id)
        .order_by(GenerationSession.started_at.desc())
    )
    
    if document_id:
        query = query.where(GenerationSession.document_id == document_id)
    if subject_id:
        query = query.where(GenerationSession.subject_id == subject_id)
    
    # Count total
    count_q = select(func.count()).select_from(
        select(GenerationSession.id)
        .where(GenerationSession.user_id == current_user.id)
    )
    if document_id:
        count_q = select(func.count()).select_from(
            select(GenerationSession.id)
            .where(GenerationSession.user_id == current_user.id, GenerationSession.document_id == document_id)
        )
    if subject_id:
        count_q = select(func.count()).select_from(
            select(GenerationSession.id)
            .where(GenerationSession.user_id == current_user.id, GenerationSession.subject_id == subject_id)
        )
    count_result = await db.execute(count_q)
    total = count_result.scalar_one()
    
    # Paginate
    offset = (page - 1) * limit
    result = await db.execute(query.offset(offset).limit(limit))
    rows = result.all()
    
    return {
        "sessions": [
            {
                "id": str(s.id),
                "document_id": str(s.document_id) if s.document_id else None,
                "subject_id": str(s.subject_id) if s.subject_id else None,
                "subject_name": subject_name or None,
                "subject_code": subject_code or None,
                "topic_id": str(s.topic_id) if s.topic_id else None,
                "topic_name": topic_name or None,
                "generation_method": s.generation_method,
                "requested_count": s.requested_count,
                "requested_difficulty": s.requested_difficulty,
                "focus_topics": s.focus_topics,
                "questions_generated": s.questions_generated,
                "questions_failed": s.questions_failed,
                "status": s.status,
                "started_at": s.started_at.isoformat() if s.started_at else None,
                "completed_at": s.completed_at.isoformat() if s.completed_at else None,
                "generation_config": s.generation_config,
            }
            for s, subject_name, subject_code, topic_name in rows
        ],
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit,
        },
    }


@router.get("/sessions/{session_id}/questions")
async def get_session_questions(
    session_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get all questions from a specific generation session.
    """
    # Verify session belongs to user
    session_result = await db.execute(
        select(GenerationSession).where(
            GenerationSession.id == session_id,
            GenerationSession.user_id == current_user.id,
        )
    )
    session = session_result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get questions for this session
    questions_result = await db.execute(
        select(Question).where(
            Question.session_id == session_id,
        ).order_by(Question.generated_at.asc())
    )
    questions = questions_result.scalars().all()
    
    # Build version chains: group questions so replacements are stacked with newest first
    # 1. Build lookup maps
    q_by_id = {q.id: q for q in questions}
    
    # 2. Find chain roots and build chains
    # A chain root is a question where replaces_id is None or points outside this session
    chains = []  # List of lists, each list is a version chain (newest first)
    visited = set()
    
    for q in questions:
        if q.id in visited:
            continue
        
        # Walk backwards to find the root of this chain
        root = q
        while root.replaces_id and root.replaces_id in q_by_id:
            root = q_by_id[root.replaces_id]
        
        # Walk forwards from root to build the chain
        chain = []
        current = root
        while current:
            chain.append(current)
            visited.add(current.id)
            # Find next version (the one that has replaces_id == current.id)
            next_version = None
            for nq in questions:
                if nq.replaces_id == current.id and nq.id not in visited:
                    next_version = nq
                    break
            current = next_version
        
        # Reverse so newest is first
        chain.reverse()
        chains.append(chain)
    
    # 3. Sort chains by the original (oldest) question's generated_at
    chains.sort(key=lambda c: c[-1].generated_at if c else None)
    
    # 4. Flatten chains into final list
    ordered_questions = []
    for chain in chains:
        ordered_questions.extend(chain)
    
    return {
        "session_id": str(session_id),
        "generation_method": session.generation_method,
        "started_at": session.started_at.isoformat() if session.started_at else None,
        "questions": [
            {
                "id": str(q.id),
                "question_text": q.question_text,
                "question_type": q.question_type,
                "options": q.options,
                "correct_answer": q.correct_answer,
                "marks": q.marks,
                "difficulty_level": q.difficulty_level,
                "bloom_taxonomy_level": q.bloom_taxonomy_level,
                "course_outcome_mapping": q.course_outcome_mapping,
                "learning_outcome_id": q.learning_outcome_id,
                "topic_tags": q.topic_tags,
                "vetting_status": q.vetting_status,
                "generated_at": q.generated_at.isoformat() if q.generated_at else None,
                "version_number": q.version_number,
                "replaces_id": str(q.replaces_id) if q.replaces_id else None,
                "is_latest": q.is_latest,
                "regenerated_by_vetter": bool(
                    q.generation_metadata and q.generation_metadata.get("regenerated_by_vetter")
                ),
            }
            for q in ordered_questions
        ],
    }


@router.get("/sessions/{session_id}", response_model=GenerationSessionResponse)
async def get_generation_session(
    session_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get details of a specific generation session.
    """
    from sqlalchemy import select
    from app.models.question import GenerationSession
    
    result = await db.execute(
        select(GenerationSession).where(
            GenerationSession.id == session_id,
            GenerationSession.user_id == current_user.id,
        )
    )
    session = result.scalar_one_or_none()
    
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found",
        )
    
    return GenerationSessionResponse.model_validate(session)


@router.delete("/sessions/{session_id}")
async def delete_generation_session(
    session_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Delete a generation session and all its associated questions.
    """
    # Verify session belongs to user
    result = await db.execute(
        select(GenerationSession).where(
            GenerationSession.id == session_id,
            GenerationSession.user_id == current_user.id,
        )
    )
    session = result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found",
        )

    # Delete associated questions
    from sqlalchemy import delete as sql_delete
    await db.execute(
        sql_delete(Question).where(Question.session_id == session_id)
    )

    # Delete the session
    await db.delete(session)
    await db.commit()

    return {"message": "Session and associated questions deleted successfully"}


@router.get("/teacher/dashboard")
async def get_teacher_dashboard(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Unified teacher dashboard stats in a single call.
    Returns subjects, documents, question counts, vetting breakdown, and type distribution.
    """
    from sqlalchemy import select, func, case
    from app.models.question import Question, GenerationSession
    from app.models.document import Document
    from app.models.subject import Subject

    # Subjects owned by user
    subjects_result = await db.execute(
        select(func.count()).where(Subject.user_id == current_user.id)
    )
    subjects_count = subjects_result.scalar_one()

    # Uploaded documents (primary + processed)
    docs_result = await db.execute(
        select(func.count()).where(
            Document.user_id == current_user.id,
            Document.processing_status == "completed",
        )
    )
    documents_count = docs_result.scalar_one()

    # Generation sessions
    sessions_result = await db.execute(
        select(func.count()).where(GenerationSession.user_id == current_user.id)
    )
    sessions_count = sessions_result.scalar_one()

    # Question stats — join via session (document_id may be null for subject-based)
    # Use session.user_id as the ownership check
    q_base = (
        select(Question)
        .join(GenerationSession, Question.session_id == GenerationSession.id)
        .where(
            GenerationSession.user_id == current_user.id,
            Question.is_archived == False,
        )
    )

    total_result = await db.execute(
        select(func.count()).select_from(q_base.subquery())
    )
    total_questions = total_result.scalar_one()

    vetting_result = await db.execute(
        select(
            Question.vetting_status,
            func.count(),
        )
        .join(GenerationSession, Question.session_id == GenerationSession.id)
        .where(
            GenerationSession.user_id == current_user.id,
            Question.is_archived == False,
        )
        .group_by(Question.vetting_status)
    )
    vetting_counts = {row[0]: row[1] for row in vetting_result.all()}

    type_result = await db.execute(
        select(Question.question_type, func.count())
        .join(GenerationSession, Question.session_id == GenerationSession.id)
        .where(
            GenerationSession.user_id == current_user.id,
            Question.is_archived == False,
        )
        .group_by(Question.question_type)
    )
    questions_by_type = {row[0]: row[1] for row in type_result.all()}

    pending = vetting_counts.get("pending", 0)
    approved = vetting_counts.get("approved", 0)
    rejected = vetting_counts.get("rejected", 0)
    approval_rate = round((approved / max(approved + rejected, 1)) * 100, 1)

    return {
        "subjects_count": subjects_count,
        "documents_count": documents_count,
        "sessions_count": sessions_count,
        "total_questions": total_questions,
        "pending_questions": pending,
        "approved_questions": approved,
        "rejected_questions": rejected,
        "approval_rate": approval_rate,
        "questions_by_type": questions_by_type,
    }


@router.get("/stats/summary")
async def get_question_stats(
    document_id: Optional[str] = Query(None, description="Filter by document"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get question generation statistics.
    """
    from sqlalchemy import select, func
    from app.models.question import Question, GenerationSession
    from app.models.document import Document
    
    # Base query for user's questions (exclude archived)
    base_query = select(Question).join(Document).where(
        Document.user_id == current_user.id,
        Question.is_archived == False,
    )
    
    if document_id:
        base_query = base_query.where(Question.document_id == document_id)
    
    # Total questions
    total_result = await db.execute(
        select(func.count()).select_from(base_query.subquery())
    )
    total_questions = total_result.scalar_one()
    
    # Questions by type
    type_result = await db.execute(
        select(Question.question_type, func.count())
        .join(Document)
        .where(Document.user_id == current_user.id, Question.is_archived == False)
        .group_by(Question.question_type)
    )
    by_type = {row[0]: row[1] for row in type_result.all()}
    
    # Questions by difficulty
    diff_result = await db.execute(
        select(Question.difficulty_level, func.count())
        .join(Document)
        .where(Document.user_id == current_user.id, Question.is_archived == False)
        .group_by(Question.difficulty_level)
    )
    by_difficulty = {row[0] or "unspecified": row[1] for row in diff_result.all()}
    
    # Total sessions
    session_result = await db.execute(
        select(func.count()).where(GenerationSession.user_id == current_user.id)
    )
    total_sessions = session_result.scalar_one()
    
    # Average generation time
    avg_time_result = await db.execute(
        select(func.avg(GenerationSession.total_duration_seconds))
        .where(GenerationSession.user_id == current_user.id)
    )
    avg_time = avg_time_result.scalar_one() or 0
    
    return {
        "total_questions_generated": total_questions,
        "total_sessions": total_sessions,
        "average_questions_per_session": total_questions / max(total_sessions, 1),
        "questions_by_type": by_type,
        "questions_by_difficulty": by_difficulty,
        "average_generation_time_seconds": round(avg_time, 2),
    }


# ============== Vetting Endpoints ==============

# Rejection reason categories:
# - "modify": Keep similar concept but adjust the question (user wants same question modified)
# - "new": Generate a completely different question
REJECTION_REASON_CATEGORIES = {
    "duplicate": {"category": "new", "guidance": "Generate a completely unique question focusing on different aspects of the content"},
    "too_easy": {"category": "modify", "guidance": "Increase complexity while keeping the same topic focus. Add more analytical depth"},
    "too_hard": {"category": "modify", "guidance": "Simplify the question while keeping the same concept. Test foundational understanding"},
    "unclear": {"category": "modify", "guidance": "Rephrase the question clearly while keeping the same intent. Use precise terminology"},
    "off_topic": {"category": "new", "guidance": "Generate a question that is directly focused on the main subject matter"},
    "incorrect_answer": {"category": "modify", "guidance": "Keep the question but ensure the answer is factually correct and well-validated"},
    "poor_options": {"category": "modify", "guidance": "Keep the question but create better MCQ options that are distinct yet plausible"},
    "too_long": {"category": "modify", "guidance": "Shorten the question while preserving the core concept being tested"},
    "too_short": {"category": "modify", "guidance": "Expand the question with more context while testing the same concept"},
    "needs_improvement": {"category": "modify", "guidance": "Improve the question as specified by the user feedback"},
    "completely_new": {"category": "new", "guidance": "Generate an entirely new question from different source material"},
}

# Legacy flat mapping for quick lookup
REJECTION_REASONS = {k: v["guidance"] for k, v in REJECTION_REASON_CATEGORIES.items()}

class VettingRequest(BaseModel):
    """Schema for vetting a question."""
    status: str  # approved, rejected
    course_outcome_mapping: Optional[dict] = None  # {"CO1": 2, "CO3": 1}
    notes: Optional[str] = None
    rejection_reasons: Optional[List[str]] = None  # ["duplicate", "too_easy", ...]
    custom_feedback: Optional[str] = None  # Custom feedback from voice input or text
    regeneration_mode: Optional[str] = None  # "modify" or "new" - explicit override



class VettingStatsResponse(BaseModel):
    """Schema for vetting statistics."""
    total_generated: int
    total_approved: int
    total_rejected: int
    pending_review: int
    approval_rate: float
    change_from_last_month: float = 0.0


@router.get("/vetting/pending", response_model=QuestionListResponse)
async def get_pending_questions(
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    topic_id: Optional[str] = Query(None, description="Filter by topic"),
    question_type: Optional[str] = Query(None, description="Filter by question type"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get questions pending vetting/review.
    """
    ensure_user_can_vet(current_user)
    question_service = QuestionGenerationService(db)
    
    questions, pagination = await question_service.get_questions(
        user_id=current_user.id,
        subject_id=subject_id,
        topic_id=topic_id,
        vetting_status="pending",
        page=page,
        limit=limit,
        question_type=question_type,
        show_archived=False,  # Never show archived in pending vetting
    )
    
    return QuestionListResponse(
        questions=[QuestionResponse.from_orm_with_sources(q) for q in questions],
        pagination=pagination,
    )


@router.post("/{question_id}/vet", response_model=QuestionResponse)
async def vet_question(
    question_id: str,
    vetting_data: VettingRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Approve or reject a question.
    """
    ensure_user_can_vet(current_user)

    from datetime import datetime, timezone
    from sqlalchemy import or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    from app.schemas.question import QuestionResponse as QR
    from app.services.question_service import QuestionGenerationService
    
    # Support both document-based and rubric-based questions
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
            )
        )
    )
    question = result.scalar_one_or_none()
    
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    if vetting_data.status not in ("approved", "rejected"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status must be 'approved' or 'rejected'",
        )
    
    question.vetting_status = vetting_data.status
    question.vetted_by = current_user.id
    question.vetted_at = datetime.now(timezone.utc)
    question.vetting_notes = vetting_data.notes
    
    if vetting_data.course_outcome_mapping:
        question.course_outcome_mapping = vetting_data.course_outcome_mapping
    
    await db.commit()
    await db.refresh(question)
    
    # Eagerly snapshot the question fields BEFORE regeneration (which may do internal
    # commits/rollbacks that invalidate the ORM object's lazy-loadable attributes).
    question_snapshot = {
        "id": question.id,
        "document_id": question.document_id,
        "subject_id": question.subject_id,
        "topic_id": question.topic_id,
        "session_id": getattr(question, "session_id", None),
        "question_text": question.question_text,
        "question_type": question.question_type,
        "marks": question.marks,
        "difficulty_level": question.difficulty_level,
        "bloom_taxonomy_level": question.bloom_taxonomy_level,
        "correct_answer": question.correct_answer,
        "options": question.options,
        "explanation": question.explanation,
        "topic_tags": question.topic_tags,
        "course_outcome_mapping": question.course_outcome_mapping,
        "learning_outcome_id": question.learning_outcome_id,
        "vetting_status": question.vetting_status,
        "vetted_by": question.vetted_by,
        "vetted_at": question.vetted_at,
        "vetting_notes": question.vetting_notes,
        "answerability_score": question.answerability_score,
        "specificity_score": question.specificity_score,
        "generation_confidence": question.generation_confidence,
        "generated_at": question.generated_at,
        "times_shown": question.times_shown,
        "user_rating": question.user_rating,
        "is_archived": question.is_archived,
        "generation_metadata": question.generation_metadata,
    }

    if vetting_data.status == "rejected":
        replacement = None
        qsvc = QuestionGenerationService(db)
        
        from loguru import logger
        from app.models.question import Question as QuestionModel
        from app.models.subject import Subject, Topic
        from app.services.llm_service import LLMService
        from app.services.embedding_service import EmbeddingService
        from app.services.document_service import DocumentService
        import random

        # ── Build intelligent rejection guidance from reasons ──
        rejection_reasons = vetting_data.rejection_reasons or []
        custom_feedback = vetting_data.custom_feedback or vetting_data.notes or ""
        
        # Determine regeneration mode: "modify" (same concept, different approach) or "new" (completely different)
        regeneration_mode = vetting_data.regeneration_mode  # Explicit override
        
        if not regeneration_mode:
            # Auto-determine based on rejection reasons
            modify_count = 0
            new_count = 0
            for reason in rejection_reasons:
                if reason in REJECTION_REASON_CATEGORIES:
                    if REJECTION_REASON_CATEGORIES[reason]["category"] == "modify":
                        modify_count += 1
                    else:
                        new_count += 1
            
            # Default to "modify" unless there's a clear signal to generate new
            if new_count > modify_count or "duplicate" in rejection_reasons or "off_topic" in rejection_reasons:
                regeneration_mode = "new"
            else:
                regeneration_mode = "modify"
        
        # Build guidance based on mode
        rejection_guidance = []
        for reason in rejection_reasons:
            if reason in REJECTION_REASON_CATEGORIES:
                rejection_guidance.append(REJECTION_REASON_CATEGORIES[reason]["guidance"])
        
        # Add custom feedback to guidance
        if custom_feedback:
            rejection_guidance.append(f"User feedback: {custom_feedback}")
        
        rejection_guidance_str = "\n".join(f"- {g}" for g in rejection_guidance) if rejection_guidance else ""
        
        logger.info(f"Rejection reasons: {rejection_reasons}, mode: {regeneration_mode}, "
                    f"custom: {custom_feedback[:100] if custom_feedback else 'None'}, "
                    f"guidance: {rejection_guidance_str[:200]}...")

        # ── Detect original generation method ──
        gen_method = None
        gen_metadata = question_snapshot.get("generation_metadata") or {}

        # 1. Try session-based detection
        session_id = question_snapshot.get("session_id")
        if session_id:
            try:
                sess_res = await db.execute(
                    select(GenerationSession.generation_method).where(GenerationSession.id == session_id)
                )
                gen_method = sess_res.scalar_one_or_none()
            except Exception:
                pass

        # 2. Fallback to generation_metadata.source
        if not gen_method:
            source = gen_metadata.get("source", "")
            if "chapter" in source:
                gen_method = "chapter"
            elif gen_metadata.get("rubric_id"):
                gen_method = "rubric"
            elif source == "import":
                gen_method = "import"
            elif question_snapshot["document_id"]:
                gen_method = "quick"
            elif question_snapshot["subject_id"]:
                gen_method = "rubric"

        logger.info(f"Question rejected: {question_id}, detected method={gen_method}, "
                     f"doc={question_snapshot['document_id']}, subj={question_snapshot['subject_id']}, topic={question_snapshot['topic_id']}")

        q_type = question_snapshot["question_type"] or "mcq"
        marks = question_snapshot["marks"] or 2

        # ═══════════════════════════════════════════
        # PATH A: Quick Generate (document-based RAG)
        # ═══════════════════════════════════════════
        if gen_method == "quick" and question_snapshot["document_id"]:
            logger.info(f"Regenerating via direct generation (method=quick)")

            llm_service = LLMService()
            embedding_service = EmbeddingService()
            doc_service = DocumentService(db, embedding_service)

            # Fetch document chunks for RAG context
            try:
                from app.models.document import DocumentChunk
                chunk_res = await db.execute(
                    select(DocumentChunk)
                    .where(DocumentChunk.document_id == question_snapshot["document_id"])
                    .order_by(DocumentChunk.chunk_index)
                    .limit(10)
                )
                doc_chunks = chunk_res.scalars().all()
                content_context = "\n\n".join(c.chunk_text for c in doc_chunks if c.chunk_text)[:4000]
            except Exception as e:
                logger.warning(f"Failed to load document chunks for regen: {e}")
                content_context = ""

            if content_context:
                system_prompt = _get_rubric_system_prompt(q_type)
                
                # Preload existing embeddings for subject-level dedupe
                _vetter_existing_embs = []
                if question_snapshot["subject_id"]:
                    _vetter_existing_embs = await _preload_subject_embeddings(db, question_snapshot["subject_id"])

                # Retry loop for robustness
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        import random as _r
                        regen_starter = _r.choice(_QUESTION_STARTERS.get(q_type, _QUESTION_STARTERS["default"]))
                        
                        # Build prompt based on regeneration mode
                        if regeneration_mode == "modify":
                            # Keep same concept, improve the question based on feedback
                            prompt = f"""Content:\n{content_context}\n

TASK: Improve the following question based on user feedback.

ORIGINAL QUESTION:
"{question_snapshot["question_text"]}"

ORIGINAL ANSWER: {question_snapshot["correct_answer"] or "Not specified"}

IMPROVEMENT INSTRUCTIONS:
{rejection_guidance_str if rejection_guidance_str else "- Improve clarity and quality while keeping the same concept"}

REQUIREMENTS:
- Generate a {q_type.replace('_', ' ')} question
- Marks: {marks}
- Difficulty: {question_snapshot["difficulty_level"] or "medium"}
- KEEP the same core concept/topic as the original question
- APPLY the improvement feedback above
- Your question MUST start with "{regen_starter}" or a similar interrogative
- FORBIDDEN: Do NOT start with scenario setups like "A clinician", "A patient", etc.
- CRITICAL: The question must be FULLY SELF-CONTAINED without referencing source material.

Output valid JSON only."""
                        else:
                            # Generate completely new question from different content
                            prompt = f"""Content:\n{content_context}\n

Generate a COMPLETELY NEW {q_type.replace('_', ' ')} question based on this content.

PREVIOUSLY REJECTED QUESTION (DO NOT regenerate similar):
"{question_snapshot["question_text"][:200]}..."

REQUIREMENTS:
- Marks: {marks}
- Difficulty: {question_snapshot["difficulty_level"] or "medium"}
- REQUIRED: Your question MUST start with "{regen_starter}" — this is non-negotiable
- Generate a question about a DIFFERENT TOPIC or ASPECT than the rejected question
- FORBIDDEN: Do NOT start with "A clinician", "A patient", "A doctor", etc.
- CRITICAL: The question must be FULLY SELF-CONTAINED. NEVER reference the source material.
{f'''
FEEDBACK TO ADDRESS:
{rejection_guidance_str}
''' if rejection_guidance_str else ''}
Output valid JSON only."""

                        logger.info(f"Quick regen attempt {attempt + 1}/{max_retries}, mode={regeneration_mode}")
                        response = await llm_service.generate_json(
                            prompt=prompt, system_prompt=system_prompt, temperature=0.7 + (attempt * 0.1),
                        )
                        response = _normalize_llm_response(response)
                        logger.info(f"Quick regen got response: {bool(response)}, keys={list(response.keys()) if response else []}")
                        
                        if response and "question_text" in response:
                            question_text = response["question_text"]
                            if len(question_text) >= 15 and question_text.strip() != (question_snapshot["question_text"] or "").strip():
                                # Validate MCQ has at least 4 options
                                if q_type == "mcq":
                                    norm_opts = _normalize_mcq_options(response.get("options"), "mcq")
                                    if not norm_opts or len(norm_opts) < 4:
                                        logger.warning(f"Quick regen attempt {attempt + 1}: MCQ has only {len(norm_opts) if norm_opts else 0} options, retrying")
                                        continue

                                question_embedding = await embedding_service.get_embedding(question_text)

                                # Dedupe: reject if too similar to existing questions
                                if _is_duplicate_embedding(question_embedding, _vetter_existing_embs, []):
                                    logger.warning(f"Quick regen attempt {attempt + 1}: duplicate of existing question, retrying")
                                    continue

                                _s = _sanitize_question_fields({
                                    "learning_outcome_id": question_snapshot["learning_outcome_id"],
                                    "course_outcome_mapping": question_snapshot["course_outcome_mapping"],
                                    "bloom_taxonomy_level": response.get("bloom_level", question_snapshot["bloom_taxonomy_level"] or "understand"),
                                    "difficulty_level": question_snapshot["difficulty_level"] or "medium",
                                })

                                new_question = QuestionModel(
                                    document_id=question_snapshot["document_id"],
                                    subject_id=question_snapshot["subject_id"],
                                    topic_id=question_snapshot["topic_id"],
                                    session_id=session_id,
                                    question_text=question_text,
                                    question_embedding=question_embedding,
                                    question_type=q_type,
                                    marks=marks,
                                    difficulty_level=_s["difficulty_level"],
                                    bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                                    correct_answer=response.get("correct_answer") or response.get("expected_answer"),
                                    options=_normalize_mcq_options(response.get("options"), q_type),
                                    explanation=response.get("explanation"),
                                    topic_tags=response.get("topic_tags", question_snapshot["topic_tags"]),
                                    learning_outcome_id=_s["learning_outcome_id"],
                                    course_outcome_mapping=_s["course_outcome_mapping"],
                                    generation_confidence=0.75,
                                    vetting_status="pending",
                                    # Version control fields
                                    replaces_id=question_id,
                                    version_number=(question.version_number or 1) + 1,
                                    is_latest=True,
                                    generation_metadata={
                                        "source": "quick_generation",
                                        "regenerated_from": str(question_id),
                                        "source_info": _build_rubric_source_info(
                                            doc_chunks,
                                            question_snapshot.get("topic_tags", ["Document"])[0] if question_snapshot.get("topic_tags") else "Document"
                                        ),
                                    },
                                )
                                db.add(new_question)
                                await db.commit()
                                await db.refresh(new_question)

                                # Update old question to link to replacement
                                question.replaced_by_id = new_question.id
                                question.is_latest = False
                                await db.commit()

                                replacement = new_question
                                logger.info(f"Quick regen: replacement={replacement.id}")
                                break  # Success - exit retry loop
                            else:
                                logger.warning(f"Quick regen attempt {attempt + 1}: question too short or duplicate")
                        else:
                            logger.warning(f"Quick regen attempt {attempt + 1}: response missing question_text")
                    except Exception as e:
                        logger.error(f"Quick regen attempt {attempt + 1} error: {e}")
                        await db.rollback()
                        if attempt == max_retries - 1:
                            logger.error(f"Quick regen failed after {max_retries} attempts")

            if replacement:
                return QR.from_orm_with_sources(replacement)

        # ═══════════════════════════════════════════════════════════
        # PATH A2: Quick-from-subject (RAG across subject reference docs)
        # ═══════════════════════════════════════════════════════════════
        elif gen_method == "quick_from_subject" and question_snapshot["subject_id"]:
            logger.info(f"Regenerating via quick_from_subject for subject_id={question_snapshot['subject_id']}")

            llm_service = LLMService()
            embedding_service = EmbeddingService()
            doc_service = DocumentService(db, embedding_service)

            # Fetch reference document chunks for this subject (primary + reference books)
            content_context = ""
            ref_chunks_a2: list = []
            try:
                rag_query = question_snapshot["question_text"][:300]
                query_emb = await embedding_service.get_query_embedding(rag_query)

                ref_doc_ids = await doc_service.get_reference_document_ids(
                    user_id=current_user.id,
                    subject_id=question_snapshot["subject_id"],
                    index_type="reference_book",
                )
                if ref_doc_ids:
                    ref_chunks_a2 = await doc_service.hybrid_search_multi_document(
                        document_ids=ref_doc_ids,
                        query=rag_query,
                        query_embedding=query_emb,
                        top_k=8,
                        alpha=0.6,
                    )
                if not ref_chunks_a2:
                    # Fallback: any uploaded documents for the subject (from any teacher)
                    from app.models.document import Document as DocModel
                    any_docs_res = await db.execute(
                        select(DocModel.id).where(
                            DocModel.subject_id == question_snapshot["subject_id"],
                            DocModel.processing_status == "completed",
                        ).limit(5)
                    )
                    any_doc_ids = [row[0] for row in any_docs_res.all()]
                    if any_doc_ids:
                        ref_chunks_a2 = await doc_service.hybrid_search_multi_document(
                            document_ids=any_doc_ids,
                            query=rag_query,
                            query_embedding=query_emb,
                            top_k=5,
                            alpha=0.6,
                        )
                if ref_chunks_a2:
                    content_context = "\n\n".join(c.chunk_text for c in ref_chunks_a2 if c.chunk_text)[:4000]
            except Exception as e:
                logger.warning(f"RAG lookup for quick_from_subject regen failed: {e}")

            # Fallback: use topic syllabus content if no reference docs found
            if not content_context and question_snapshot["topic_id"]:
                try:
                    tres = await db.execute(select(Topic).where(Topic.id == question_snapshot["topic_id"]))
                    t_obj = tres.scalar_one_or_none()
                    if t_obj and t_obj.syllabus_content:
                        content_context = t_obj.syllabus_content[:4000]
                except Exception:
                    pass

            if content_context:
                system_prompt = _get_rubric_system_prompt(q_type)
                _vetter_existing_embs_a2 = await _preload_subject_embeddings(db, question_snapshot["subject_id"])

                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        import random as _r
                        regen_starter = _r.choice(_QUESTION_STARTERS.get(q_type, _QUESTION_STARTERS["default"]))

                        if regeneration_mode == "modify":
                            prompt = f"""Content:\n{content_context}\n

TASK: Improve the following question based on user feedback.

ORIGINAL QUESTION:
"{question_snapshot["question_text"]}"

ORIGINAL ANSWER: {question_snapshot["correct_answer"] or "Not specified"}

IMPROVEMENT INSTRUCTIONS:
{rejection_guidance_str if rejection_guidance_str else "- Improve clarity and quality while keeping the same concept"}

REQUIREMENTS:
- Generate a {q_type.replace('_', ' ')} question
- Marks: {marks}
- Difficulty: {question_snapshot["difficulty_level"] or "medium"}
- KEEP the same core concept/topic as the original question
- APPLY the improvement feedback above
- Your question MUST start with "{regen_starter}" or a similar interrogative
- FORBIDDEN: Do NOT start with scenario setups like "A clinician", "A patient", etc.
- CRITICAL: The question must be FULLY SELF-CONTAINED without referencing source material.

Output valid JSON only."""
                        else:
                            feedback_block = (
                                f"\nFEEDBACK TO ADDRESS:\n{rejection_guidance_str}\n"
                                if rejection_guidance_str else ""
                            )
                            prompt = (
                                f"Content:\n{content_context}\n\n"
                                f"Generate a COMPLETELY NEW {q_type.replace('_', ' ')} question based on this content.\n\n"
                                f"PREVIOUSLY REJECTED QUESTION (DO NOT regenerate similar):\n"
                                f'"{question_snapshot["question_text"][:200]}..."\n\n'
                                f"REQUIREMENTS:\n"
                                f"- Marks: {marks}\n"
                                f"- Difficulty: {question_snapshot['difficulty_level'] or 'medium'}\n"
                                f'- Your question MUST start with "{regen_starter}"\n'
                                f"- Generate a question about a DIFFERENT TOPIC or ASPECT than the rejected question\n"
                                f"- FORBIDDEN: Do NOT start with \"A clinician\", \"A patient\", \"A doctor\", etc.\n"
                                f"- CRITICAL: The question must be FULLY SELF-CONTAINED.\n"
                                f"{feedback_block}"
                                f"Output valid JSON only."
                            )

                        logger.info(f"quick_from_subject regen attempt {attempt + 1}/{max_retries}, mode={regeneration_mode}")
                        response = await llm_service.generate_json(
                            prompt=prompt, system_prompt=system_prompt, temperature=0.7 + (attempt * 0.1),
                        )
                        response = _normalize_llm_response(response)

                        if response and "question_text" in response:
                            question_text = response["question_text"]
                            if len(question_text) >= 15 and question_text.strip() != (question_snapshot["question_text"] or "").strip():
                                if q_type == "mcq":
                                    norm_opts = _normalize_mcq_options(response.get("options"), "mcq")
                                    if not norm_opts or len(norm_opts) < 4:
                                        logger.warning(f"quick_from_subject regen attempt {attempt + 1}: MCQ has only {len(norm_opts) if norm_opts else 0} options, retrying")
                                        continue

                                question_embedding = await embedding_service.get_embedding(question_text)
                                if _is_duplicate_embedding(question_embedding, _vetter_existing_embs_a2, []):
                                    logger.warning(f"quick_from_subject regen attempt {attempt + 1}: duplicate, retrying")
                                    continue

                                _s = _sanitize_question_fields({
                                    "learning_outcome_id": question_snapshot["learning_outcome_id"],
                                    "course_outcome_mapping": question_snapshot["course_outcome_mapping"],
                                    "bloom_taxonomy_level": response.get("bloom_level", question_snapshot["bloom_taxonomy_level"] or "understand"),
                                    "difficulty_level": question_snapshot["difficulty_level"] or "medium",
                                })

                                new_question = QuestionModel(
                                    subject_id=question_snapshot["subject_id"],
                                    topic_id=question_snapshot["topic_id"],
                                    session_id=session_id,
                                    question_text=question_text,
                                    question_embedding=question_embedding,
                                    question_type=q_type,
                                    marks=marks,
                                    difficulty_level=_s["difficulty_level"],
                                    bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                                    correct_answer=response.get("correct_answer") or response.get("expected_answer"),
                                    options=_normalize_mcq_options(response.get("options"), q_type),
                                    explanation=response.get("explanation"),
                                    topic_tags=response.get("topic_tags", question_snapshot["topic_tags"]),
                                    learning_outcome_id=_s["learning_outcome_id"],
                                    course_outcome_mapping=_s["course_outcome_mapping"],
                                    generation_confidence=0.75,
                                    vetting_status="pending",
                                    replaces_id=question_id,
                                    version_number=(question.version_number or 1) + 1,
                                    is_latest=True,
                                    generation_metadata={
                                        "source": "quick_from_subject",
                                        "regenerated_by_vetter": False,
                                        "regenerated_from": str(question_id),
                                        "source_info": _build_rubric_source_info(
                                            ref_chunks_a2,
                                            question_snapshot.get("topic_tags", ["Subject"])[0] if question_snapshot.get("topic_tags") else "Subject",
                                        ),
                                    },
                                )
                                db.add(new_question)
                                await db.commit()
                                await db.refresh(new_question)

                                question.replaced_by_id = new_question.id
                                question.is_latest = False
                                await db.commit()

                                replacement = new_question
                                logger.info(f"quick_from_subject regen: replacement={replacement.id}")
                                break
                            else:
                                logger.warning(f"quick_from_subject regen attempt {attempt + 1}: question too short or duplicate")
                        else:
                            logger.warning(f"quick_from_subject regen attempt {attempt + 1}: response missing question_text")
                    except Exception as e:
                        logger.error(f"quick_from_subject regen attempt {attempt + 1} error: {e}")
                        await db.rollback()
                        if attempt == max_retries - 1:
                            logger.error(f"quick_from_subject regen failed after {max_retries} attempts")

            if replacement:
                return QR.from_orm_with_sources(replacement)

        # ═══════════════════════════════════════════════
        # PATH B: Chapter Generate (syllabus content + RAG)
        # ═══════════════════════════════════════════════
        elif gen_method == "chapter" and question_snapshot["subject_id"]:
            topic_id_for_regen = question_snapshot["topic_id"]
            logger.info(f"Regenerating via chapter method for topic_id={topic_id_for_regen}")

            topic_obj = None
            if topic_id_for_regen:
                tres = await db.execute(select(Topic).where(Topic.id == topic_id_for_regen))
                topic_obj = tres.scalar_one_or_none()

            if topic_obj and topic_obj.syllabus_content:
                llm_service = LLMService()
                embedding_service = EmbeddingService()
                syllabus_content = topic_obj.syllabus_content
                topic_name = topic_obj.name
                topic_lo_mappings = topic_obj.learning_outcome_mappings or {}

                # RAG: fetch reference-book chunks using hybrid search
                reference_context = ""
                try:
                    doc_service = DocumentService(db, embedding_service)
                    rag_query = f"{topic_name}: {syllabus_content[:500]}"
                    query_emb = await embedding_service.get_query_embedding(rag_query)
                    
                    # Get reference book IDs for hybrid search
                    ref_doc_ids = await doc_service.get_reference_document_ids(
                        user_id=current_user.id,
                        subject_id=question_snapshot["subject_id"],
                        index_type="reference_book",
                    )
                    
                    ref_chunks = []
                    if ref_doc_ids:
                        ref_chunks = await doc_service.hybrid_search_multi_document(
                            document_ids=ref_doc_ids,
                            query=rag_query,
                            query_embedding=query_emb,
                            top_k=5,
                            alpha=0.6,
                        )
                    
                    if not ref_chunks:
                        # Fallback to vector-only search
                        ref_chunks = await doc_service.get_reference_chunks(
                            user_id=current_user.id,
                            subject_id=question_snapshot["subject_id"],
                            index_type="reference_book",
                            query_embedding=query_emb,
                            top_k=3,
                        )
                    
                    if ref_chunks:
                        reference_context = "\n\n".join(
                            f"[Reference {i+1}]: {c.chunk_text}" for i, c in enumerate(ref_chunks)
                        )
                except Exception as e:
                    logger.warning(f"RAG lookup for chapter regen failed: {e}")

                system_prompt = _get_rubric_system_prompt(q_type)
                
                # Preload existing embeddings for subject-level dedupe
                _vetter_existing_embs_ch = await _preload_subject_embeddings(db, question_snapshot["subject_id"])

                # Retry loop for robustness
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        import random as _r
                        regen_starter = _r.choice(_QUESTION_STARTERS.get(q_type, _QUESTION_STARTERS["default"]))
                        prompt = f"Chapter: \"{topic_name}\"\n\nSyllabus Content:\n{syllabus_content[:3000]}\n"
                        if reference_context:
                            prompt += f"\nBackground knowledge (use to inform the question, do NOT cite or reference it):\n{reference_context}\n"
                        prompt += f"""
Generate a {q_type.replace('_', ' ')} question based on this chapter content.
- Marks: {marks}
- Difficulty: {question_snapshot["difficulty_level"] or "medium"}
- REQUIRED: Your question MUST start with the word/phrase "{regen_starter}" — this is non-negotiable
- FORBIDDEN: Do NOT start all the questions with "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", or similar scenario setups. Also start directly with the required starter word.
- CRITICAL: The question must be FULLY SELF-CONTAINED. NEVER reference the source material in ANY way. Do NOT use ANY of these phrases or similar: "the text", "the reference", "the provided material", "the provided content", "the context", "the passage", "according to", "based on the content", "as stated", "as mentioned", "in the material", "from the content". The question must stand completely alone as if no source was given.
- IMPORTANT: Generate a DIFFERENT question from: "{question_snapshot["question_text"][:150]}..."
{f'''
REJECTION FEEDBACK (address these issues in the new question):
{rejection_guidance_str}
''' if rejection_guidance_str else ''}
Output valid JSON only."""

                        logger.info(f"Chapter regen attempt {attempt + 1}/{max_retries}")
                        response = await llm_service.generate_json(
                            prompt=prompt, system_prompt=system_prompt, temperature=0.7 + (attempt * 0.1),
                        )
                        response = _normalize_llm_response(response)
                        logger.info(f"Chapter regen got response: {bool(response)}, keys={list(response.keys()) if response else []}")
                        
                        if response and "question_text" in response:
                            question_text = response["question_text"]
                            if len(question_text) >= 15 and question_text.strip() != (question_snapshot["question_text"] or "").strip():
                                question_embedding = await embedding_service.get_embedding(question_text)

                                # Dedupe: reject if too similar to existing questions
                                if _is_duplicate_embedding(question_embedding, _vetter_existing_embs_ch, []):
                                    logger.warning(f"Chapter regen attempt {attempt + 1}: duplicate of existing question, retrying")
                                    continue

                                assigned_lo = response.get("learning_outcome_id") or response.get("learning_outcome")
                                if not assigned_lo and topic_lo_mappings:
                                    try:
                                        assigned_lo = max(topic_lo_mappings.items(), key=lambda kv: (kv[1] or 0))[0]
                                    except Exception:
                                        pass

                                assigned_co_map = {}
                                co_val = response.get("course_outcome")
                                if isinstance(co_val, str):
                                    assigned_co_map = {co_val: 1}
                                elif isinstance(co_val, dict):
                                    assigned_co_map = co_val
                                elif assigned_lo:
                                    subj_res = await db.execute(select(Subject).where(Subject.id == question_snapshot["subject_id"]))
                                    subj = subj_res.scalar_one_or_none()
                                    if subj:
                                        s_lo_map = getattr(subj, "learning_outcome_mappings", {}) or {}
                                        if isinstance(s_lo_map.get(assigned_lo), dict):
                                            assigned_co_map = s_lo_map[assigned_lo]

                                _s = _sanitize_question_fields({
                                    "learning_outcome_id": assigned_lo,
                                    "course_outcome_mapping": assigned_co_map,
                                    "bloom_taxonomy_level": response.get("bloom_level", "understand"),
                                    "difficulty_level": question_snapshot["difficulty_level"] or "medium",
                                })

                                new_question = QuestionModel(
                                    subject_id=question_snapshot["subject_id"],
                                    topic_id=topic_id_for_regen,
                                    session_id=session_id,
                                    question_text=question_text,
                                    question_embedding=question_embedding,
                                    question_type=q_type,
                                    marks=marks,
                                    difficulty_level=_s["difficulty_level"],
                                    bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                                    correct_answer=response.get("correct_answer") or response.get("expected_answer"),
                                    options=_normalize_mcq_options(response.get("options"), q_type),
                                    explanation=response.get("explanation"),
                                    topic_tags=response.get("topic_tags", [topic_name]),
                                    learning_outcome_id=_s["learning_outcome_id"],
                                    course_outcome_mapping=_s["course_outcome_mapping"],
                                    generation_confidence=0.75,
                                    vetting_status="pending",
                                    # Version control fields
                                    replaces_id=question_id,
                                    version_number=(question.version_number or 1) + 1,
                                    is_latest=True,
                                    generation_metadata={
                                        "source": "chapter_generation",
                                        "regenerated_from": str(question_id),
                                        "topic_name": topic_name,
                                        "has_reference_context": bool(reference_context),
                                        "source_info": _build_rubric_source_info(
                                            ref_chunks if ref_chunks else [],
                                            topic_name,
                                        ),
                                    },
                                )
                                db.add(new_question)
                                await db.commit()
                                await db.refresh(new_question)
                                
                                # Update old question to link to replacement
                                question.replaced_by_id = new_question.id
                                question.is_latest = False
                                await db.commit()
                                
                                replacement = new_question
                                logger.info(f"Chapter regen: replacement={replacement.id}")
                                break  # Success - exit retry loop
                            else:
                                logger.warning(f"Chapter regen attempt {attempt + 1}: question too short or duplicate")
                        else:
                            logger.warning(f"Chapter regen attempt {attempt + 1}: response missing question_text")
                    except Exception as e:
                        logger.error(f"Chapter regen attempt {attempt + 1} error: {e}")
                        await db.rollback()
                        if attempt == max_retries - 1:
                            logger.error(f"Chapter regen failed after {max_retries} attempts")

            if replacement:
                return QR.from_orm_with_sources(replacement)

        # ═══════════════════════════════════════════════
        # PATH C: Rubric Generate (syllabus chunks + LO/CO)
        # ═══════════════════════════════════════════════
        elif gen_method in ("rubric", None) and question_snapshot["subject_id"]:
            logger.info(f"Regenerating via rubric method for subject_id={question_snapshot['subject_id']}")

            result = await db.execute(
                select(Topic).where(
                    Topic.subject_id == question_snapshot["subject_id"],
                    Topic.has_syllabus == True,
                ).order_by(Topic.order_index)
            )
            topics = result.scalars().all()

            if topics:
                all_chunks = []
                for t in topics:
                    if t.syllabus_content:
                        if question_snapshot["topic_id"] and t.id == question_snapshot["topic_id"] and len(topics) > 1:
                            continue
                        content = t.syllabus_content
                        paragraphs = re.split(r'\n\n+', content)
                        chunk_size = 1500
                        current_chunk = ""
                        for para in paragraphs:
                            if len(current_chunk) + len(para) < chunk_size:
                                current_chunk += "\n\n" + para if current_chunk else para
                            else:
                                if current_chunk:
                                    all_chunks.append({"topic_id": str(t.id), "topic_name": t.name, "content": current_chunk.strip()})
                                current_chunk = para
                        if current_chunk:
                            all_chunks.append({"topic_id": str(t.id), "topic_name": t.name, "content": current_chunk.strip()})

                if all_chunks:
                    llm_service = LLMService()
                    embedding_service = EmbeddingService()
                    system_prompt = _get_rubric_system_prompt(q_type)
                    
                    # Preload existing embeddings for subject-level dedupe
                    _vetter_existing_embs_rb = await _preload_subject_embeddings(db, question_snapshot["subject_id"])

                    # Retry loop for robustness
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            selected_chunk = random.choice(all_chunks)
                            import random as _r
                            regen_starter = _r.choice(_QUESTION_STARTERS.get(q_type, _QUESTION_STARTERS["default"]))
                            prompt = f"""Context from "{selected_chunk['topic_name']}":
{selected_chunk['content']}

Generate a {q_type.replace('_', ' ')} question based on this content.
- Marks: {marks}
- REQUIRED: Your question MUST start with the word/phrase "{regen_starter}" — this is non-negotiable
- FORBIDDEN: Do NOT start all the questions with "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", or similar scenario setups. Also start directly with the required starter word.
- CRITICAL: The question must be FULLY SELF-CONTAINED. NEVER reference the source material in ANY way. Do NOT use ANY of these phrases or similar: "the text", "the reference", "the provided material", "the provided content", "the context", "the passage", "according to", "based on the content", "as stated", "as mentioned", "in the material", "from the content". The question must stand completely alone as if no source was given.
- IMPORTANT: Generate a DIFFERENT question from: "{question_snapshot["question_text"][:100]}..."
{f'''
REJECTION FEEDBACK (address these issues in the new question):
{rejection_guidance_str}
''' if rejection_guidance_str else ''}
Output valid JSON only."""

                            logger.info(f"Rubric regen attempt {attempt + 1}/{max_retries}")
                            response = await llm_service.generate_json(
                                prompt=prompt, system_prompt=system_prompt, temperature=0.7 + (attempt * 0.1),
                            )
                            response = _normalize_llm_response(response)
                            logger.info(f"Rubric regen got response: {bool(response)}, keys={list(response.keys()) if response else []}")
                            
                            if response and "question_text" in response:
                                question_text = response["question_text"]
                                if len(question_text) >= 15 and question_text.strip() != (question_snapshot["question_text"] or "").strip():
                                    # Validate MCQ has at least 4 options
                                    if q_type == "mcq":
                                        norm_opts = _normalize_mcq_options(response.get("options"), "mcq")
                                        if not norm_opts or len(norm_opts) < 4:
                                            logger.warning(f"Rubric regen attempt {attempt + 1}: MCQ has only {len(norm_opts) if norm_opts else 0} options, retrying")
                                            continue

                                    question_embedding = await embedding_service.get_embedding(question_text)

                                    # Dedupe: reject if too similar to existing questions
                                    if _is_duplicate_embedding(question_embedding, _vetter_existing_embs_rb, []):
                                        logger.warning(f"Rubric regen attempt {attempt + 1}: duplicate of existing question, retrying")
                                        continue

                                    assigned_lo = response.get("learning_outcome_id") or response.get("learning_outcome")
                                    topic_obj = None
                                    try:
                                        tres = await db.execute(select(Topic).where(Topic.id == str(selected_chunk["topic_id"])))
                                        topic_obj = tres.scalar_one_or_none()
                                    except Exception:
                                        pass

                                    topic_lo_map = (topic_obj.learning_outcome_mappings if topic_obj and getattr(topic_obj, "learning_outcome_mappings", None) else {}) or {}
                                    if not assigned_lo and topic_lo_map:
                                        try:
                                            assigned_lo = max(topic_lo_map.items(), key=lambda kv: (kv[1] or 0))[0]
                                        except Exception:
                                            pass

                                    assigned_co_map = {}
                                    subj_res = await db.execute(select(Subject).where(Subject.id == question_snapshot["subject_id"]))
                                    subj = subj_res.scalar_one_or_none()
                                    if subj and assigned_lo:
                                        s_lo_map = getattr(subj, "learning_outcome_mappings", {}) or {}
                                        if isinstance(s_lo_map.get(assigned_lo), dict):
                                            assigned_co_map = s_lo_map[assigned_lo]
                                        else:
                                            cos = (subj.course_outcomes or {}).get("outcomes") if subj.course_outcomes else None
                                            if cos and isinstance(cos, list) and len(cos) > 0:
                                                m = re.search(r"LO(\d+)", str(assigned_lo))
                                                if m:
                                                    idx = (int(m.group(1)) - 1) % len(cos)
                                                    assigned_co_map = {f"CO{idx+1}": 1}

                                    _s = _sanitize_question_fields({
                                        "learning_outcome_id": assigned_lo,
                                        "course_outcome_mapping": assigned_co_map,
                                        "bloom_taxonomy_level": response.get("bloom_level", "understand"),
                                        "difficulty_level": question_snapshot["difficulty_level"] or "medium",
                                    })

                                    new_question = QuestionModel(
                                        subject_id=question_snapshot["subject_id"],
                                        topic_id=str(selected_chunk["topic_id"]),
                                        session_id=session_id,
                                        question_text=question_text,
                                        question_embedding=question_embedding,
                                        question_type=q_type,
                                        marks=marks,
                                        difficulty_level=_s["difficulty_level"],
                                        bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                                        correct_answer=response.get("correct_answer") or response.get("expected_answer"),
                                        options=_normalize_mcq_options(response.get("options"), q_type),
                                        explanation=response.get("explanation"),
                                        topic_tags=response.get("topic_tags", [selected_chunk["topic_name"]]),
                                        learning_outcome_id=_s["learning_outcome_id"],
                                        course_outcome_mapping=_s["course_outcome_mapping"],
                                        generation_confidence=0.75,
                                        vetting_status="pending",
                                        # Version control fields
                                        replaces_id=question_id,
                                        version_number=(question.version_number or 1) + 1,
                                        is_latest=True,
                                        generation_metadata={
                                            "regenerated_from": str(question_id),
                                            "topic_name": selected_chunk["topic_name"],
                                            "source_info": {
                                                "sources": [{
                                                    "document_name": f"Syllabus: {selected_chunk['topic_name']}",
                                                    "page_number": None,
                                                    "page_range": None,
                                                    "position_in_page": None,
                                                    "position_percentage": None,
                                                    "section_heading": selected_chunk["topic_name"],
                                                    "content_snippet": selected_chunk["content"][:500],
                                                    "relevance_reason": f"Syllabus content used for regeneration of {selected_chunk['topic_name']}",
                                                }],
                                                "generation_reasoning": f"Question regenerated from '{selected_chunk['topic_name']}' syllabus content",
                                                "content_coverage": f"Rubric-based regeneration covering {selected_chunk['topic_name']} concepts",
                                            },
                                        },
                                    )
                                    db.add(new_question)
                                    await db.commit()
                                    await db.refresh(new_question)
                                    
                                    # Update old question to link to replacement
                                    question.replaced_by_id = new_question.id
                                    question.is_latest = False
                                    await db.commit()
                                    
                                    replacement = new_question
                                    logger.info(f"Rubric regen: replacement={replacement.id}")
                                    break  # Success - exit retry loop
                                else:
                                    logger.warning(f"Rubric regen attempt {attempt + 1}: question too short or duplicate")
                            else:
                                logger.warning(f"Rubric regen attempt {attempt + 1}: response missing question_text")
                        except Exception as e:
                            logger.error(f"Rubric regen attempt {attempt + 1} error: {e}")
                            await db.rollback()
                            if attempt == max_retries - 1:
                                logger.error(f"Rubric regen failed after {max_retries} attempts")

            if replacement:
                return QR.from_orm_with_sources(replacement)

        # ═══════════════════════════════════════════════
        # PATH D: Import — no automatic replacement
        # ═══════════════════════════════════════════════
        elif gen_method == "import":
            logger.info("Imported question rejected — no automatic replacement")

        # If no replacement was generated, restore the original question to pending
        if not replacement:
            logger.warning(f"Regeneration failed for question {question_id}, restoring to pending status")
            try:
                fresh_q = await db.execute(select(Question).where(Question.id == question_id))
                original_question = fresh_q.scalar_one_or_none()
                if original_question:
                    original_question.vetting_status = "pending"
                    original_question.vetted_at = None
                    original_question.vetting_notes = "Regeneration failed - restored to pending"
                    await db.commit()
                    logger.info(f"Restored question {question_id} to pending status")
            except Exception as restore_err:
                logger.error(f"Failed to restore question to pending: {restore_err}")
    
    # Re-fetch the question from DB to get a clean ORM object with all attributes loaded
    fresh = await db.execute(
        select(Question).where(Question.id == question_snapshot["id"])
    )
    fresh_question = fresh.scalar_one_or_none()
    if fresh_question:
        return QuestionResponse.from_orm_with_sources(fresh_question)
    # Fallback: build response from snapshot dict
    return QuestionResponse(**question_snapshot)


@router.put("/{question_id}/co-mapping", response_model=QuestionResponse)
async def update_co_mapping(
    question_id: str,
    mapping: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Update Course Outcome mapping for a question.
    """
    from sqlalchemy import or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
            )
        )
    )
    question = result.scalar_one_or_none()
    
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    question.course_outcome_mapping = mapping
    await db.commit()
    await db.refresh(question)
    
    return QuestionResponse.from_orm_with_sources(question)


@router.put("/{question_id}", response_model=QuestionResponse)
async def update_question(
    question_id: str,
    update_data: dict,  # Using dict to accept partial updates
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Update a question's editable fields including marks, subject, topic, etc.
    
    Updateable fields:
    - marks: int (1-100)
    - difficulty_level: easy/medium/hard
    - bloom_taxonomy_level: remember/understand/apply/analyze/evaluate/create
    - subject_id: UUID - link to a subject
    - topic_id: UUID - link to a specific chapter/topic
    - learning_outcome_id: string (e.g., LO1, LO2)
    - course_outcome_mapping: dict (e.g., {"CO1": 2, "CO3": 1})
    - question_text: string
    - correct_answer: string
    - options: list of strings (for MCQ)
    """
    from sqlalchemy import or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject, Topic
    
    # Find the question with ownership check
    result = await db.execute(
        select(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(
            Question.id == question_id,
            or_(
                Document.user_id == current_user.id,
                Subject.user_id == current_user.id,
                # Also allow if no document/subject (orphaned questions)
                Question.document_id.is_(None) & Question.subject_id.is_(None),
            )
        )
    )
    question = result.scalar_one_or_none()
    
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Question not found",
        )
    
    # Validate subject_id if provided
    if 'subject_id' in update_data and update_data['subject_id']:
        subject_result = await db.execute(
            select(Subject).where(
                Subject.id == update_data['subject_id'],
            )
        )
        if not subject_result.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Subject not found",
            )
    
    # Validate topic_id if provided
    if 'topic_id' in update_data and update_data['topic_id']:
        topic_result = await db.execute(
            select(Topic)
            .where(
                Topic.id == update_data['topic_id'],
            )
        )
        topic = topic_result.scalar_one_or_none()
        if not topic:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Topic not found",
            )
        # If topic is provided but subject isn't, auto-set the subject
        if 'subject_id' not in update_data or not update_data['subject_id']:
            update_data['subject_id'] = topic.subject_id
    
    # Allowed fields to update
    allowed_fields = {
        'marks', 'difficulty_level', 'bloom_taxonomy_level', 
        'subject_id', 'topic_id', 'learning_outcome_id',
        'course_outcome_mapping', 'question_text', 'correct_answer', 'options'
    }
    
    # Apply updates
    for field, value in update_data.items():
        if field in allowed_fields:
            setattr(question, field, value)
    
    await db.commit()
    await db.refresh(question)
    
    return QuestionResponse.from_orm_with_sources(question)


@router.get("/vetting/stats")
async def get_vetting_stats(
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get vetting/review statistics.
    """
    from sqlalchemy import or_
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    
    # Base query - support both document-based and subject-based questions
    base_where = [
        or_(
            Document.user_id == current_user.id,
            Subject.user_id == current_user.id,
        ),
        Question.is_archived == False,  # Exclude archived from stats
    ]
    if subject_id:
        base_where.append(Question.subject_id == subject_id)
    
    # Total generated
    total_result = await db.execute(
        select(func.count())
        .select_from(Question)
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(*base_where)
    )
    total_generated = total_result.scalar_one()
    
    # By status
    status_result = await db.execute(
        select(Question.vetting_status, func.count())
        .outerjoin(Document, Question.document_id == Document.id)
        .outerjoin(Subject, Question.subject_id == Subject.id)
        .where(*base_where)
        .group_by(Question.vetting_status)
    )
    by_status = {row[0]: row[1] for row in status_result.all()}
    
    total_approved = by_status.get("approved", 0)
    total_rejected = by_status.get("rejected", 0)
    pending_review = by_status.get("pending", 0)
    
    # Approval rate (out of reviewed questions)
    reviewed = total_approved + total_rejected
    approval_rate = (total_approved / reviewed * 100) if reviewed > 0 else 0.0
    
    return {
        "total_generated": total_generated,
        "total_approved": total_approved,
        "total_rejected": total_rejected,
        "pending_review": pending_review,
        "approval_rate": round(approval_rate, 1),
        "change_from_last_month": 5.2,  # Placeholder - would need historical data
    }


@router.get("/analytics/by-subject")
async def get_analytics_by_subject(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get question analytics grouped by subject.
    """
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    from sqlalchemy import case
    
    result = await db.execute(
        select(
            Subject.id,
            Subject.name,
            Subject.code,
            func.count(Question.id).label("total_questions"),
            func.sum(case((Question.vetting_status == "approved", 1), else_=0)).label("approved"),
            func.sum(case((Question.vetting_status == "rejected", 1), else_=0)).label("rejected"),
            func.sum(case((Question.vetting_status == "pending", 1), else_=0)).label("pending"),
        )
        .outerjoin(Question, (Question.subject_id == Subject.id) & (Question.is_archived == False))
        .where(Subject.user_id == current_user.id)
        .group_by(Subject.id, Subject.name, Subject.code)
    )
    
    subjects = []
    for row in result.all():
        subjects.append({
            "id": str(row.id),
            "name": row.name,
            "code": row.code,
            "total_questions": row.total_questions or 0,
            "approved": row.approved or 0,
            "rejected": row.rejected or 0,
            "pending": row.pending or 0,
        })
    
    return {"subjects": subjects}


@router.get("/analytics/by-learning-outcome")
async def get_analytics_by_lo(
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get question distribution by Learning Outcome.
    Covers all question sources: document-based, rubric/chapter/subject-generated.
    """
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    from sqlalchemy import or_

    # Subqueries for user ownership
    user_doc_ids = select(Document.id).where(Document.user_id == current_user.id).scalar_subquery()
    user_subject_ids = select(Subject.id).where(Subject.user_id == current_user.id).scalar_subquery()

    base_where = [
        or_(
            Question.document_id.in_(user_doc_ids),
            Question.subject_id.in_(user_subject_ids),
        ),
        Question.is_archived == False,
    ]
    if subject_id:
        base_where.append(Question.subject_id == subject_id)

    result = await db.execute(
        select(Question.learning_outcome_id, func.count())
        .where(*base_where)
        .group_by(Question.learning_outcome_id)
    )

    distribution = {}
    for row in result.all():
        lo_id = row[0] or "Unassigned"
        distribution[lo_id] = row[1]

    return {"learning_outcomes": distribution}


@router.get("/analytics/by-bloom")
async def get_analytics_by_bloom(
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get question distribution by Bloom's Taxonomy level.
    Covers all question sources: document-based, rubric/chapter/subject-generated.
    """
    from app.models.question import Question
    from app.models.document import Document
    from app.models.subject import Subject
    from sqlalchemy import or_

    # Subqueries for user ownership
    user_doc_ids = select(Document.id).where(Document.user_id == current_user.id).scalar_subquery()
    user_subject_ids = select(Subject.id).where(Subject.user_id == current_user.id).scalar_subquery()

    base_where = [
        or_(
            Question.document_id.in_(user_doc_ids),
            Question.subject_id.in_(user_subject_ids),
        ),
        Question.is_archived == False,
    ]
    if subject_id:
        base_where.append(Question.subject_id == subject_id)

    result = await db.execute(
        select(Question.bloom_taxonomy_level, func.count())
        .where(*base_where)
        .group_by(Question.bloom_taxonomy_level)
    )

    distribution = {}
    for row in result.all():
        level = row[0] or "unspecified"
        distribution[level] = row[1]

    return {"bloom_levels": distribution}


# ============== Rubric-Based Generation ==============

class RubricGenerationRequest(BaseModel):
    """Schema for generating questions from a rubric."""
    rubric_id: str
    topic_id: Optional[str] = None  # If set, restrict generation to this chapter only
    count_override: Optional[int] = None  # If set, stop after this many questions (for retry-failed)


@router.post("/generate-from-rubric")
async def generate_from_rubric(
    request: RubricGenerationRequest,
    raw_request: Request,
    current_user: User = Depends(rate_limit(requests=50, window_seconds=86400)),  # 50/day
    db: AsyncSession = Depends(get_db),
):
    """
    Generate questions based on a rubric configuration.
    Uses topics' syllabus content from the rubric's subject.
    
    Flow:
    1. Load rubric with question type & LO distribution
    2. Collect all topics with syllabus content
    3. Chunk syllabus content for better RAG
    4. Generate questions distributed across topics/LOs
    5. Validate & save questions with duplicate detection
    
    Returns Server-Sent Events (SSE) stream with progress updates.
    """
    ensure_user_can_generate(current_user)

    from app.models.rubric import Rubric
    from app.models.subject import Subject, Topic
    from app.services.llm_service import LLMService
    from app.services.embedding_service import EmbeddingService
    from app.models.question import Question
    from loguru import logger
    import random
    import math
    
    request_id = uuid.uuid4().hex[:8]
    
    # Get rubric
    result = await db.execute(
        select(Rubric).where(
            Rubric.id == request.rubric_id,
            Rubric.user_id == current_user.id,
        )
    )
    rubric = result.scalar_one_or_none()
    
    if not rubric:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rubric not found",
        )
    
    # Extract IDs early to avoid SQLAlchemy lazy loading issues
    rubric_id = rubric.id
    rubric_name = rubric.name
    subject_id = rubric.subject_id
    question_distribution = rubric.question_type_distribution or {}
    lo_distribution = rubric.learning_outcomes_distribution or {}
    total_questions = rubric.total_questions or 0
    
    # Get subject
    result = await db.execute(
        select(Subject).where(Subject.id == subject_id)
    )
    subject = result.scalar_one_or_none()
    
    if not subject:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Subject not found",
        )
    
    subject_id_str = str(subject.id)
    subject_name = subject.name
    
    # Get topics with syllabus content
    topic_query = select(Topic).where(
        Topic.subject_id == subject.id,
        Topic.has_syllabus == True,
    )
    # If a specific topic_id was requested, restrict to that chapter only
    if request.topic_id:
        topic_query = topic_query.where(Topic.id == request.topic_id)
    topic_query = topic_query.order_by(Topic.order_index)
    result = await db.execute(topic_query)
    topics = result.scalars().all()
    
    if not topics:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No topics with syllabus content found. Please add content to topics first.",
        )
    
    # Extract topic data early
    topic_data = []
    for t in topics:
        topic_data.append({
            "id": str(t.id),
            "name": t.name,
            "content": t.syllabus_content or "",
            "lo_mappings": t.learning_outcome_mappings or {},
        })
    
    llm_service = LLMService()
    embedding_service = EmbeddingService()
    
    # Pre-fetch subject attributes that might not be loaded in the new session
    subject_lo_mappings = getattr(subject, "learning_outcome_mappings", {}) or {}
    subject_cos = (subject.course_outcomes or {}).get("outcomes") if subject.course_outcomes else None
    # Build LO description map so the prompt can name what each LO actually means
    subject_lo_descriptions: dict = {}
    _los_raw = (subject.learning_outcomes or {})
    if isinstance(_los_raw, dict):
        for _item in _los_raw.get("outcomes", []):
            if isinstance(_item, dict):
                _lo_id = _item.get("id") or _item.get("code")
                _lo_desc = _item.get("description") or _item.get("name") or ""
                if _lo_id:
                    subject_lo_descriptions[_lo_id] = _lo_desc

    logger.info(f"[{request_id}] Rubric generation starting: {rubric_name}, {total_questions} questions, topic_id={request.topic_id}, topics_found={len(topic_data)}")
    for td in topic_data:
        logger.info(f"[{request_id}]   -> Topic: {td['name']} ({td['id']})")
    
    async def event_generator():
        """Generate SSE events with questions."""
        from app.core.database import AsyncSessionLocal
        db = AsyncSessionLocal()
        try:
            yield f"data: {json.dumps({'status': 'processing', 'progress': 5, 'message': 'Analyzing syllabus content...'})}\n\n"
            
            # Prepare chunked content for all topics
            all_chunks = []
            for topic in topic_data:
                if topic["content"]:
                    # Simple chunking by paragraphs
                    content = topic["content"]
                    paragraphs = re.split(r'\n\n+', content)
                    chunk_size = 1500
                    current_chunk = ""
                    
                    for para in paragraphs:
                        if len(current_chunk) + len(para) < chunk_size:
                            current_chunk += "\n\n" + para if current_chunk else para
                        else:
                            if current_chunk:
                                all_chunks.append({
                                    "topic_id": topic["id"],
                                    "topic_name": topic["name"],
                                    "content": current_chunk.strip(),
                                    "lo_mappings": topic["lo_mappings"],
                                })
                            current_chunk = para
                    
                    if current_chunk:
                        all_chunks.append({
                            "topic_id": topic["id"],
                            "topic_name": topic["name"],
                            "content": current_chunk.strip(),
                            "lo_mappings": topic["lo_mappings"],
                        })
            
            if not all_chunks:
                yield f"data: {json.dumps({'status': 'error', 'progress': 0, 'message': 'No syllabus content found in topics'})}\n\n"
                return

            # ── RAG: fetch per-topic reference-book chunks using hybrid search ──
            # Build a map of topic_id -> reference_context so each question gets
            # reference material relevant to its specific topic (not a single blob
            # that may be dominated by one topic).
            topic_reference_contexts: dict[str, str] = {}  # topic_id -> reference text
            topic_reference_chunks: dict[str, list] = {}    # topic_id -> list of DocumentChunk objects (for source_info)
            ref_count = 0
            ref_doc_ids = []
            try:
                from app.services.document_service import DocumentService
                doc_service = DocumentService(db, embedding_service)
                
                # Get reference book IDs once
                ref_doc_ids = await doc_service.get_reference_document_ids(
                    user_id=current_user.id,
                    subject_id=str(subject_id_str),
                    index_type="reference_book",
                )
                
                if ref_doc_ids:
                    for td in topic_data:
                        if not td["content"]:
                            continue
                        topic_rag_query = f"{td['name']}: {td['content'][:400]}"
                        topic_query_emb = await embedding_service.get_query_embedding(topic_rag_query)
                        
                        topic_ref_chunks = await doc_service.hybrid_search_multi_document(
                            document_ids=ref_doc_ids,
                            query=topic_rag_query,
                            query_embedding=topic_query_emb,
                            top_k=3,
                            alpha=0.6,
                        )
                        
                        if topic_ref_chunks:
                            topic_reference_contexts[td["id"]] = "\n\n".join(
                                f"[Reference {i+1}]: {c.chunk_text}" for i, c in enumerate(topic_ref_chunks)
                            )
                            topic_reference_chunks[td["id"]] = topic_ref_chunks
                            ref_count += len(topic_ref_chunks)
                    
                    if topic_reference_contexts:
                        logger.info(f"[{request_id}] RAG found {ref_count} reference chunks across {len(topic_reference_contexts)} topics for rubric generation")
                
                if not topic_reference_contexts and ref_doc_ids:
                    # Fallback: single combined query if per-topic failed
                    rag_query_parts = [td["name"] + ": " + (td["content"] or "")[:200] for td in topic_data if td["content"]]
                    rag_query = " ".join(rag_query_parts)[:1000]
                    query_emb = await embedding_service.get_query_embedding(rag_query)
                    fallback_chunks = await doc_service.get_reference_chunks(
                        user_id=current_user.id,
                        subject_id=str(subject_id_str),
                        index_type="reference_book",
                        query_embedding=query_emb,
                        top_k=5,
                    )
                    if fallback_chunks:
                        fallback_ctx = "\n\n".join(
                            f"[Reference {i+1}]: {c.chunk_text}" for i, c in enumerate(fallback_chunks)
                        )
                        for td in topic_data:
                            topic_reference_contexts[td["id"]] = fallback_ctx
                            topic_reference_chunks[td["id"]] = fallback_chunks
                        ref_count = len(fallback_chunks)
                        logger.info(f"[{request_id}] RAG fallback: {ref_count} reference chunks shared across all topics")
            except Exception as e:
                logger.warning(f"[{request_id}] RAG lookup failed ({e}); continuing without reference context")

            # Build CO context string for prompts
            co_context = ""
            if subject_cos and isinstance(subject_cos, list):
                co_ids = [co.get('id', f'CO{i+1}') if isinstance(co, dict) else str(co) for i, co in enumerate(subject_cos)]
                co_context = f"\nAvailable Course Outcomes (pick 1–2 most relevant): {', '.join(co_ids)}"
            elif subject_lo_mappings:
                all_cos = set()
                for lo_key, co_map in subject_lo_mappings.items():
                    if isinstance(co_map, dict):
                        all_cos.update(co_map.keys())
                if all_cos:
                    co_context = f"\nAvailable Course Outcomes (pick 1–2 most relevant): {', '.join(sorted(all_cos))}"

            # Pre-compute per-slot LO assignments from the rubric distribution
            all_lo_slots = _compute_lo_question_slots(lo_distribution, total_questions)
            lo_slot_index = 0  # global counter across all question types

            yield f"data: {json.dumps({'status': 'processing', 'progress': 10, 'message': f'Prepared {len(all_chunks)} content sections from {len(topic_data)} topics + {ref_count} reference excerpts'})}\n\n"
            
            # Create a GenerationSession so this shows up in history
            rubric_session_id = uuid.uuid4()
            rubric_gen_session = GenerationSession(
                id=rubric_session_id,
                user_id=current_user.id,
                subject_id=str(subject_id_str),
                topic_id=str(str(request.topic_id)) if request.topic_id else None,
                generation_method="rubric",
                requested_count=total_questions,
                status="in_progress",
                started_at=datetime.now(timezone.utc),
                generation_config={
                    "rubric_id": str(rubric_id),
                    "rubric_name": rubric_name,
                    "question_distribution": question_distribution,
                },
            )
            db.add(rubric_gen_session)
            await db.commit()
            await db.refresh(rubric_gen_session)
            
            questions_generated = 0
            questions_failed = 0
            generated_questions = []  # Track for duplicate detection
            generated_embeddings = []  # Track embeddings for semantic dedupe within this session
            used_starters_rubric: list[str] = []  # Track starters for variety

            # Preload existing question embeddings for subject-level dedupe
            existing_question_embeddings = await _preload_subject_embeddings(db, str(subject_id_str))

            # Type mapping
            type_mapping = {
                "mcq": "mcq",
                "short_notes": "short_answer",
                "short_answer": "short_answer",
                "essay": "long_answer",
                "long_answer": "long_answer",
            }

            RETRY_LIMIT = 5

            cancelled = False

            # ── Topic setup (shared globally across ALL question types) ──
            # This ensures topic spread is even across the whole exam, not per-type.
            topic_ids_ordered = list(dict.fromkeys(c["topic_id"] for c in all_chunks))
            chunks_by_topic: dict[str, list] = {tid: [] for tid in topic_ids_ordered}
            for c in all_chunks:
                chunks_by_topic[c["topic_id"]].append(c)
            # Rotating chunk cursor per topic (advances as questions are taken from that topic)
            topic_chunk_cursors: dict[str, int] = {tid: 0 for tid in topic_ids_ordered}
            # Count how many questions have been assigned to each topic (for LO-aware balancing)
            topic_question_counts: dict[str, int] = {tid: 0 for tid in topic_ids_ordered}

            # If count_override is set (retry-failed), cap total generation
            effective_total = request.count_override if request.count_override else total_questions

            # Generate questions for each type
            for q_type, config in question_distribution.items():
                if cancelled:
                    break
                if request.count_override and questions_generated >= request.count_override:
                    break
                count = config.get("count", 0)
                marks = config.get("marks_each", 2)
                mapped_type = type_mapping.get(q_type, "mcq")
                
                if count == 0:
                    continue
                
                logger.info(f"[{request_id}] Generating {count} {mapped_type} questions")
                
                for i in range(count):
                    if request.count_override and questions_generated >= request.count_override:
                        break
                    if await raw_request.is_disconnected():
                        logger.info(f"[{request_id}] Client disconnected, cancelling generation")
                        cancelled = True
                        break

                    try:
                        # ── LO-aware topic selection ──
                        # Consume the next pre-assigned LO slot from the global list
                        required_lo = all_lo_slots[lo_slot_index] if lo_slot_index < len(all_lo_slots) else None
                        lo_slot_index += 1

                        # Pick a topic that has content mapped to this LO (least-used first)
                        selected_topic_id = _select_lo_aware_topic(
                            required_lo, topic_ids_ordered, chunks_by_topic, topic_question_counts
                        )
                        topic_question_counts[selected_topic_id] = topic_question_counts.get(selected_topic_id, 0) + 1

                        topic_chunks = chunks_by_topic[selected_topic_id]
                        cursor = topic_chunk_cursors[selected_topic_id]
                        selected_chunk = topic_chunks[cursor % len(topic_chunks)]
                        topic_chunk_cursors[selected_topic_id] = cursor + 1
                        
                        # Pick a unique starter for this question
                        starter = _pick_starter(mapped_type, used_starters_rubric)

                        # Build LO-specific requirement for the prompt
                        lo_requirement = ""
                        if required_lo:
                            lo_desc = subject_lo_descriptions.get(required_lo, "")
                            lo_requirement = (
                                f"\n- REQUIRED Learning Outcome: This question MUST specifically assess "
                                f"{required_lo}"
                            )
                            if lo_desc:
                                lo_requirement += f' — "{lo_desc}"'
                            lo_requirement += (
                                "\n- Design the question to test the competency described by this LO, "
                                "not just mention the topic"
                            )

                        # Build enhanced prompt
                        system_prompt = _get_rubric_system_prompt(mapped_type)
                        
                        prompt = f"""Context from "{selected_chunk['topic_name']}":
{selected_chunk['content']}
"""
                        # Use topic-specific reference context (not a single blob)
                        chunk_ref_context = topic_reference_contexts.get(selected_chunk["topic_id"], "")
                        if chunk_ref_context:
                            prompt += f"""
Background knowledge (use to inform the question, do NOT cite or reference it):
{chunk_ref_context}
"""
                        prompt += f"""
Generate a {mapped_type.replace('_', ' ')} question based on this content.
- Marks: {marks}
- REQUIRED: Your question MUST start with the word/phrase "{starter}" — this is non-negotiable
- FORBIDDEN: Do NOT start all the questions with "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", "An engineer", or similar scenario setups. Also start directly with the required starter word.
- The question text must be fully self-contained — do NOT mention "the text", "the reference", "the provided material", "reference [N]", or any source
- Avoid questions that are too similar to: {', '.join([q[:50] for q in generated_questions[-5:]]) if generated_questions else 'None yet'}{lo_requirement}
{co_context}
- For "course_outcome_mapping" in your JSON: assign 1 CO for most questions, at most 2 COs for genuinely cross-cutting questions. Do NOT map to 3 or more COs. Use Bloom level as value (1=remember/understand, 2=apply/analyze, 3=evaluate/create).

Output valid JSON only."""

                        accepted = False
                        response_obj = None

                        for attempt in range(RETRY_LIMIT):
                            try:
                                resp = await llm_service.generate_json(
                                    prompt=prompt,
                                    system_prompt=system_prompt,
                                    temperature=0.75 + (0.05 * attempt),
                                )
                                resp = _normalize_llm_response(resp)
                            except Exception as e:
                                logger.warning(f"[{request_id}] LLM call exception (attempt {attempt+1}): {e}")
                                resp = None

                            if not resp or "question_text" not in resp:
                                resp_keys = list(resp.keys()) if resp else []
                                resp_preview = str(resp)[:300] if resp else "None"
                                logger.warning(f"[{request_id}] LLM returned invalid response (attempt {attempt+1}), keys={resp_keys}, preview={resp_preview}")
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            candidate_text = resp["question_text"]
                            if len(candidate_text) < 15:
                                logger.debug(f"[{request_id}] Generated text too short (len={len(candidate_text)})")
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            # Compute embedding for candidate
                            try:
                                candidate_emb = await embedding_service.get_embedding(candidate_text)
                            except Exception:
                                logger.exception(f"[{request_id}] Embedding generation failed on attempt {attempt+1}")
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            # Validate MCQ has at least 4 options
                            if mapped_type == "mcq":
                                raw_opts = resp.get("options")
                                normalized_opts = _normalize_mcq_options(raw_opts, "mcq")
                                if not normalized_opts or len(normalized_opts) < 4:
                                    logger.debug(f"[{request_id}] MCQ has only {len(normalized_opts) if normalized_opts else 0} options (attempt {attempt+1}), retrying")
                                    if attempt == RETRY_LIMIT - 1:
                                        questions_failed += 1
                                    continue

                            # Dedupe: reject if too similar to existing or session-generated questions
                            if _is_duplicate_embedding(candidate_emb, existing_question_embeddings, generated_embeddings):
                                logger.debug(f"[{request_id}] Rubric candidate rejected as duplicate (attempt {attempt+1})")
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            # Accepted candidate
                            question_text = candidate_text
                            question_embedding = candidate_emb
                            response_obj = resp
                            accepted = True
                            break

                        if not accepted:
                            # All attempts failed for this slot
                            continue

                        # --- Determine Learning Outcome (LO) and Course Outcome (CO) mapping ---
                        assigned_lo = None
                        assigned_co_map = {}

                        # 1) Use the pre-assigned LO slot (from rubric distribution) as the primary source.
                        #    Accept a matching LO from the LLM only if it agrees.
                        if required_lo:
                            # LLM may return a more specific LO; only override if it reports the same one
                            llm_lo = (response_obj or {}).get("learning_outcome_id") or (response_obj or {}).get("learning_outcome")
                            assigned_lo = llm_lo if llm_lo == required_lo else required_lo
                        else:
                            # No pre-assigned slot: trust whatever the LLM returned
                            if response_obj and isinstance(response_obj.get("learning_outcome_id"), str):
                                assigned_lo = response_obj.get("learning_outcome_id")
                            elif response_obj and isinstance(response_obj.get("learning_outcome"), str):
                                assigned_lo = response_obj.get("learning_outcome")
                            # Fallback to topic mappings
                            topic_lo_map = selected_chunk.get("lo_mappings") or {}
                            if not assigned_lo and topic_lo_map:
                                try:
                                    assigned_lo = max(topic_lo_map.items(), key=lambda kv: (kv[1] or 0))[0]
                                except Exception:
                                    assigned_lo = None

                        # 2) CO mapping — prefer LLM's multi-CO dict (course_outcome_mapping)
                        if response_obj and isinstance(response_obj.get("course_outcome_mapping"), dict) and response_obj["course_outcome_mapping"]:
                            assigned_co_map = response_obj["course_outcome_mapping"]
                        elif response_obj and isinstance(response_obj.get("course_outcome"), dict):
                            assigned_co_map = response_obj["course_outcome"]
                        elif response_obj and isinstance(response_obj.get("course_outcome"), str) and response_obj["course_outcome"]:
                            assigned_co_map = {response_obj["course_outcome"]: 1}
                        # 3) Build CO map from subject LO→CO mappings when LLM didn't supply one
                        if not assigned_co_map and assigned_lo and subject_lo_mappings and isinstance(subject_lo_mappings.get(assigned_lo), dict):
                            assigned_co_map = subject_lo_mappings.get(assigned_lo)
                        if not assigned_co_map and assigned_lo and subject_cos and isinstance(subject_cos, list) and len(subject_cos) > 0:
                            m = re.search(r"LO(\d+)", str(assigned_lo))
                            if m:
                                idx = (int(m.group(1)) - 1) % len(subject_cos)
                                co_id = subject_cos[idx].get('id', f'CO{idx+1}') if isinstance(subject_cos[idx], dict) else f'CO{idx+1}'
                                assigned_co_map = {co_id: 1}

                        # Sanitize fields to fit DB column limits
                        _s = _sanitize_question_fields({
                            "learning_outcome_id": assigned_lo,
                            "course_outcome_mapping": assigned_co_map,
                            "bloom_taxonomy_level": (response_obj.get("bloom_level") if response_obj else "understand"),
                            "difficulty_level": "medium",
                        })
                        assigned_lo = _s["learning_outcome_id"]
                        assigned_co_map = _s["course_outcome_mapping"]

                        # Create question (accepted candidate) with LO/CO and metadata
                        question = Question(
                            session_id=rubric_session_id,
                            subject_id=str(subject_id_str),
                            topic_id=str(selected_chunk["topic_id"]),
                            question_text=question_text,
                            question_embedding=question_embedding,
                            question_type=mapped_type,
                            marks=marks,
                            difficulty_level=_s["difficulty_level"],
                            bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                            correct_answer=(response_obj.get("correct_answer") or (response_obj.get("expected_answer") if response_obj else None)),
                            options=_normalize_mcq_options(response_obj.get("options"), mapped_type) if response_obj else None,
                            explanation=(response_obj.get("explanation") if response_obj else None),
                            topic_tags=(response_obj.get("topic_tags", [selected_chunk["topic_name"]]) if response_obj else [selected_chunk["topic_name"]]),
                            learning_outcome_id=assigned_lo,
                            course_outcome_mapping=assigned_co_map,
                            generation_confidence=0.75,
                            vetting_status="pending",
                            generation_metadata={
                                "rubric_id": str(rubric_id),
                                "topic_name": selected_chunk["topic_name"],
                                "lo_mappings": selected_chunk["lo_mappings"],
                                "assigned_learning_outcome": assigned_lo,
                                "assigned_course_outcome_mapping": assigned_co_map,
                                "source_info": _build_rubric_source_info(
                                    topic_reference_chunks.get(selected_chunk["topic_id"], []),
                                    selected_chunk["topic_name"],
                                ),
                            },
                        )
                        # Use an isolated session per question so a DB error on one
                        # question does not corrupt the main session's transaction state.
                        async with AsyncSessionLocal() as q_session:
                            q_session.add(question)
                            await q_session.commit()
                            await q_session.refresh(question)

                        # Track for duplicate detection
                        generated_questions.append(question_text)
                        generated_embeddings.append(question_embedding)
                        questions_generated += 1

                        # Calculate progress
                        progress = 10 + int((questions_generated / max(effective_total, 1)) * 85)

                        yield f"data: {json.dumps({'status': 'generating', 'progress': min(progress, 95), 'current_question': questions_generated, 'total_questions': effective_total, 'message': f'Generated {questions_generated}/{effective_total} questions', 'question': {'id': str(question.id), 'question_text': question.question_text[:100] + '...' if len(question.question_text) > 100 else question.question_text, 'question_type': question.question_type, 'marks': question.marks, 'learning_outcome_id': question.learning_outcome_id, 'course_outcome_mapping': question.course_outcome_mapping}})}\n\n"
                    except Exception as e:
                        logger.error(f"[{request_id}] Error generating question: {e}")
                        questions_failed += 1
                        continue
            
            # Finalize rubric session
            try:
                rubric_gen_session.questions_generated = questions_generated
                rubric_gen_session.questions_failed = max(0, total_questions - questions_generated)
                rubric_gen_session.status = "completed"
                rubric_gen_session.completed_at = datetime.now(timezone.utc)
                await db.commit()
            except Exception as e:
                logger.error(f"[{request_id}] Error finalizing rubric session: {e}")

            # Update subject stats
            try:
                result = await db.execute(
                    select(Subject).where(Subject.id == str(subject_id_str))
                )
                subj = result.scalar_one_or_none()
                if subj:
                    subj.total_questions = (subj.total_questions or 0) + questions_generated
                    await db.commit()
            except Exception as e:
                logger.error(f"[{request_id}] Error updating subject stats: {e}")
            
            questions_failed_count = max(0, total_questions - questions_generated)
            yield f"data: {json.dumps({'status': 'complete', 'progress': 100, 'current_question': questions_generated, 'total_questions': total_questions, 'questions_failed': questions_failed_count, 'message': f'Generated {questions_generated} questions' + (f' ({questions_failed_count} shortfall)' if questions_failed_count > 0 else '')})}\n\n"
            
            logger.info(f"[{request_id}] Rubric generation complete: {questions_generated} generated, {questions_failed_count} shortfall")
            
        except Exception as e:
            logger.error(f"[{request_id}] Rubric generation failed: {e}")
            yield f"data: {json.dumps({'status': 'error', 'progress': 0, 'message': f'Generation failed: {str(e)}'})}\n\n"
        finally:
            try:
                await db.rollback()
            except Exception:
                pass
            try:
                await db.close()
            except Exception:
                pass
    return StreamingResponse(
        sse_with_heartbeat(event_generator()),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ============== Per-Chapter Generation with RAG ==============

class QuestionTypeSpec(BaseModel):
    """Specification for a single question type."""
    count: int
    marks_each: int = 2

class ChapterGenerationRequest(BaseModel):
    """Schema for generating questions from a single chapter."""
    topic_id: str
    question_types: dict  # e.g. {"mcq": {"count": 5, "marks_each": 2}, ...}
    difficulty: Optional[str] = "medium"  # easy | medium | hard
    lo_filter: Optional[List[str]] = None  # restrict to these LO ids, e.g. ["LO1", "LO2"]
    lo_distribution: Optional[dict] = None  # {"LO1": 25, "LO2": 25, ...} percentage weights
    count_override: Optional[int] = None  # If set, stop after this many questions (for retry-failed)


@router.post("/generate-chapter")
async def generate_chapter(
    request: ChapterGenerationRequest,
    raw_request: Request,
    current_user: User = Depends(rate_limit(requests=200, window_seconds=86400)),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate questions strictly from a single chapter.
    Uses the chapter's syllabus content + RAG from reference books.
    Streams SSE progress updates.
    """
    ensure_user_can_generate(current_user)

    from app.models.subject import Subject, Topic
    from app.services.llm_service import LLMService
    from app.services.embedding_service import EmbeddingService
    from app.services.document_service import DocumentService
    from app.models.question import Question
    from app.core.database import AsyncSessionLocal
    from loguru import logger
    import random

    request_id = uuid.uuid4().hex[:8]

    # ── Load the topic ──
    result = await db.execute(
        select(Topic).where(Topic.id == request.topic_id)
    )
    topic = result.scalar_one_or_none()
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")

    if not topic.has_syllabus or not topic.syllabus_content:
        raise HTTPException(status_code=400, detail="This chapter has no syllabus content. Upload syllabus first.")

    topic_id_str = str(topic.id)
    topic_name = topic.name
    topic_content = topic.syllabus_content
    topic_lo_mappings = topic.learning_outcome_mappings or {}
    subject_id = topic.subject_id

    # ── Load the subject ──
    result = await db.execute(select(Subject).where(Subject.id == subject_id))
    subject = result.scalar_one_or_none()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")

    subject_id_str = str(subject.id)
    subject_name = subject.name
    subject_lo_mappings_data = getattr(subject, "learning_outcome_mappings", {}) or {}
    subject_cos = (subject.course_outcomes or {}).get("outcomes") if subject.course_outcomes else None
    user_id = current_user.id

    # Build LO description map from subject.learning_outcomes
    subject_lo_descriptions_ch: dict = {}
    _los_raw_ch = (subject.learning_outcomes or {})
    if isinstance(_los_raw_ch, dict):
        for _item in _los_raw_ch.get("outcomes", []):
            if isinstance(_item, dict):
                _lo_id = _item.get("id") or _item.get("code")
                _lo_desc = _item.get("description") or _item.get("name") or ""
                if _lo_id:
                    subject_lo_descriptions_ch[_lo_id] = _lo_desc

    # ── Parse question types ──
    type_mapping = {
        "mcq": "mcq", "short_notes": "short_answer",
        "short_answer": "short_answer", "essay": "long_answer",
        "long_answer": "long_answer",
    }
    generation_plan = []
    total_questions = 0
    for q_type, spec in request.question_types.items():
        count = spec.get("count", 0) if isinstance(spec, dict) else 0
        marks = spec.get("marks_each", 2) if isinstance(spec, dict) else 2
        mapped = type_mapping.get(q_type, q_type)
        if count > 0:
            generation_plan.append({"type": mapped, "count": count, "marks": marks})
            total_questions += count

    if total_questions == 0:
        raise HTTPException(status_code=400, detail="Add at least one question to generate.")

    llm_service = LLMService()
    embedding_service = EmbeddingService()

    # Pre-compute LO slots from distribution (if provided)
    chapter_lo_slots = _compute_lo_question_slots(
        request.lo_distribution or {}, total_questions
    )
    chapter_lo_slot_index = 0  # consumed inside event_generator

    logger.info(f"[{request_id}] Chapter gen: topic={topic_name} ({topic_id_str}), {total_questions} questions, lo_distribution={request.lo_distribution}")

    # ── Chunk the syllabus ──
    # Multi-level splitting: try double newlines first, then single newlines,
    # then sentence boundaries. Also break long single blocks into ~1000-char chunks.
    CHUNK_TARGET = 1000
    CHUNK_MAX = 1800

    def _smart_chunk(text: str) -> list:
        """Split text into chunks of ~CHUNK_TARGET chars, using natural boundaries."""
        # Try double-newline paragraphs first
        parts = re.split(r'\n\n+', text)
        # If only 1 big block, try single newlines
        if len(parts) <= 1:
            parts = re.split(r'\n', text)
        # If still only 1, try sentence boundaries
        if len(parts) <= 1:
            parts = re.split(r'(?<=[.!?])\s+', text)

        chunks = []
        current = ""
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if len(current) + len(part) < CHUNK_TARGET:
                current += "\n\n" + part if current else part
            else:
                if current:
                    chunks.append(current.strip())
                # If single part exceeds max, split by sentences
                if len(part) > CHUNK_MAX:
                    sentences = re.split(r'(?<=[.!?])\s+', part)
                    sub = ""
                    for s in sentences:
                        if len(sub) + len(s) < CHUNK_TARGET:
                            sub += " " + s if sub else s
                        else:
                            if sub:
                                chunks.append(sub.strip())
                            sub = s
                    current = sub
                else:
                    current = part
        if current and current.strip():
            chunks.append(current.strip())
        return chunks if chunks else [text]

    syllabus_chunks = _smart_chunk(topic_content)

    async def event_generator():
        db = AsyncSessionLocal()
        try:
            yield f"data: {json.dumps({'status': 'processing', 'progress': 5, 'message': 'Analyzing chapter content...'})}\n\n"

            # ── RAG: build a varied reference chunk pool in parallel ──────
            # Build multiple query variants covering different angles of the
            # topic, run all hybrid searches concurrently, deduplicate, then
            # rerank once.  Per-question we slice a different window of the
            # pool so each question gets contextually distinct reference text.
            ref_chunk_pool: list = []
            ref_doc_ids: list = []
            try:
                import asyncio as _asyncio
                doc_service = DocumentService(db, embedding_service)
                ref_doc_ids = await doc_service.get_reference_document_ids(
                    user_id=user_id,
                    subject_id=str(subject_id_str),
                    index_type="reference_book",
                )

                if ref_doc_ids:
                    # Extract section headings from syllabus for richer queries
                    heading_lines = [
                        ln.strip() for ln in topic_content.splitlines()
                        if ln.strip() and (ln.strip().isupper() or re.match(r'^(\d+[\.\)]|#+)\s', ln.strip()))
                    ][:3]

                    rag_query_variants = [
                        f"{topic_name}: {topic_content[:400]}",
                        topic_name,
                        f"examples of {topic_name}",
                        f"compare and contrast {topic_name}",
                        f"applications of {topic_name}",
                    ]
                    for h in heading_lines:
                        rag_query_variants.append(f"{topic_name} {h}")

                    async def _rag_search(q_variant: str):
                        try:
                            q_emb = await embedding_service.get_query_embedding(q_variant)
                            return await doc_service.hybrid_search_multi_document(
                                document_ids=ref_doc_ids,
                                query=q_variant,
                                query_embedding=q_emb,
                                top_k=6,
                                alpha=0.6,
                            )
                        except Exception as _re:
                            logger.warning(f"[{request_id}] RAG variant search failed: {_re}")
                            return []

                    all_ref_results = await _asyncio.gather(
                        *[_rag_search(v) for v in rag_query_variants[:6]]
                    )

                    seen_ref_ids: set = set()
                    for result_list in all_ref_results:
                        for c in result_list:
                            if c.id not in seen_ref_ids:
                                seen_ref_ids.add(c.id)
                                ref_chunk_pool.append(c)

                if not ref_chunk_pool and ref_doc_ids:
                    # Fallback: vector-only with generic topic query
                    fallback_emb = await embedding_service.get_query_embedding(topic_name)
                    ref_chunk_pool = await doc_service.get_reference_chunks(
                        user_id=user_id,
                        subject_id=str(subject_id_str),
                        index_type="reference_book",
                        query_embedding=fallback_emb,
                        top_k=10,
                    )

                logger.info(f"[{request_id}] RAG pool: {len(ref_chunk_pool)} unique chunks from {len(ref_doc_ids)} reference doc(s)")
            except Exception as e:
                logger.warning(f"[{request_id}] RAG pool build failed ({e}); continuing without reference context")

            ref_count = len(ref_chunk_pool)

            # Resolve a stable document_id for all generated questions in this chapter session.
            # Some deployments enforce questions.document_id as NOT NULL at the DB level.
            chapter_question_document_id = None
            try:
                if ref_chunk_pool:
                    chapter_question_document_id = next(
                        (str(getattr(c, "document_id", "")) for c in ref_chunk_pool if getattr(c, "document_id", None)),
                        None,
                    )
                if not chapter_question_document_id and ref_doc_ids:
                    chapter_question_document_id = str(ref_doc_ids[0])
                if not chapter_question_document_id:
                    fallback_doc_res = await db.execute(
                        select(Document.id)
                        .where(
                            Document.user_id == user_id,
                            Document.subject_id == str(subject_id_str),
                            Document.processing_status == "completed",
                        )
                        .order_by(Document.upload_timestamp.desc())
                        .limit(1)
                    )
                    fallback_doc_id = fallback_doc_res.scalar_one_or_none()
                    if fallback_doc_id:
                        chapter_question_document_id = str(fallback_doc_id)
            except Exception as e:
                logger.warning(f"[{request_id}] Failed to resolve chapter document_id: {e}")

            if not chapter_question_document_id:
                yield f"data: {json.dumps({'status': 'error', 'progress': 0, 'message': 'No processed document found for this subject. Upload at least one processed document before chapter generation.'})}\n\n"
                return

            # Helper: build reference_context string for a specific pool slice
            def _ref_context_for_index(q_index: int, window: int = 4) -> str:
                if not ref_chunk_pool:
                    return ""
                pool_size = len(ref_chunk_pool)
                offset = (q_index * window) % max(1, pool_size - window + 1)
                sliced = ref_chunk_pool[offset:offset + window]
                if not sliced:
                    sliced = ref_chunk_pool[:window]
                return "\n\n".join(
                    f"[Reference {i+1}]: {c.chunk_text}" for i, c in enumerate(sliced)
                )

            # Helper: build source_info for a specific pool slice
            def build_source_info_for_index(q_index: int, window: int = 4) -> dict:
                if not ref_chunk_pool:
                    return {"sources": [], "generation_reasoning": None, "content_coverage": None}
                pool_size = len(ref_chunk_pool)
                offset = (q_index * window) % max(1, pool_size - window + 1)
                sliced = ref_chunk_pool[offset:offset + window]
                if not sliced:
                    sliced = ref_chunk_pool[:window]

                sources = []
                for chunk in sliced:
                    chunk_meta = chunk.chunk_metadata or {}
                    source_meta = chunk_meta.get("source_info", {})
                    document_name = source_meta.get("document_name")
                    if not document_name and hasattr(chunk, 'document') and chunk.document:
                        document_name = chunk.document.filename
                    if not document_name:
                        document_name = "Unknown Document"
                    content_snippet = chunk.chunk_text
                    sources.append({
                        "document_name": document_name,
                        "page_number": chunk.page_number or source_meta.get("page_number"),
                        "page_range": source_meta.get("page_range"),
                        "position_in_page": source_meta.get("position_in_page"),
                        "position_percentage": source_meta.get("position_percentage"),
                        "section_heading": chunk.section_heading,
                        "content_snippet": content_snippet,
                        "highlighted_phrase": _extract_highlighted_phrase(content_snippet, topic_name),
                        "relevance_reason": f"Retrieved as reference material for {topic_name}",
                    })
                return {
                    "sources": sources,
                    "generation_reasoning": f"Question generated from chapter '{topic_name}' syllabus content with reference material support",
                    "content_coverage": f"Chapter-based question covering {topic_name} concepts",
                }


            yield f"data: {json.dumps({'status': 'processing', 'progress': 10, 'message': f'Prepared {len(syllabus_chunks)} content sections + {ref_count} reference excerpts'})}\n\n"

            # ── Create a GenerationSession for chapter generation ──
            chapter_session_id = uuid.uuid4()
            chapter_gen_session = GenerationSession(
                id=chapter_session_id,
                user_id=user_id,
                subject_id=str(subject_id_str),
                topic_id=str(topic_id_str),
                generation_method="chapter",
                requested_count=total_questions,
                status="in_progress",
                started_at=datetime.now(timezone.utc),
                generation_config={
                    "topic_name": topic_name,
                    "question_types": request.question_types,
                },
            )
            db.add(chapter_gen_session)
            await db.commit()
            await db.refresh(chapter_gen_session)

            # ── Preload existing embeddings ──
            existing_embeddings = await _preload_subject_embeddings(db, str(subject_id_str))

            questions_generated = 0
            questions_failed = 0
            generated_texts = []
            used_starters_chapter: list[str] = []  # Track starters for variety
            generated_embeddings = []
            lo_slot_idx_ch = chapter_lo_slot_index  # local counter for this session

            RETRY_LIMIT = 5

            # Build CO context string once (reused per question)
            co_context_ch = ""
            if subject_cos and isinstance(subject_cos, list):
                co_ids_ch = [co.get('id', f'CO{i+1}') if isinstance(co, dict) else str(co) for i, co in enumerate(subject_cos)]
                co_context_ch = f"\nAvailable Course Outcomes (pick 1–2 most relevant): {', '.join(co_ids_ch)}"
            elif subject_lo_mappings_data:
                all_cos_ch = set()
                for lo_key, co_map in subject_lo_mappings_data.items():
                    if isinstance(co_map, dict):
                        all_cos_ch.update(co_map.keys())
                if all_cos_ch:
                    co_context_ch = f"\nAvailable Course Outcomes (pick 1–2 most relevant): {', '.join(sorted(all_cos_ch))}"

            cancelled = False
            # If count_override is set (retry-failed), cap total generation
            effective_total_ch = request.count_override if request.count_override else total_questions

            for plan in generation_plan:
                if cancelled:
                    break
                if request.count_override and questions_generated >= request.count_override:
                    break
                q_type = plan["type"]
                count = plan["count"]
                marks = plan["marks"]
                chunk_idx = 0

                for i in range(count):
                    if request.count_override and questions_generated >= request.count_override:
                        break
                    # Check if client has disconnected (cancelled)
                    if await raw_request.is_disconnected():
                        logger.info(f"[{request_id}] Client disconnected, cancelling chapter generation")
                        cancelled = True
                        break

                    try:
                        sel_idx = (chunk_idx + random.randint(0, min(2, len(syllabus_chunks) - 1))) % len(syllabus_chunks)
                        selected_content = syllabus_chunks[sel_idx]
                        chunk_idx = (chunk_idx + max(1, len(syllabus_chunks) // count)) % len(syllabus_chunks)

                        system_prompt = _get_rubric_system_prompt(q_type)

                        # Consume next pre-assigned LO slot
                        required_lo_ch = chapter_lo_slots[lo_slot_idx_ch] if lo_slot_idx_ch < len(chapter_lo_slots) else None
                        lo_slot_idx_ch += 1

                        prompt = f"""Chapter: "{topic_name}"

Syllabus content:
{selected_content}
"""
                        q_ref_context = _ref_context_for_index(questions_generated)
                        if q_ref_context:
                            prompt += f"""
Background knowledge (use to inform the question, do NOT cite or reference it):
{q_ref_context}
"""
                        # LO requirement — either mandated by distribution or restricted by lo_filter
                        lo_requirement_ch = ""
                        if required_lo_ch:
                            lo_desc_ch = subject_lo_descriptions_ch.get(required_lo_ch, "")
                            lo_requirement_ch = (
                                f"\n- REQUIRED Learning Outcome: This question MUST specifically assess {required_lo_ch}"
                            )
                            if lo_desc_ch:
                                lo_requirement_ch += f' \u2014 "{lo_desc_ch}"'
                            lo_requirement_ch += (
                                "\n- Design the question to test the competency described by this LO, not just mention the topic"
                            )
                        elif request.lo_filter:
                            lo_requirement_ch = f"\n- REQUIRED: This question MUST map to one of the following Learning Outcomes: {', '.join(request.lo_filter)}. Choose the most relevant one."
                        elif topic_lo_mappings:
                            lo_list = ', '.join(topic_lo_mappings.keys())
                            lo_requirement_ch = f"\nAvailable Learning Outcomes for this chapter: {lo_list}"

                        # Pick a unique starter for this question
                        starter = _pick_starter(q_type, used_starters_chapter)

                        difficulty_label = (request.difficulty or "medium").capitalize()
                        difficulty_guidance = {
                            "easy": "straightforward recall or recognition; suitable for introductory assessment",
                            "medium": "application or understanding; requires connecting concepts",
                            "hard": "analysis, synthesis, or evaluation; requires deep understanding and critical thinking",
                        }.get((request.difficulty or "medium").lower(), "application or understanding")

                        prompt += f"""
Generate a {q_type.replace('_', ' ')} question based STRICTLY on the chapter "{topic_name}" content above.
- Marks: {marks}
- Difficulty level: {difficulty_label} — {difficulty_guidance}
- The question MUST be about the topic "{topic_name}" — do NOT use content from other chapters
- REQUIRED: Your question MUST start with the word/phrase "{starter}" — this is non-negotiable
- FORBIDDEN: Do NOT start all the questions with "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", "An engineer", or similar scenario setups. Also start directly with the required starter word.
- The question text must be fully self-contained — do NOT mention "the text", "the reference", "the provided material", "reference [N]", or any source
- Avoid questions similar to: {', '.join([q[:50] for q in generated_texts[-5:]]) if generated_texts else 'None yet'}{lo_requirement_ch}
{co_context_ch}
- For "course_outcome_mapping" in your JSON: assign 1 CO for most questions, at most 2 COs for genuinely cross-cutting questions. Do NOT map to 3 or more COs. Use Bloom level as value (1=remember/understand, 2=apply/analyze, 3=evaluate/create).
- Also include "bloom_level" (one of: remember, understand, apply, analyze, evaluate, create) matching the difficulty.

Output valid JSON only."""

                        accepted = False
                        for attempt in range(RETRY_LIMIT):
                            try:
                                resp = await llm_service.generate_json(
                                    prompt=prompt,
                                    system_prompt=system_prompt,
                                    temperature=0.75 + (0.05 * attempt),
                                )
                            except Exception:
                                resp = None

                            if not resp or "question_text" not in resp:
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            candidate_text = resp["question_text"]
                            if len(candidate_text) < 15:
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            try:
                                candidate_emb = await embedding_service.get_embedding(candidate_text)
                            except Exception:
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            # Validate MCQ has at least 4 options
                            if q_type == "mcq":
                                raw_opts = resp.get("options")
                                normalized_opts = _normalize_mcq_options(raw_opts, "mcq")
                                if not normalized_opts or len(normalized_opts) < 4:
                                    logger.debug(f"[{request_id}] Chapter MCQ has only {len(normalized_opts) if normalized_opts else 0} options (attempt {attempt+1}), retrying")
                                    if attempt == RETRY_LIMIT - 1:
                                        questions_failed += 1
                                    continue

                            # Dedupe: only check within the current session's generated questions.
                            # Skipping the full DB (existing_embeddings) here because chapter
                            # generation is expected to produce new questions even for topics
                            # that already have some questions in the bank. Comparing against
                            # all DB embeddings with a small LLM causes near-universal rejection.
                            if _is_duplicate_embedding(candidate_emb, [], generated_embeddings):
                                logger.debug(f"[{request_id}] Chapter candidate rejected as duplicate (attempt {attempt+1})")
                                if attempt == RETRY_LIMIT - 1:
                                    questions_failed += 1
                                continue

                            accepted = True

                            # LO / CO assignment
                            # 1) If a required LO was pre-assigned from distribution, honour it
                            if required_lo_ch:
                                llm_lo_ch = resp.get("learning_outcome_id") or resp.get("learning_outcome")
                                assigned_lo = llm_lo_ch if llm_lo_ch == required_lo_ch else required_lo_ch
                            else:
                                assigned_lo = None
                                if isinstance(resp.get("learning_outcome_id"), str):
                                    assigned_lo = resp["learning_outcome_id"]
                                elif isinstance(resp.get("learning_outcome"), str):
                                    assigned_lo = resp["learning_outcome"]
                                if not assigned_lo and topic_lo_mappings:
                                    try:
                                        assigned_lo = max(topic_lo_mappings.items(), key=lambda kv: (kv[1] or 0))[0]
                                    except Exception:
                                        pass

                            # CO — prefer multi-CO dict from LLM, then fallbacks
                            assigned_co_map = {}
                            if isinstance(resp.get("course_outcome_mapping"), dict) and resp["course_outcome_mapping"]:
                                assigned_co_map = resp["course_outcome_mapping"]
                            elif isinstance(resp.get("course_outcome"), dict):
                                assigned_co_map = resp["course_outcome"]
                            elif isinstance(resp.get("course_outcome"), str) and resp["course_outcome"]:
                                assigned_co_map = {resp["course_outcome"]: 1}
                            if not assigned_co_map and assigned_lo and subject_lo_mappings_data and isinstance(subject_lo_mappings_data.get(assigned_lo), dict):
                                assigned_co_map = subject_lo_mappings_data.get(assigned_lo)
                            if not assigned_co_map and assigned_lo and subject_cos and isinstance(subject_cos, list) and len(subject_cos) > 0:
                                m = re.search(r"LO(\d+)", str(assigned_lo))
                                if m:
                                    idx = (int(m.group(1)) - 1) % len(subject_cos)
                                    co_id = subject_cos[idx].get('id', f'CO{idx+1}') if isinstance(subject_cos[idx], dict) else f'CO{idx+1}'
                                    assigned_co_map = {co_id: 1}

                            # Sanitize fields to fit DB column limits
                            _s = _sanitize_question_fields({
                                "learning_outcome_id": assigned_lo,
                                "course_outcome_mapping": assigned_co_map,
                                "bloom_taxonomy_level": (resp.get("bloom_level") if resp else "understand"),
                                "difficulty_level": (request.difficulty or "medium"),
                            })
                            assigned_lo = _s["learning_outcome_id"]
                            assigned_co_map = _s["course_outcome_mapping"]

                            question = Question(
                                session_id=chapter_session_id,
                                document_id=chapter_question_document_id,
                                subject_id=str(subject_id_str),
                                topic_id=str(topic_id_str),
                                question_text=candidate_text,
                                question_embedding=candidate_emb,
                                question_type=q_type,
                                marks=marks,
                                difficulty_level=_s["difficulty_level"],
                                bloom_taxonomy_level=_s["bloom_taxonomy_level"],
                                correct_answer=(resp.get("correct_answer") or resp.get("expected_answer")),
                                options=_normalize_mcq_options(resp.get("options"), q_type),
                                explanation=resp.get("explanation"),
                                topic_tags=resp.get("topic_tags", [topic_name]),
                                learning_outcome_id=assigned_lo,
                                course_outcome_mapping=assigned_co_map,
                                generation_confidence=0.75,
                                vetting_status="pending",
                                generation_metadata={
                                    "source": "chapter_generation",
                                    "topic_name": topic_name,
                                    "has_reference_context": bool(ref_chunk_pool),
                                    "raw_response": resp,
                                    "source_info": build_source_info_for_index(questions_generated),
                                },
                            )
                            # Use an isolated session per question so a DB error on one
                            # question does not corrupt the main session's transaction state.
                            async with AsyncSessionLocal() as q_session:
                                q_session.add(question)
                                await q_session.commit()
                                await q_session.refresh(question)

                            generated_texts.append(candidate_text)
                            generated_embeddings.append(candidate_emb)
                            questions_generated += 1

                            progress = 10 + int((questions_generated / max(effective_total_ch, 1)) * 85)
                            yield f"data: {json.dumps({'status': 'generating', 'progress': min(progress, 95), 'current_question': questions_generated, 'total_questions': effective_total_ch, 'message': f'Generated {questions_generated}/{effective_total_ch} questions', 'question': {'id': str(question.id), 'question_text': question.question_text, 'question_type': question.question_type, 'marks': question.marks, 'difficulty_level': question.difficulty_level, 'bloom_taxonomy_level': question.bloom_taxonomy_level, 'options': question.options, 'correct_answer': question.correct_answer, 'explanation': question.explanation, 'topic_name': topic_name, 'topic_tags': question.topic_tags, 'learning_outcome_id': question.learning_outcome_id, 'course_outcome_mapping': question.course_outcome_mapping, 'vetting_status': question.vetting_status}})}\n\n"
                            break

                        if not accepted:
                            continue
                    except Exception as e:
                        logger.error(f"[{request_id}] Error generating question: {e}")
                        questions_failed += 1
                        continue

            # Finalize chapter session
            try:
                chapter_gen_session.questions_generated = questions_generated
                chapter_gen_session.questions_failed = max(0, total_questions - questions_generated)
                chapter_gen_session.status = "completed"
                chapter_gen_session.completed_at = datetime.now(timezone.utc)
                await db.commit()
            except Exception as e:
                logger.error(f"[{request_id}] Error finalizing chapter session: {e}")

            # Update subject stats
            try:
                from app.models.subject import Subject as SubjModel
                result = await db.execute(select(SubjModel).where(SubjModel.id == str(subject_id_str)))
                subj = result.scalar_one_or_none()
                if subj:
                    subj.total_questions = (subj.total_questions or 0) + questions_generated
                    await db.commit()
            except Exception as e:
                logger.error(f"[{request_id}] Error updating subject stats: {e}")

            questions_failed_count = max(0, total_questions - questions_generated)
            yield f"data: {json.dumps({'status': 'complete', 'progress': 100, 'current_question': questions_generated, 'total_questions': total_questions, 'questions_failed': questions_failed_count, 'message': f'Generated {questions_generated} questions' + (f' ({questions_failed_count} shortfall)' if questions_failed_count > 0 else '')})}\n\n"
            logger.info(f"[{request_id}] Chapter gen complete: {questions_generated} generated, {questions_failed_count} shortfall")

        except Exception as e:
            logger.error(f"[{request_id}] Chapter generation failed: {e}")
            yield f"data: {json.dumps({'status': 'error', 'progress': 0, 'message': f'Generation failed: {str(e)}'})}\n\n"
        finally:
            try:
                await db.rollback()
            except Exception:
                pass
            try:
                await db.close()
            except Exception:
                pass
    return StreamingResponse(
        sse_with_heartbeat(event_generator()),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _extract_highlighted_phrase(chunk_text: str, query_text: str, max_chars: int = 300) -> str:
    """Extract the most relevant sentence(s) from a chunk matching the query."""
    import re
    STOPWORDS = {
        'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
        'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
        'should', 'may', 'might', 'shall', 'can', 'of', 'in', 'on', 'at',
        'to', 'for', 'with', 'by', 'from', 'as', 'into', 'through', 'during',
        'what', 'which', 'who', 'how', 'when', 'where', 'why', 'if', 'and',
        'or', 'but', 'not', 'that', 'this', 'these', 'those', 'it', 'its',
    }
    sentences = re.split(r'(?<=[.!?])\s+', chunk_text.strip())
    sentences = [s.strip() for s in sentences if len(s.strip()) >= 15]
    if not sentences:
        return chunk_text[:max_chars]
    q_words = set(re.sub(r'[^\w\s]', '', query_text.lower()).split()) - STOPWORDS
    if not q_words:
        return sentences[0][:max_chars]
    scored = [(s, len(q_words & set(re.sub(r'[^\w\s]', '', s.lower()).split()))) for s in sentences]
    scored.sort(key=lambda x: x[1], reverse=True)
    result = scored[0][0]
    if len(scored) > 1 and scored[1][1] > 0:
        candidate = result + ' ' + scored[1][0]
        if len(candidate) <= max_chars:
            result = candidate
    if len(result) > max_chars:
        result = result[:max_chars].rsplit(' ', 1)[0] + '\u2026'
    return result


def _build_rubric_source_info(ref_chunks: list, topic_name: str) -> dict:
    """Build source_info dict from reference chunks for rubric-generated questions."""
    if not ref_chunks:
        return {"sources": [], "generation_reasoning": None, "content_coverage": None}
    
    sources = []
    for chunk in ref_chunks:
        chunk_meta = getattr(chunk, "chunk_metadata", None) or {}
        source_meta = chunk_meta.get("source_info", {})
        document_name = source_meta.get("document_name")
        if not document_name and hasattr(chunk, 'document') and chunk.document:
            document_name = chunk.document.filename
        if not document_name:
            document_name = "Unknown Document"
        
        source = {
            "document_name": document_name,
            "page_number": getattr(chunk, "page_number", None) or source_meta.get("page_number"),
            "page_range": source_meta.get("page_range"),
            "position_in_page": source_meta.get("position_in_page"),
            "position_percentage": source_meta.get("position_percentage"),
            "section_heading": getattr(chunk, "section_heading", None),
            "content_snippet": chunk.chunk_text,
            "highlighted_phrase": _extract_highlighted_phrase(chunk.chunk_text, topic_name),
            "relevance_reason": f"Retrieved as reference material for {topic_name}",
        }
        sources.append(source)
    
    return {
        "sources": sources,
        "generation_reasoning": f"Question generated from '{topic_name}' syllabus content with reference material support",
        "content_coverage": f"Rubric-based question covering {topic_name} concepts",
    }


def _normalize_mcq_options(options, question_type: str = "mcq"):
    """
    Normalize and validate MCQ options.
    Ensures options are a list of properly formatted strings like 'A) text'.
    Returns None for non-MCQ types.
    """
    if question_type != "mcq" or not options:
        return options
    
    labels = ['A', 'B', 'C', 'D', 'E', 'F']
    normalized = []
    
    if isinstance(options, list):
        for i, opt in enumerate(options):
            label = labels[i] if i < len(labels) else str(i + 1)
            if isinstance(opt, str):
                opt = opt.strip()
                if not opt:
                    continue
                # Check if it already has a label prefix (e.g. "A) ..." or "A. ...")
                if len(opt) > 1 and opt[0].isalpha() and opt[1] in ').:':
                    normalized.append(opt)
                else:
                    normalized.append(f"{label}) {opt}")
            elif isinstance(opt, dict):
                text = opt.get("text") or opt.get("description") or opt.get("option") or opt.get("value") or str(opt)
                normalized.append(f"{label}) {text}")
            else:
                normalized.append(f"{label}) {str(opt)}")
    elif isinstance(options, dict):
        # Handle dict format like {"A": "text", "B": "text", ...}
        for key, val in options.items():
            label = str(key).strip().rstrip(').:')
            text = str(val).strip() if val else ""
            if text:
                normalized.append(f"{label}) {text}")
    
    return normalized if normalized else options


def _compute_lo_question_slots(lo_distribution: dict, total_questions: int) -> list:
    """
    Pre-assign an LO to each question slot based on the rubric LO distribution.

    The lo_distribution values are percentage weights (may sum > 100% when questions
    address multiple LOs). We normalize them and use the largest-remainder method
    so every slot has a mandated LO that the LLM MUST target.

    Returns a shuffled list of LO IDs of length == total_questions.
    """
    import random
    if not lo_distribution or total_questions == 0:
        return [None] * total_questions
    lo_items = [(k, float(v or 0)) for k, v in lo_distribution.items() if (v or 0) > 0]
    if not lo_items:
        return [None] * total_questions
    total_weight = sum(w for _, w in lo_items)
    if total_weight == 0:
        return [None] * total_questions
    # Largest remainder method for fair rounding
    quotients = [(k, w / total_weight * total_questions) for k, w in lo_items]
    floor_counts = {k: int(q) for k, q in quotients}
    remainders = sorted(quotients, key=lambda x: -(x[1] - int(x[1])))
    extra = total_questions - sum(floor_counts.values())
    for k, _ in remainders[:extra]:
        floor_counts[k] = floor_counts.get(k, 0) + 1
    # Build interleaved list
    slots: list = []
    for lo_id, count in floor_counts.items():
        slots.extend([lo_id] * count)
    random.shuffle(slots)
    return slots


def _select_lo_aware_topic(
    required_lo: "str | None",
    topic_ids_ordered: list,
    chunks_by_topic: dict,
    topic_question_counts: dict,
) -> str:
    """
    Select the best topic ID for a given required LO.

    Priority:
    1. Topics that have a mapping for required_lo → pick the least-used one.
    2. Fall back to the globally least-used topic.
    """
    if required_lo and topic_ids_ordered:
        lo_mapped = [
            tid for tid in topic_ids_ordered
            if required_lo in (chunks_by_topic.get(tid, [{}])[0].get("lo_mappings") or {})
        ]
        if lo_mapped:
            return min(lo_mapped, key=lambda t: topic_question_counts.get(t, 0))
    return min(topic_ids_ordered, key=lambda t: topic_question_counts.get(t, 0))


def _get_rubric_system_prompt(question_type: str) -> str:
    """Get system prompt for rubric-based question generation."""
    if question_type == "mcq":
        return """You are an expert educator creating CHALLENGING examination questions.
Generate a high-quality multiple-choice question (MCQ) that tests APPLICATION or ANALYSIS, not just recall.

Critical Guidelines:
- COGNITIVE DEPTH: Focus on applying concepts, analyzing scenarios, or evaluating options - avoid simple recall
- Use question patterns like:
  * "Which approach would be MOST effective when..."
  * "What would be the PRIMARY consequence of..."
  * "Which statement BEST explains the relationship between..."
  * "What distinguishes X from Y in terms of..."
- Start with interrogative words (What, Which, How, When, Why) or imperative (Calculate, Find, Determine, Identify)
- All 4 options must be plausible and challenging - avoid obviously wrong distractors
- FORBIDDEN: Do NOT start with scenario setups like "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", "An engineer", etc. Start directly with interrogatives.
- CRITICAL: The question text must be fully self-contained. NEVER mention, cite, or reference the source material, context, references, or any provided text. Do NOT use phrases like "According to the text", "As discussed in the reference", "Based on the provided material", "As mentioned in reference", "According to reference [N]", "As stated in the context", or any similar phrasing. Write it as a standalone exam question.

Output format (JSON):
{
    "question_text": "The question text ending with ?",
    "options": ["A) Option 1", "B) Option 2", "C) Option 3", "D) Option 4"],
    "correct_answer": "A",
    "explanation": "Brief explanation of why A is correct",
    "topic_tags": ["topic1"],
    "bloom_level": "apply",
    "learning_outcome": "LO1",
    "course_outcome_mapping": {"CO1": 2, "CO2": 1}
}
NOTE: course_outcome_mapping must be a JSON object mapping CO IDs to Bloom level (1=remember/understand, 2=apply/analyze, 3=evaluate/create). Assign 1 CO for most questions; use 2 COs only when the question is genuinely cross-cutting. NEVER assign 3 or more COs."""
    elif question_type == "short_answer":
        return """You are an expert educator creating THOUGHT-PROVOKING examination questions.
Generate a short-answer question requiring ANALYSIS or EVALUATION, not just description.

Critical Guidelines:
- COGNITIVE DEPTH: Questions should require reasoning, comparison, or justification
- Use higher-order starters:
  * "Explain WHY... rather than just..."
  * "Compare and contrast..."
  * "Analyze the relationship between..."
  * "Justify the use of... in..."
- Expected answers should demonstrate reasoning and understanding, not just list facts
- FORBIDDEN: Do NOT start with scenario setups like "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", "An engineer", etc. Start with action verbs.
- CRITICAL: Both the question AND expected_answer must be fully self-contained. NEVER mention, cite, or reference the source material, context, references, or any provided text. Do NOT use phrases like "According to the text", "As discussed in the reference", "Based on the provided material", "As mentioned in reference", "According to reference [N]", "As stated in the context", or any similar phrasing. Write as standalone exam content.

Output format (JSON):
{
    "question_text": "The question text",
    "expected_answer": "Model answer (3-5 sentences) - demonstrating reasoning and analysis",
    "key_points": ["point1", "point2"],
    "topic_tags": ["topic1"],
    "bloom_level": "analyze",
    "learning_outcome": "LO1",
    "course_outcome_mapping": {"CO1": 2, "CO2": 1}
}
NOTE: course_outcome_mapping must be a JSON object mapping CO IDs to Bloom level (1=remember/understand, 2=apply/analyze, 3=evaluate/create). Assign 1 CO for most questions; use 2 COs only when the question is genuinely cross-cutting. NEVER assign 3 or more COs."""
    else:
        return """You are an expert educator creating ANALYTICALLY RIGOROUS examination questions.
Generate a long-answer question requiring CRITICAL ANALYSIS or EVALUATION.

Critical Guidelines:
- HIGHEST COGNITIVE LEVEL: Questions must require synthesis, evaluation, or critical thinking
- Use analytical prompts:
  * "Critically evaluate the advantages and limitations of..."
  * "Analyze how... differs from... in terms of..."
  * "Discuss the implications of... on..."
  * "Assess the effectiveness of..."
- Expected answers should include structured argumentation and justification
- FORBIDDEN: Do NOT start with scenario setups like "A clinician", "A patient", "A doctor", "A dentist", "A practitioner", "A student", "A researcher", "An engineer", etc. Start with analytical verbs.
- CRITICAL: Both the question AND expected_answer must be fully self-contained. NEVER mention, cite, or reference the source material, context, references, or any provided text. Do NOT use phrases like "According to the text", "As discussed in the reference", "Based on the provided material", "As mentioned in reference", "According to reference [N]", "As stated in the context", or any similar phrasing. Write as standalone professional exam content.

Output format (JSON):
{
    "question_text": "The question text",
    "expected_answer": "Model answer outline covering key arguments and analysis",
    "topic_tags": ["topic1"],
    "bloom_level": "evaluate",
    "learning_outcome": "LO1",
    "course_outcome_mapping": {"CO1": 3, "CO2": 2}
}
NOTE: course_outcome_mapping must be a JSON object mapping CO IDs to Bloom level (1=remember/understand, 2=apply/analyze, 3=evaluate/create). Assign 1 CO for most questions; use 2 COs only when the question is genuinely cross-cutting. NEVER assign 3 or more COs."""




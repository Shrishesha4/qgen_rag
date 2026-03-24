"""
Document management API endpoints.
"""

import asyncio
import os
from typing import Optional, Literal

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Form
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.config import settings
from app.schemas.document import (
    DocumentResponse,
    DocumentListResponse,
    DocumentListItem,
    DocumentUploadResponse,
    PaginationInfo,
)
from app.services.document_service import DocumentService
from app.api.v1.deps import get_current_user, rate_limit
from app.models.user import User


router = APIRouter()


MIME_TYPE_MAPPING = {
    ".pdf": "application/pdf",
    ".txt": "text/plain",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


@router.post("/upload", response_model=DocumentUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_document(
    file: UploadFile = File(...),
    current_user: User = Depends(rate_limit(requests=10, window_seconds=3600)),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a document for processing.
    Supported formats: PDF, DOCX, TXT
    """
    # Validate file extension
    filename = file.filename or "document"
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Unsupported file type. Allowed: {', '.join(settings.ALLOWED_EXTENSIONS)}",
        )
    
    # Read file content
    content = await file.read()
    
    # Validate file size
    if len(content) > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Maximum size: {settings.MAX_UPLOAD_SIZE_MB}MB",
        )
    
    # Determine mime type
    mime_type = MIME_TYPE_MAPPING.get(ext, "application/octet-stream")
    
    # Upload document
    document_service = DocumentService(db)
    
    try:
        document = await document_service.upload_document(
            user_id=current_user.id,
            filename=filename,
            file_content=content,
            mime_type=mime_type,
        )
        
        return DocumentUploadResponse(
            document_id=document.id,
            filename=document.filename,
            status=document.processing_status,
            message="Document uploaded successfully. Processing started.",
        )
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=str(e),
        )


@router.get("", response_model=DocumentListResponse)
async def list_documents(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None, description="Filter by processing status"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    List all documents for the current user.
    """
    document_service = DocumentService(db)
    
    documents, pagination = await document_service.get_documents(
        user_id=current_user.id,
        page=page,
        limit=limit,
        status=status,
    )
    
    return DocumentListResponse(
        documents=[
            DocumentListItem(
                id=doc.id,
                filename=doc.filename,
                processing_status=doc.processing_status,
                upload_timestamp=doc.upload_timestamp,
                total_chunks=doc.total_chunks,
                questions_generated=getattr(doc, "questions_generated", 0),
            )
            for doc in documents
        ],
        pagination=pagination,
    )


@router.get("/{document_id}", response_model=DocumentResponse)
async def get_document(
    document_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get document details by ID.
    """
    document_service = DocumentService(db)
    
    document = await document_service.get_document(document_id, current_user.id)
    
    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found",
        )
    
    return DocumentResponse.model_validate(document)


@router.delete("/{document_id}")
async def delete_document(
    document_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Delete a document while keeping generated questions.
    """
    document_service = DocumentService(db)
    
    success = await document_service.delete_document(document_id, current_user.id)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found",
        )
    
    return {"message": "Document deleted successfully"}


@router.get("/{document_id}/chunks")
async def get_document_chunks(
    document_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get all chunks for a document (for debugging/preview).
    """
    document_service = DocumentService(db)
    
    try:
        chunks = await document_service.get_document_chunks(document_id, current_user.id)
        
        return {
            "document_id": str(document_id),
            "total_chunks": len(chunks),
            "chunks": [
                {
                    "index": chunk.chunk_index,
                    "text": chunk.chunk_text[:500] + "..." if len(chunk.chunk_text) > 500 else chunk.chunk_text,
                    "token_count": chunk.token_count,
                    "page_number": chunk.page_number,
                }
                for chunk in chunks
            ],
        }
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e),
        )


@router.get("/{document_id}/status")
async def get_document_status(
    document_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get document processing status.
    """
    document_service = DocumentService(db)
    
    document = await document_service.get_document(document_id, current_user.id)
    
    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found",
        )
    
    metadata = document.document_metadata or {}
    return {
        "document_id": str(document.id),
        "filename": document.filename,
        "status": document.processing_status,
        "total_chunks": document.total_chunks,
        "total_tokens": document.total_tokens,
        "processed_at": document.processed_at,
        "processing_step": metadata.get("processing_step"),
        "processing_progress": metadata.get("processing_progress", 0),
        "processing_detail": metadata.get("processing_detail"),
        "total_pages": metadata.get("total_pages"),
        "extraction_method": metadata.get("extraction_method"),
        "used_ocr": metadata.get("used_ocr", False),
        "error": metadata.get("error"),
    }


# ============== Reference Document Management ==============


def _parse_reference_questions_from_spreadsheet(ext: str, content: bytes) -> list:
    """Parse reference questions from Excel/CSV file content."""
    import io
    
    parsed_questions = []
    
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        row1_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        row2_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[2]] if ws.max_row >= 2 else []
        
        if any("question" in v for v in row2_values):
            headers = row2_values
            data_start = 3
        elif any("question" in v for v in row1_values):
            headers = row1_values
            data_start = 2
        else:
            headers = row1_values
            data_start = 2
        
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if row and any(v is not None for v in row):
                cells = [str(v).strip() if v is not None else "" for v in row]
                parsed_questions.append(dict(zip(headers, cells)))
    
    elif ext == ".csv":
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if len(all_rows) >= 2:
            headers = [v.strip().lower() for v in all_rows[0]]
            for row in all_rows[1:]:
                if any(v.strip() for v in row):
                    cells = [v.strip() for v in row]
                    parsed_questions.append(dict(zip(headers, cells)))
    
    # Build structured reference question list
    ref_questions = []
    for pq in parsed_questions:
        q_text = ""
        for key in pq:
            if "question" in key and pq[key]:
                q_text = pq[key]
                break
        if not q_text:
            continue
        
        # Extract options
        options = []
        for opt_key in ["option_a", "option a", "a", "option_b", "option b", "b", 
                         "option_c", "option c", "c", "option_d", "option d", "d"]:
            if opt_key in pq and pq[opt_key]:
                options.append(pq[opt_key])
        
        # Extract correct answer
        correct = ""
        for ans_key in ["option_correct", "correct", "answer", "correct_answer", "correct answer"]:
            if ans_key in pq and pq[ans_key]:
                correct = pq[ans_key]
                break
        
        # Extract other fields
        difficulty = ""
        for d_key in ["difficulty", "difficulty_level"]:
            if d_key in pq and pq[d_key]:
                difficulty = pq[d_key]
                break
        
        marks_val = ""
        for m_key in ["marks", "mark", "points"]:
            if m_key in pq and pq[m_key]:
                marks_val = pq[m_key]
                break
        
        co_val = ""
        for c_key in ["co", "course_outcome", "course outcome"]:
            if c_key in pq and pq[c_key]:
                co_val = pq[c_key]
                break
        
        lo_val = ""
        for l_key in ["lo", "lo mapping", "lo_mapping", "learning_outcome", "learning outcome"]:
            if l_key in pq and pq[l_key]:
                lo_val = pq[l_key]
                break
        
        ref_questions.append({
            "question_text": q_text,
            "options": options if options else None,
            "correct_answer": correct or None,
            "question_type": "mcq" if len(options) >= 3 else "short_answer",
            "difficulty": difficulty or None,
            "marks": marks_val or None,
            "co": co_val or None,
            "lo": lo_val or None,
        })
    
    return ref_questions


async def _process_pdf_reference_questions(
    document_service: DocumentService,
    content: bytes,
    filename: str,
    document_id: str,
) -> None:
    """
    Background task: Extract text from a PDF (with OCR fallback for scanned docs),
    then use the LLM to parse questions into structured format.
    Uses short-lived DB sessions to avoid holding connections during long operations.
    """
    import logging
    from sqlalchemy import update
    from app.models.document import Document
    from app.services.llm_service import LLMService
    from app.core.database import AsyncSessionLocal
    import fitz
    import re
    
    logger = logging.getLogger(__name__)
    
    async def _update_doc(values: dict):
        """Helper to update document with a short-lived session."""
        async with AsyncSessionLocal() as session:
            await session.execute(
                update(Document).where(Document.id == document_id).values(**values)
            )
            await session.commit()
    
    try:
        await _update_doc({
            "processing_status": "processing",
            "document_metadata": {
                "processing_step": "extracting_text",
                "processing_progress": 10,
                "processing_detail": "Fast parsing mode",
            },
        })

        max_pages = 40
        pages_data = []
        with fitz.open(stream=content, filetype="pdf") as doc:
            total_pages = min(len(doc), max_pages)
            for i in range(total_pages):
                txt = doc[i].get_text("text").strip()
                if txt:
                    pages_data.append({"text": txt, "extraction_method": "native"})

        full_text = "\n\n".join(p["text"] for p in pages_data if p.get("text"))
        used_ocr = False

        # If native extraction yields almost nothing, fall back to existing OCR-capable extractor.
        if len(full_text) < 500:
            await _update_doc({
                "document_metadata": {
                    "processing_step": "extracting_text",
                    "processing_progress": 35,
                    "processing_detail": "Low native text, trying OCR fallback...",
                }
            })
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            try:
                async with AsyncSessionLocal() as tmp_session:
                    bg_document_service = DocumentService(tmp_session)
                    pages_data = await bg_document_service._extract_text_with_pages(
                        tmp_path, "application/pdf", progress_callback=None
                    )
                full_text = "\n\n".join(p.get("text", "") for p in pages_data if p.get("text", "").strip())
                used_ocr = any(p.get("extraction_method") == "ocr" for p in pages_data)
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        if not full_text.strip():
            await _update_doc({
                "processing_status": "completed",
                "total_chunks": 0,
                "document_metadata": {
                    "parsed_questions": [],
                    "total_parsed": 0,
                    "error": "No text could be extracted from the PDF.",
                    "used_ocr": used_ocr,
                    "total_pages": len(pages_data),
                },
            })
            return

        await _update_doc({
            "document_metadata": {
                "processing_step": "heuristic_parsing",
                "processing_progress": 65,
                "processing_detail": "Extracting questions with fast parser...",
                "used_ocr": used_ocr,
                "total_pages": len(pages_data),
            }
        })

        question_pattern = re.compile(r"(?:^|\n)\s*(?:Q\.?\s*\d+|\d+[\.)])\s+(.+?)(?=(?:\n\s*(?:Q\.?\s*\d+|\d+[\.)])\s+)|\Z)", re.IGNORECASE | re.DOTALL)
        heuristic_questions = []
        for match in question_pattern.finditer(full_text):
            q_text = re.sub(r"\s+", " ", match.group(1)).strip()
            if len(q_text) < 12:
                continue
            heuristic_questions.append({
                "question_text": q_text[:2000],
                "options": None,
                "correct_answer": None,
                "question_type": "short_answer",
                "difficulty": None,
                "marks": None,
                "co": None,
                "lo": None,
            })

        all_questions = heuristic_questions

        # Optional single LLM pass if heuristic parser found too few questions.
        if len(all_questions) < 3:
            await _update_doc({
                "document_metadata": {
                    "processing_step": "llm_parsing",
                    "processing_progress": 80,
                    "processing_detail": "Refining extraction with AI...",
                    "used_ocr": used_ocr,
                    "total_pages": len(pages_data),
                }
            })
            llm = LLMService()
            prompt_text = full_text[:14000]
            result = await llm.generate_json(
                prompt=(
                    "Extract exam questions from this text and return ONLY JSON with key 'questions'.\n\n"
                    f"TEXT:\n{prompt_text}"
                ),
                system_prompt=(
                    "Return JSON: {\"questions\":[{\"question_text\":str,\"options\":list|null,"
                    "\"correct_answer\":str|null,\"question_type\":str,\"difficulty\":str|null,"
                    "\"marks\":str|null,\"co\":str|null,\"lo\":str|null}]}"
                ),
                temperature=0.0,
            )
            llm_questions = result.get("questions", []) if isinstance(result, dict) else []
            if llm_questions:
                all_questions = []
                for q in llm_questions:
                    qt = (q.get("question_text") or "").strip()
                    if not qt:
                        continue
                    all_questions.append({
                        "question_text": qt[:2000],
                        "options": q.get("options"),
                        "correct_answer": q.get("correct_answer"),
                        "question_type": q.get("question_type", "short_answer"),
                        "difficulty": q.get("difficulty"),
                        "marks": q.get("marks"),
                        "co": q.get("co"),
                        "lo": q.get("lo"),
                    })

        await _update_doc({
            "processing_status": "completed",
            "total_chunks": len(all_questions),
            "document_metadata": {
                "parsed_questions": all_questions,
                "total_parsed": len(all_questions),
                "used_ocr": used_ocr,
                "total_pages": len(pages_data),
            },
        })
        logger.info(f"PDF reference questions parsed quickly: {len(all_questions)} questions from {filename}")

    except Exception as e:
        logger.error(f"Background PDF reference question processing failed: {e}")
        try:
            await _update_doc({
                "processing_status": "failed",
                "document_metadata": {"error": str(e)},
            })
        except Exception:
            pass

@router.post("/reference/upload", response_model=DocumentUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_reference_document(
    file: UploadFile = File(..., description="Reference book PDF, template paper PDF, or reference questions PDF/Excel/CSV"),
    subject_id: str = Form(..., description="Subject ID to associate the document with"),
    index_type: Literal["primary", "reference_book", "template_paper", "reference_questions"] = Form(..., description="Type of reference material"),
    topic_id: Optional[str] = Form(None, description="Topic ID to associate the document with (for per-topic references)"),
    current_user: User = Depends(rate_limit(requests=20, window_seconds=3600)),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a reference document (book, template paper, or reference questions).
    
    - reference_book: PDF used for conceptual inspiration when generating questions
    - template_paper: PDF used for detecting similarity with existing exam questions
    - reference_questions: PDF/Excel/CSV of past exam questions used as quality/style templates
    """
    import io
    
    filename = file.filename or "document.pdf"
    ext = os.path.splitext(filename)[1].lower()
    
    # Validate file type based on index_type
    if index_type == "reference_questions":
        if ext not in (".xlsx", ".csv", ".pdf"):
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=f"Reference questions must be .xlsx, .csv, or .pdf files, got: {ext}",
            )
    else:
        if ext not in settings.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=f"Unsupported file type. Allowed: {', '.join(settings.ALLOWED_EXTENSIONS)}",
            )
    
    content = await file.read()
    
    if len(content) > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Maximum size: {settings.MAX_UPLOAD_SIZE_MB}MB",
        )
    
    mime_type = MIME_TYPE_MAPPING.get(ext, "application/octet-stream")
    
    document_service = DocumentService(db)
    
    try:
        document = await document_service.upload_reference_document(
            user_id=current_user.id,
            filename=filename,
            file_content=content,
            mime_type=mime_type,
            subject_id=subject_id,
            index_type=index_type,
            topic_id=topic_id,
        )
        linked_from = (document.document_metadata or {}).get("linked_from_document_id")
        if linked_from:
            return DocumentUploadResponse(
                document_id=document.id,
                filename=document.filename,
                status=document.processing_status,
                message=f"{index_type.replace('_', ' ').title()} linked from existing vector index. Ready to use.",
            )
        
        # For reference_questions, parse Excel/CSV/PDF and store as structured chunks
        if index_type == "reference_questions":
            from sqlalchemy import update
            from app.models.document import Document
            
            if ext == ".pdf":
                # PDF: Process in background (OCR + LLM takes time)
                # Return immediately and let frontend poll for status
                asyncio.create_task(
                    _process_pdf_reference_questions(
                        document_service, content, filename, document.id
                    )
                )
                return DocumentUploadResponse(
                    document_id=document.id,
                    filename=document.filename,
                    status="processing",
                    message=f"PDF uploaded. Extracting and parsing questions from {filename}...",
                )
            else:
                # Excel/CSV: Direct spreadsheet parsing (fast, do inline)
                try:
                    ref_questions = _parse_reference_questions_from_spreadsheet(
                        ext, content
                    )
                    
                    await db.execute(
                        update(Document)
                        .where(Document.id == document.id)
                        .values(
                            processing_status="completed",
                            total_chunks=len(ref_questions),
                            document_metadata={
                                "parsed_questions": ref_questions,
                                "total_parsed": len(ref_questions),
                            },
                        )
                    )
                    await db.commit()
                    
                    return DocumentUploadResponse(
                        document_id=document.id,
                        filename=document.filename,
                        status="completed",
                        message=f"Reference questions uploaded: {len(ref_questions)} questions parsed from {filename}.",
                    )
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error(f"Failed to parse reference questions: {e}")
                    await db.execute(
                        update(Document)
                        .where(Document.id == document.id)
                        .values(processing_status="failed")
                    )
                    await db.commit()
        else:
            # reference_book / template_paper: run standard extract → embed → store pipeline
            asyncio.create_task(document_service._process_document(document.id))
        
        return DocumentUploadResponse(
            document_id=document.id,
            filename=document.filename,
            status=document.processing_status,
            message=f"{index_type.replace('_', ' ').title()} uploaded successfully. Processing started.",
        )
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=str(e),
        )


@router.get("/reference/list")
async def list_reference_documents(
    subject_id: Optional[str] = Query(None, description="Filter by subject"),
    index_type: Optional[str] = Query(None, description="Filter by type: reference_book, template_paper, reference_questions"),
    topic_id: Optional[str] = Query(None, description="Filter by topic"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    List reference documents (books, template papers, and reference questions).
    """
    document_service = DocumentService(db)
    
    documents = await document_service.get_reference_documents(
        user_id=current_user.id,
        subject_id=subject_id,
        index_type=index_type,
        topic_id=topic_id,
    )
    
    doc_responses = [
        {
            "id": str(doc.id),
            "filename": doc.filename,
            "file_size_bytes": doc.file_size_bytes,
            "mime_type": doc.mime_type,
            "index_type": doc.index_type,
            "subject_id": str(doc.subject_id) if doc.subject_id else None,
            "topic_id": str(doc.topic_id) if doc.topic_id else None,
            "processing_status": doc.processing_status,
            "total_chunks": doc.total_chunks,
            "upload_timestamp": doc.upload_timestamp.isoformat() if doc.upload_timestamp else None,
            "processed_at": doc.processed_at.isoformat() if doc.processed_at else None,
            "parsed_question_count": (doc.document_metadata or {}).get("total_parsed", 0) if doc.index_type == "reference_questions" else None,
        }
        for doc in documents
    ]
    
    reference_books = [d for d in doc_responses if d["index_type"] == "reference_book"]
    template_papers = [d for d in doc_responses if d["index_type"] == "template_paper"]
    reference_questions = [d for d in doc_responses if d["index_type"] == "reference_questions"]
    
    return {
        "reference_books": reference_books,
        "template_papers": template_papers,
        "reference_questions": reference_questions,
    }

